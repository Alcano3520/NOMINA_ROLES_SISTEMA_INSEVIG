import io

import openpyxl

from core.excel.nomina_builders import consolidado_xlsx
from core.repos import nomina


def _movs():
    # dos empleados, período junio
    yield {"EMPLEADO": "1012", "CLASE": 100, "VALOR": 800.0, "ASENTADO": True, "DIAS": 30}
    yield {"EMPLEADO": "1012", "CLASE": 200, "VALOR": 75.6, "ASENTADO": True, "DIAS": None}
    yield {"EMPLEADO": "1012", "CLASE": 205, "VALOR": 50.0, "ASENTADO": True, "DIAS": None}
    yield {"EMPLEADO": "2050", "CLASE": 100, "VALOR": 1200.0, "ASENTADO": True, "DIAS": 30}
    yield {"EMPLEADO": "2050", "CLASE": 105, "VALOR": 999.0, "ASENTADO": True, "DIAS": None}  # ignorada


_CATALOGOS = {"FNC": {"01": "GUARDIA"}, "DPT": {"10": "OPERACIONES"}, "SEC": {}}
_EMPLEADOS = {
    "1012": {"APELLIDOS": "PEREIRA", "NOMBRES": "JUAN", "CEDULA": 920116811.0, "CARGO": "01", "DEPTO": "10", "SECCION": "", "SUELDO": 800},
    "2050": {"APELLIDOS": "LOPEZ", "NOMBRES": "ANA", "CEDULA": "1712345678", "CARGO": "01", "DEPTO": "10", "SECCION": "", "SUELDO": 1200},
}


def test_consolidar_una_fila_por_empleado():
    filas = nomina.consolidar(_movs(), _CATALOGOS, _EMPLEADOS)
    assert [f.empleado for f in filas] == ["1012", "2050"]
    f1 = filas[0]
    assert f1.apellidos_nombres == "PEREIRA JUAN"
    assert f1.cedula == "0920116811"
    assert f1.cargo == "GUARDIA"
    assert f1.total_ingresos == 800.0
    assert f1.total_egresos == 125.6  # APORT_IESS 75.6 + PRESTAMOS_COMPANIA 50
    assert f1.total_recibir == 674.4


def test_to_row_tiene_totales_al_final():
    fila = nomina.consolidar(_movs(), _CATALOGOS, _EMPLEADOS)[0].to_row()
    assert fila["TOTAL_RECIBIR"] == 674.4
    assert fila["SUELDO"] == 800.0
    assert "TOTAL_INGRESOS" in fila


def test_leer_movimientos_fuente_invalida():
    import pytest

    with pytest.raises(ValueError, match="Fuente desconocida"):
        list(nomina.leer_movimientos("2026-06", historico=False, fuente="x", _lectores={}))


def test_comparar_detecta_diferencias():
    sql = [{"TOTAL_INGRESOS": 100.0, "TOTAL_EGRESOS": 10.0, "TOTAL_RECIBIR": 90.0, "A": 1}]
    sup = [
        {"TOTAL_INGRESOS": 100.0, "TOTAL_EGRESOS": 10.0, "TOTAL_RECIBIR": 90.0, "A": 1},
        {"TOTAL_INGRESOS": 5.0, "TOTAL_EGRESOS": 0.0, "TOTAL_RECIBIR": 5.0},
    ]
    difs = nomina.comparar(sql, sup)
    tipos = {d.tipo for d in difs}
    assert "conteo_filas" in tipos
    assert "suma" in tipos  # TOTAL_INGRESOS 100 vs 105


def test_comparar_sin_diferencias():
    rows = [{"TOTAL_INGRESOS": 1.0, "TOTAL_EGRESOS": 0.0, "TOTAL_RECIBIR": 1.0}]
    assert nomina.comparar(rows, list(rows)) == []


def test_consolidado_xlsx_valido():
    filas = [f.to_row() for f in nomina.consolidar(_movs(), _CATALOGOS, _EMPLEADOS)]
    data = consolidado_xlsx(filas)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Por departamento" in wb.sheetnames
    ws = wb["Consolidado"]
    headers = [c.value for c in ws[1]]
    assert headers[:3] == ["EMPLEADO", "APELLIDOS_NOMBRES", "CEDULA"]
    assert headers[-3:] == ["TOTAL_INGRESOS", "TOTAL_EGRESOS", "TOTAL_RECIBIR"]
    assert ws.max_row == 3  # header + 2 empleados
    # subtotales por depto
    dep = wb["Por departamento"]
    assert [c.value for c in dep[1]][:2] == ["DEPARTAMENTO", "EMPLEADOS"]
