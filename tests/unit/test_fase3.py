"""Fase 3: helpers de empleados + parsers de Excel (sin BD)."""

import io

import openpyxl

from core.excel.parsers import (
    _col_letra_a_idx,
    biess_autodetectar,
    parse_biess_manual,
    parse_biess_quirografarios,
    parse_carga_masiva_empleados,
)
from core.repos import empleados


def _xlsx(filas: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for f in filas:
        ws.append(f)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_token_cambia_si_cambia_un_campo():
    base = {c: "" for c in empleados.CAMPOS_EDITABLES}
    t1 = empleados._token(base)
    base["SUELDO"] = "800"
    assert empleados._token(base) != t1
    assert empleados._token({c: "" for c in empleados.CAMPOS_EDITABLES}) == t1


def test_normalizar_numericos_y_vacios():
    n = empleados._normalizar({"SUELDO": "800", "NOMBRES": "  JUAN ", "TELEFONO": "", "XXX": 1})
    assert n["SUELDO"] == 800.0
    assert n["NOMBRES"] == "JUAN"
    assert n["TELEFONO"] is None
    assert "XXX" not in n  # campo no editable


def test_normalizar_codificacion_de_flags():
    n = empleados._normalizar({
        "INCL_ROL": "1", "INCL_BAN": "",           # S/N texto
        "CAT_PROYECT_7": "S", "RPCAM2": "",        # '1'/'0' texto
        "PRIMARIA": "1", "SER_MIL": "",            # 1/0 entero
        "SEXO": "1",
    })
    assert n["INCL_ROL"] == "S" and n["INCL_BAN"] == "N"
    assert n["CAT_PROYECT_7"] == "1" and n["RPCAM2"] == "0"
    assert n["PRIMARIA"] == 1 and n["SER_MIL"] == 0
    assert n["SEXO"] == "1"


def test_estados_filtro():
    assert empleados._ESTADOS_FILTRO["ACTIVOS"] == ("ACT",)
    assert set(empleados._ESTADOS_FILTRO["INACTIVOS"]) == {"LIQ", "SUS"}
    assert empleados._ESTADOS_FILTRO["TODOS"] == ()


def test_ficha_empleado_pdf_parseable():
    import io

    import pypdf

    from core.pdf.ficha_empleado import ficha_empleado_pdf

    data = ficha_empleado_pdf("1012", {
        "NOMBRES": "JUAN", "APELLIDOS": "PEREZ", "CEDULA": "0920116811",
        "SUELDO": "800", "DIRECCION": "GUAYAQUIL", "CONYUGUE": "MARIA",
    })
    assert data[:4] == b"%PDF"
    txt = pypdf.PdfReader(io.BytesIO(data)).pages[0].extract_text()
    assert "PEREZ JUAN" in txt
    assert "Ficha de empleado" in txt


def test_documentos_empleado_pdf():
    import io

    import pypdf

    from core.pdf.documentos_empleado import DOCUMENTOS

    c = {
        "NOMBRES": "JUAN", "APELLIDOS": "PEREZ MERA", "CEDULA": "0920116811",
        "CARGO": "GUARDIA", "SUELDO": "470", "FECHA_ING": "2023-01-15", "ESTADO": "ACT",
    }
    for tipo, (_nombre, fn) in DOCUMENTOS.items():
        data = fn("1012", c)
        assert data[:4] == b"%PDF", tipo
        txt = pypdf.PdfReader(io.BytesIO(data)).pages[0].extract_text()
        assert "PEREZ MERA" in txt or "PEREZ" in txt, tipo
    # el certificado debe nombrar el cargo
    cert = DOCUMENTOS["certificado"][1]("1012", c)
    assert "GUARDIA" in pypdf.PdfReader(io.BytesIO(cert)).pages[0].extract_text()


def test_fotos_guardar_leer_borrar(app_db, tmp_path, monkeypatch):
    from core.repos import fotos

    png = bytes.fromhex("89504e470d0a1a0a") + b"x" * 50  # cabecera PNG + relleno
    fotos.guardar_foto("9999", png, "cara.png")
    r = fotos.leer_foto("9999")
    assert r is not None and r[1] == "image/png"
    assert fotos.borrar_foto("9999") is True
    assert fotos.leer_foto("9999") is None


def test_busqueda_avanzada_y_catalogos_xlsx():
    import io

    import openpyxl

    from core.excel.empleados_builders import busqueda_avanzada_xlsx, catalogos_xlsx

    filas = [{
        "empleado": "1012", "apellidos": "PEREZ", "nombres": "JUAN", "cedula": "0920116811",
        "cargo": "09", "cargo_nombre": "GUARDIA", "depto": "10", "depto_nombre": "OPERACIONES",
        "sueldo": 470.0, "telefono": "099", "email": "j@x.com", "estado": "ACT",
    }]
    wb = openpyxl.load_workbook(io.BytesIO(busqueda_avanzada_xlsx(filas)))
    ws = wb["EMPLEADOS"]
    assert [c.value for c in ws[1]][:4] == ["CÓDIGO", "APELLIDOS", "NOMBRES", "CÉDULA"]
    assert ws["F2"].value == "GUARDIA"

    wb2 = openpyxl.load_workbook(io.BytesIO(catalogos_xlsx({
        "FNC": [{"codigo": "09", "nombre": "GUARDIA"}],
        "DPT": [{"codigo": "10", "nombre": "OPS"}],
    })))
    assert "Cargos" in wb2.sheetnames and "Departamentos" in wb2.sheetnames


def test_grupos_cubren_las_6_pestanas_del_legado():
    assert set(empleados.GRUPOS) == {
        "Datos generales", "Ingresos / descuentos", "Otros datos",
        "Certificados / familiares", "Referencias",
    }
    # campos clave del legado presentes
    for c in ("COMPEN", "HOR100", "FONRESER", "CONYUGUE", "INCL_BAN", "NOM_FAM", "CERTVINF", "NUM_AFIL"):
        assert c in empleados.CAMPOS_EDITABLES


def test_parse_carga_masiva_ok_y_errores():
    data = _xlsx([
        ["EMPLEADO", "SUELDO", "TELEFONO"],
        ["1012", 850, "099"],
        ["", 900, ""],          # sin código -> se ignora
        ["2050", None, None],   # sin campos -> error
    ])
    filas, errores = parse_carga_masiva_empleados(data)
    assert len(filas) == 1
    assert filas[0]["EMPLEADO"] == "1012" and filas[0]["SUELDO"] == 850
    assert any("2050" in e for e in errores)


def test_parse_carga_masiva_sin_columna_empleado():
    _, errores = parse_carga_masiva_empleados(_xlsx([["CODIGO", "SUELDO"], ["1", 2]]))
    assert errores and "EMPLEADO" in errores[0]


def test_parse_biess_autodetecta_columnas():
    data = _xlsx([
        ["nombre", "cedula", "monto"],
        ["PEREIRA", "0920116811", "45.50"],
        ["LOPEZ", 1712345678, 120],
        ["basura", "abc", "xyz"],
    ])
    filas, _errores = parse_biess_quirografarios(data)
    ceds = {f["cedula"] for f in filas}
    assert "0920116811" in ceds and "1712345678" in ceds
    assert all(f["valor"] > 0 for f in filas)


def test_col_letra_a_idx():
    assert _col_letra_a_idx("A") == 0
    assert _col_letra_a_idx("E") == 4
    assert _col_letra_a_idx("AA") == 26


def test_parse_biess_manual_columnas_explicitas_y_consolida():
    data = _xlsx([
        ["encabezado", "", ""],
        ["basura", "x", "y"],
        ["", "0920116811", "10.50"],
        ["", "0920116811", "5"],       # misma cédula -> se suma
        ["", "1712345678", "$1,020.00"],  # $ y coma de miles como el legado
        ["", "", ""],
    ])
    filas, _err = parse_biess_manual(data, fila_inicio=3, col_cedula="B", col_valor="C")
    d = {f["cedula"]: f["valor"] for f in filas}
    assert d["0920116811"] == 15.5
    assert d["1712345678"] == 1020.0


def test_biess_autodetectar_devuelve_columnas():
    data = _xlsx([[f"h{c}" for c in range(4)]] + [
        ["PEREIRA", str(9200000000 + i), "aaa", round(10.5 + i, 2)] for i in range(8)
    ])
    det = biess_autodetectar(data)
    assert det["col_cedula"] == "B" and det["col_valor"] == "D"
    assert det["fila"] == 2 and det["confianza"] > 0
