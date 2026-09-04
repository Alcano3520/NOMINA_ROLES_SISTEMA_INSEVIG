"""Módulo 9: fórmulas legales de liquidación (Ecuador). Portadas de
`nucleo_modular` (LIQUIDACIONES_SISTEMA_INSEVIG), extracción fiel y ya
probada del `.pyw` que la empresa usa hoy en producción
(`Generador_Liquidaciones_INSEVIG.pyw`) — deben coincidir exactamente.

Nota histórica: la primera versión de este archivo portaba
`Liquidaciones_generador_CON_VACACIONES.pyw` (versión vieja/deprecada), con
un ancla de vacaciones en el día 1 del mes y un tope de 2 periodos para la
Décima Cuarta. Los tests de `test_periodos_vacaciones_por_mes_de_ingreso` y
`test_periodo_decima_tercera_dic_a_nov` se actualizaron para reflejar las
correcciones ya validadas en producción (ver docstrings de
`core.repos.liquidaciones.periodos_vacaciones`/`periodos_decima_tercera`).
"""

import datetime as dt

from core.excel.liquidaciones_builders import liquidaciones_xlsx
from core.repos import liquidaciones as lq


def test_dias360_como_excel():
    assert lq.dias360(dt.date(2024, 1, 15), dt.date(2024, 2, 15)) == 30
    assert lq.dias360(dt.date(2023, 2, 4), dt.date(2026, 2, 4)) == 3 * 360
    assert lq.dias360(dt.date(2024, 1, 31), dt.date(2024, 3, 31)) == 60


def test_periodos_vacaciones_por_mes_de_ingreso():
    # Ingresó 04/02/2023: el periodo ancla en el DÍA EXACTO de ingreso (04),
    # no en el día 1 del mes -- y como las vacaciones no caducan en Ecuador,
    # se devuelven TODOS los periodos pendientes desde el ingreso, no solo
    # los últimos 2 (el último queda recortado a la fecha de salida).
    p = lq.periodos_vacaciones(dt.date(2023, 2, 4), dt.date(2026, 3, 15))
    assert p == [
        (dt.date(2023, 2, 4), dt.date(2024, 2, 3)),
        (dt.date(2024, 2, 4), dt.date(2025, 2, 3)),
        (dt.date(2025, 2, 4), dt.date(2026, 2, 3)),
        (dt.date(2026, 2, 4), dt.date(2026, 3, 15)),  # periodo en curso, recortado a la salida
    ]
    # sale ANTES del aniversario (04/02) -> el periodo en curso llega solo
    # hasta la fecha de salida, no hasta el aniversario completo.
    p2 = lq.periodos_vacaciones(dt.date(2023, 2, 4), dt.date(2026, 1, 20))
    assert p2[-1] == (dt.date(2025, 2, 4), dt.date(2026, 1, 20))


def test_periodo_decima_tercera_dic_a_nov():
    # El periodo ACTUAL (en curso) se recorta a la fecha de salida -- no se
    # cuentan movimientos de meses que todavía no pasaron.
    p = lq.periodos_decima_tercera(dt.date(2022, 1, 1), dt.date(2026, 3, 10))
    assert p[-1] == (dt.date(2025, 12, 1), dt.date(2026, 3, 10), False)  # pago 24/12/26 > salida


def test_decima_cuarta_costa_vs_sierra():
    costa = lq.periodos_decima_cuarta(dt.date(2024, 1, 1), dt.date(2026, 5, 1), "COSTA")
    assert (dt.date(2025, 3, 1), dt.date(2026, 2, 28), True) in costa  # pagado 15/03/26 < 01/05
    sierra = lq.periodos_decima_cuarta(dt.date(2024, 1, 1), dt.date(2026, 5, 1), "SIERRA")
    assert any(i.month == 8 for i, _, _ in sierra)


def test_decima_cuarta_no_infla_con_mucha_antiguedad():
    # Corregido: antes se recorrían TODOS los años desde el ingreso -- para
    # alguien con 14 años de antigüedad esto sumaba ~13 periodos de décima
    # cuarta (todos ya pagados año a año en su momento) en vez de solo los
    # últimos 2. Caso real documentado: inflaba ~$5895 en vez de los ~$220
    # del único periodo realmente pendiente.
    p = lq.periodos_decima_cuarta(dt.date(2012, 1, 1), dt.date(2026, 5, 1), "COSTA")
    assert len(p) == 2


def test_decima_tercera_recorta_reingreso():
    # Reingreso a mitad del periodo calendario (01/12 -> 30/11): el inicio
    # real a considerar es la fecha de ingreso ACTUAL, no el 01/12 -- si no,
    # se suma sueldo de un ingreso anterior ya liquidado por separado.
    p = lq.periodos_decima_tercera(dt.date(2026, 2, 1), dt.date(2026, 6, 1))
    assert p == [(dt.date(2026, 2, 1), dt.date(2026, 6, 1), False)]


def test_desahucio_menos_de_un_anio_es_cero():
    assert lq.desahucio(dt.date(2025, 6, 1), dt.date(2026, 1, 1), 800.0) == 0.0  # 214 días


def test_desahucio_formula():
    # 3 años completos, sueldo 800 -> (800/4)*3 = 600
    d = lq.desahucio(dt.date(2023, 1, 1), dt.date(2026, 2, 1), 800.0)
    assert d == 600.0


def test_indemnizacion_despido():
    # motivo con DESPIDO, <3 años -> 3×sueldo
    assert lq.indemnizacion_despido(dt.date(2025, 1, 1), dt.date(2026, 6, 1), 800.0, "DESPIDO INTEMPESTIVO") == 2400.0
    # >=3 años -> años×sueldo
    assert lq.indemnizacion_despido(dt.date(2020, 1, 1), dt.date(2026, 6, 1), 800.0, "despido") == 6 * 800.0
    # motivo normal -> 0
    assert lq.indemnizacion_despido(dt.date(2020, 1, 1), dt.date(2026, 6, 1), 800.0, "RENUNCIA VOLUNTARIA") == 0.0


def test_sbu_por_anio_fallback():
    cfg = lq.ConfigLiquidacion()
    assert cfg.sbu(2026) == 482.0
    assert cfg.sbu(1999) == cfg.sbu(2020)   # antes del más antiguo
    assert cfg.sbu(2099) == cfg.sbu(2027)   # después del más reciente


def test_parse_linea():
    assert lq._parse_linea("0920116811, 15/02/2026, DESPIDO") == ("0920116811", "15/02/2026", "DESPIDO", "")
    assert lq._parse_linea("0920116811, 15/02/2026") == ("0920116811", "15/02/2026", "", "")
    assert lq._parse_linea("092, 15/02/2026, RENUNCIA, 01/01/2020") == ("092", "15/02/2026", "RENUNCIA", "01/01/2020")
    assert lq._parse_linea("solo_esto") is None


def test_vacaciones_pagadas_gozadas_degradan_sin_supabase(monkeypatch):
    """Si Supabase no responde/no está configurado, se devuelve None (no un
    dict vacío) -- el llamador no debe interpretar esto como "nada pagado"."""
    def _sin_conexion():
        raise RuntimeError("sin conexión")

    monkeypatch.setattr(lq.supabase_client, "get_client", _sin_conexion)
    assert lq.vacaciones_pagadas("0920116811") is None
    assert lq.vacaciones_gozadas("0920116811") is None


def test_total_vacaciones_sin_verificacion_solo_calcula_el_ultimo(monkeypatch):
    """Sin poder verificar contra vac_registros, solo se autocalcula el
    periodo MÁS RECIENTE -- cualquier periodo más antiguo con saldo se deja
    fuera (posible doble pago) y se alerta para revisión manual."""
    monkeypatch.setattr(lq, "vacaciones_pagadas", lambda cedula: None)
    monkeypatch.setattr(lq, "vacaciones_gozadas", lambda cedula: None)
    periodos = [
        (dt.date(2023, 2, 4), dt.date(2024, 2, 3)),
        (dt.date(2024, 2, 4), dt.date(2025, 2, 3)),
        (dt.date(2025, 2, 4), dt.date(2026, 2, 3)),
    ]
    total, alertas, detalle = lq.total_vacaciones_a_pagar("0920116811", [500.0, 500.0, 500.0], periodos)
    assert total == 500.0  # solo el último periodo
    assert len(alertas) == 2  # los 2 periodos más antiguos, con saldo, alertados
    assert detalle[0].estado == "SIN_VERIFICAR"
    assert detalle[-1].estado == "PENDIENTE"


def test_total_vacaciones_descarta_periodo_ya_pagado(monkeypatch):
    monkeypatch.setattr(lq, "vacaciones_pagadas", lambda cedula: {"2025-2026": True})
    monkeypatch.setattr(lq, "vacaciones_gozadas", lambda cedula: {})
    periodos = [(dt.date(2025, 2, 4), dt.date(2026, 2, 3))]
    total, alertas, detalle = lq.total_vacaciones_a_pagar("0920116811", [500.0], periodos)
    assert total == 0.0
    assert detalle[0].estado == "PAGADO"
    assert detalle[0].incluido is False


def test_total_vacaciones_prorratea_goce_parcial(monkeypatch):
    # 3 de los 15 días base ya gozados -> se paga (15-3)/15 = 12/15 del bruto.
    monkeypatch.setattr(lq, "vacaciones_pagadas", lambda cedula: {})
    monkeypatch.setattr(lq, "vacaciones_gozadas", lambda cedula: {"2025-2026": 3})
    periodos = [(dt.date(2025, 2, 4), dt.date(2026, 2, 3))]
    total, alertas, detalle = lq.total_vacaciones_a_pagar("0920116811", [300.0], periodos)
    assert total == 240.0  # 300 * 12/15
    assert detalle[0].estado == "GOZADO_PARCIAL"
    assert len(alertas) == 1


def test_total_vacaciones_suma_todos_los_periodos_pendientes(monkeypatch):
    """Las vacaciones NO caducan: con verificación disponible y ningún
    periodo pagado/gozado, se suman TODOS (no solo los últimos 2)."""
    monkeypatch.setattr(lq, "vacaciones_pagadas", lambda cedula: {})
    monkeypatch.setattr(lq, "vacaciones_gozadas", lambda cedula: {})
    periodos = [
        (dt.date(2023, 2, 4), dt.date(2024, 2, 3)),
        (dt.date(2024, 2, 4), dt.date(2025, 2, 3)),
        (dt.date(2025, 2, 4), dt.date(2026, 2, 3)),
    ]
    total, alertas, _detalle = lq.total_vacaciones_a_pagar("0920116811", [100.0, 200.0, 300.0], periodos)
    assert total == 600.0
    assert alertas == []


def test_decimo_anterior_se_incluye_por_defecto_en_el_total():
    """Corregido: la extracción inicial de este archivo nunca sumaba el
    décimo ANTERIOR al total (subpagaba la liquidación). Por defecto ahora
    SÍ se incluye -- se puede excluir explícitamente pasando
    incluir_dec13_anterior/incluir_dec14_anterior=False."""
    import inspect

    firma = inspect.signature(lq.procesar_empleado)
    assert firma.parameters["incluir_dec13_anterior"].default is True
    assert firma.parameters["incluir_dec14_anterior"].default is True


def test_excel_liquidaciones_valido():
    import io

    import openpyxl

    liq = lq.Liquidacion(
        empleado="1012", nombre="PEREIRA JUAN", cedula="0920116811", cargo="01",
        depto="10", seccion="", sueldo_base=800.0, fecha_ingreso="2023-01-01",
        fecha_salida="2026-02-01", motivo_salida="RENUNCIA", dias_trabajados=1127,
        campos={"TOTAL_INGRESOS": 1500.0, "TOTAL_DESCUENTOS": 200.0, "TOTAL_A_RECIBIR": 1300.0,
                "DESAHUCIO": 600.0, "VACACIONES_CALCULADAS": 400.0},
    )
    data = liquidaciones_xlsx([liq])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["FORMATO"]
    headers = [c.value for c in ws[1]]
    assert "MOTIVO DE SALIDA" in headers
    assert "INDEM. DESPIDO" in headers
    assert "TOTAL VALORES A LIQUIDAR" in headers
    assert ws.max_row == 2


def _liq_ejemplo(**overrides) -> "lq.Liquidacion":
    base = dict(
        empleado="1012", nombre="PEREIRA JUAN", cedula="0920116811",
        cargo="GUARDIA", depto="OPERACIONES", seccion="SEC1", sueldo_base=460.0,
        fecha_ingreso="2020-03-15", fecha_salida="2026-06-30", motivo_salida="RENUNCIA VOLUNTARIA",
        dias_trabajados=2298, apellidos="PEREIRA", nombres="JUAN",
        campos={
            "SUELDO": 460.0, "HORAS_25": 10, "HORAS_50": 0, "HORAS_100": 0,
            "VAL_SOBT_25": 15.6, "VAL_SOBT_50": 0.0, "VAL_SOBT_100": 0.0,
            "FONDO_RESERVA": 38.3, "MANIOBRAS": 0.0, "MOVILIZACION": 0.0,
            "REEMBOLSOS": 0.0, "BONIFICACION": 0.0,
            "VACACIONES_ANTERIOR": 200.0, "VACACIONES_ULTIMO": 220.0, "VACACIONES_CALCULADAS": 18.33,
            "DECIMA_TERCERA_ANTERIOR": 38.3, "DECIMA_TERCERA_ACTUAL": 38.3,
            "DECIMA_CUARTA_ANTERIOR": 0.0, "DECIMA_CUARTA_ACTUAL": 240.0,
            "DESAHUCIO": 690.0, "INDEM_DESPIDO": 0.0,
            "APORT_IESS": 43.47, "PRESTAMOS_QUIROGRAFARIOS": 0.0, "PRESTAMOS_COMPANIA": 0.0,
            "ANTICIPO_SUELDO": 0.0, "ANTICIPOS_OTROS": 0.0, "ANTICIPOS_SURTIDOS": 0.0,
            "APORT_IESS_CONYUGE": 0.0, "PENSION_ALIMENTICIA": 0.0, "PRESTAMO_HIPOTECARIO": 0.0,
            "IMPUESTO_RENTA": 0.0, "ANTICIPOS_OTROS_L": 0.0, "ANTICIPO_L_DESAHUCIO": 0.0,
            "TOTAL_INGRESOS": 1000.0, "TOTAL_DESCUENTOS": 43.47, "TOTAL_A_RECIBIR": 956.53,
        },
    )
    base.update(overrides)
    return lq.Liquidacion(**base)


def test_clasificar_tipo_liquidacion():
    assert lq.clasificar_tipo_liquidacion("RENUNCIA VOLUNTARIA") == "renuncia"
    assert lq.clasificar_tipo_liquidacion("DESPIDO INTEMPESTIVO") == "despido"
    assert lq.clasificar_tipo_liquidacion("VISTO BUENO") == "visto_bueno"
    assert lq.clasificar_tipo_liquidacion("FIN DE CONTRATO") == "termino_contrato"
    assert lq.clasificar_tipo_liquidacion("FALLECIMIENTO") == "muerte"
    assert lq.clasificar_tipo_liquidacion("JUBILACION") == "jubilacion"
    assert lq.clasificar_tipo_liquidacion(None) == "otro"
    assert lq.clasificar_tipo_liquidacion("motivo raro") == "otro"


def test_mapear_registro_totales_y_horas():
    liq = _liq_ejemplo()
    cfg = lq.ConfigLiquidacion()
    registro = lq._mapear_registro(liq, "generada", cfg, usuario="tester")
    assert registro["empleado_codigo"] == "1012"
    assert registro["empleado_cedula"] == "0920116811"
    assert registro["empleado_apellidos"] == "PEREIRA" and registro["empleado_nombres"] == "JUAN"
    assert registro["tipo_liquidacion"] == "renuncia"
    assert registro["estado"] == "generada"
    assert registro["total_liquido"] == 956.53
    assert registro["decimo_tercero"] == 76.6  # 38.3 + 38.3
    assert registro["horas_25_cantidad"] == 10
    assert registro["horas_25_valor_hora"] == round(15.6 / 10, 2)
    assert registro["created_by"] == "tester" and registro["updated_by"] == "tester"


def test_mapear_registro_sin_usuario_usa_sistema():
    registro = lq._mapear_registro(_liq_ejemplo(), "borrador", lq.ConfigLiquidacion(), usuario="")
    assert registro["created_by"] == "Sistema"
    assert "Simulación" in registro["observaciones"]


def test_construir_conceptos_omite_valores_en_cero():
    conceptos = lq._construir_conceptos(_liq_ejemplo())
    codigos = {c["concepto_codigo"] for c in conceptos}
    assert "SUELDO" in codigos and "DESAHUCIO" in codigos
    assert "MANIOBRAS" not in codigos  # viene en 0.0 -> se omite
    assert all(c["valor_total"] != 0 for c in conceptos)
    assert [c["orden"] for c in conceptos] == list(range(len(conceptos)))


def test_guardar_liquidacion_rechaza_estado_invalido():
    ok, msg = lq.guardar_liquidacion(
        _liq_ejemplo(), "estado_invalido", lq.ConfigLiquidacion(), usuario="t", roles=set()
    )
    assert not ok and "inválido" in msg


def test_guardar_liquidacion_rechaza_si_hay_error():
    liq = _liq_ejemplo(error="empleado no encontrado")
    ok, msg = lq.guardar_liquidacion(liq, "borrador", lq.ConfigLiquidacion(), usuario="t", roles=set())
    assert not ok and msg == "empleado no encontrado"


def test_liquidacion_pdf_genera_documento_valido():
    from core.pdf.liquidacion_individual import liquidacion_pdf

    liq = _liq_ejemplo(detalle_vacaciones=[
        lq.DetalleVacacionesPeriodo("2024-2025", "PAGADO", 0.0, 200.0, False),
        lq.DetalleVacacionesPeriodo("2025-2026", "GOZADO_PARCIAL", 5.0, 220.0, True),
    ])
    data = liquidacion_pdf(liq, mostrar_insumos=True, es_simulacion=True)
    assert data[:4] == b"%PDF"
    assert len(data) > 500

    data_real = liquidacion_pdf(liq, es_simulacion=False)
    assert data_real[:4] == b"%PDF"


def test_reconstruir_liquidacion_desde_supabase():
    registro = {
        "empleado_codigo": "1012", "empleado_cedula": "0920116811",
        "empleado_apellidos": "PEREIRA", "empleado_nombres": "JUAN",
        "cargo": "GUARDIA", "puesto_servicio": "OPERACIONES", "seccion": "SEC1",
        "fecha_ingreso": "2020-03-15", "fecha_salida": "2026-06-30",
        "motivo": "RENUNCIA VOLUNTARIA", "dias_trabajados": 2298,
        "fondo_reserva": 38.3, "vacaciones_pendientes": 18.33,
        "bonificacion_desahucio": 690.0, "total_descuentos": 43.47, "total_liquido": 956.53,
        "horas_25_cantidad": 10, "horas_50_cantidad": 0, "horas_100_cantidad": 0,
    }
    conceptos = [
        {"concepto_codigo": "SUELDO", "valor_total": 460.0},
        {"concepto_codigo": "SOBT_25", "valor_total": 15.6},
        {"concepto_codigo": "DESAHUCIO", "valor_total": 690.0},
        {"concepto_codigo": "IESS", "valor_total": 43.47},
    ]
    liq = lq.reconstruir_liquidacion(registro, conceptos)
    assert liq.empleado == "1012" and liq.cedula == "0920116811"
    assert liq.nombre == "PEREIRA JUAN"
    assert liq.campos["SUELDO"] == 460.0
    assert liq.campos["VAL_SOBT_25"] == 15.6
    assert liq.campos["DESAHUCIO"] == 690.0
    assert liq.campos["APORT_IESS"] == 43.47
    assert liq.campos["TOTAL_A_RECIBIR"] == 956.53
    assert liq.campos["HORAS_25"] == 10

    from core.pdf.liquidacion_individual import liquidacion_pdf

    data = liquidacion_pdf(liq, es_simulacion=False)
    assert data[:4] == b"%PDF"


def test_liquidacion_pdf_con_error_no_revienta():
    from core.pdf.liquidacion_individual import liquidacion_pdf

    liq = lq.Liquidacion("", "", "0900000000", "", "", "", 0.0, "", "", "", 0,
                         error="empleado no encontrado")
    data = liquidacion_pdf(liq)
    assert data[:4] == b"%PDF"
