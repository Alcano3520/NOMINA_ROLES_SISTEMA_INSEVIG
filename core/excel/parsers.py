"""Parseo de Excel subido por el usuario: carga masiva de empleados y BIESS.

Devuelven `(filas_validas, errores)` — nunca lanzan por datos malos.
"""

from __future__ import annotations

import io

import openpyxl

from core.utils import normalizar_cedula


def _hoja(datos: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    return wb[wb.sheetnames[0]]


def parse_carga_masiva_empleados(datos: bytes) -> tuple[list[dict], list[str]]:
    ws = _hoja(datos)
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return [], ["El archivo está vacío."]
    headers = [str(h).strip() if h is not None else "" for h in filas[0]]
    if "EMPLEADO" not in headers:
        return [], ["Falta la columna obligatoria 'EMPLEADO'."]
    idx_emp = headers.index("EMPLEADO")
    validas: list[dict] = []
    errores: list[str] = []
    for n, fila in enumerate(filas[1:], start=2):
        cod = fila[idx_emp]
        if cod in (None, ""):
            continue
        registro = {"EMPLEADO": str(cod).strip()}
        for i, h in enumerate(headers):
            if h and h != "EMPLEADO" and i < len(fila) and fila[i] not in (None, ""):
                registro[h] = fila[i]
        if len(registro) == 1:
            errores.append(f"Fila {n}: sin campos a actualizar para {cod}.")
            continue
        validas.append(registro)
    return validas, errores


# ── BIESS quirografarios ─────────────────────────────────────────────────────


def _es_cedula(valor: object) -> bool:
    if valor is None:
        return False
    s = str(valor).replace(".0", "").strip()
    return s.isdigit() and 8 <= len(s) <= 10


def _es_monto(valor: object) -> bool:
    try:
        return float(str(valor).replace(",", "")) > 0
    except (TypeError, ValueError):
        return False


def _col_letra_a_idx(col: str) -> int:
    """'A'->0, 'E'->4, 'AA'->26."""
    col = (col or "").strip().upper()
    if not col or not col.isalpha():
        raise ValueError(f"Columna Excel inválida: {col!r}")
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _idx_a_col_letra(idx: int) -> str:
    idx += 1
    out = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        out = chr(65 + r) + out
    return out


def biess_autodetectar(datos: bytes) -> dict:
    """Sugiere fila de inicio (1-based) y letras de columna cédula/valor por el
    CONTENIDO del Excel BIESS. Devuelve {'fila','col_cedula','col_valor','confianza'}.
    Porta `biess_autodetectar_columnas` del `.pyw`."""
    ws = _hoja(datos)
    filas = [list(f) for f in ws.iter_rows(values_only=True)]
    if not filas:
        return {"fila": 1, "col_cedula": "A", "col_valor": "B", "confianza": 0.0}
    ncols = max((len(f) for f in filas), default=0)
    mejor_c, score_c, primera = -1, 0, 0
    for c in range(ncols):
        validos, prim = 0, None
        for i, f in enumerate(filas):
            if c < len(f) and _es_cedula(f[c]):
                validos += 1
                if prim is None:
                    prim = i
        if validos > score_c:
            mejor_c, score_c, primera = c, validos, (prim or 0)
    if mejor_c < 0 or score_c < 5:
        return {"fila": 1, "col_cedula": "A", "col_valor": "B", "confianza": 0.0}
    mejor_v, score_v = -1, 0
    for c in range(ncols):
        if c == mejor_c:
            continue
        s = sum(1 for f in filas[primera:] if c < len(f) and _es_monto(f[c]))
        if s > score_v:
            mejor_v, score_v = c, s
    conf = min(1.0, score_c / max(1, len(filas) - primera))
    return {
        "fila": primera + 1,
        "col_cedula": _idx_a_col_letra(mejor_c),
        "col_valor": _idx_a_col_letra(mejor_v) if mejor_v >= 0 else "B",
        "confianza": round(conf, 2),
    }


def biess_diagnostico(datos: bytes, *, filas: int = 15, cols: int = 12) -> list[list[str]]:
    """Primeras `filas`x`cols` celdas como texto, para que el usuario ubique las
    columnas a mano. La primera fila devuelta son las letras de columna."""
    ws = _hoja(datos)
    todas = list(ws.iter_rows(values_only=True))[:filas]
    ancho = min(cols, max((len(f) for f in todas), default=0))
    out = [["#", *[_idx_a_col_letra(c) for c in range(ancho)]]]
    for n, f in enumerate(todas, 1):
        out.append([str(n), *["" if (c >= len(f) or f[c] is None) else str(f[c])[:22] for c in range(ancho)]])
    return out


def parse_biess_manual(
    datos: bytes, *, fila_inicio: int, col_cedula: str, col_valor: str,
) -> tuple[list[dict], list[str]]:
    """Lee el Excel BIESS con la fila de inicio y las columnas indicadas a mano
    (letras de Excel). Consolida cédulas duplicadas sumando el valor."""
    ic, iv = _col_letra_a_idx(col_cedula), _col_letra_a_idx(col_valor)
    ws = _hoja(datos)
    filas = list(ws.iter_rows(values_only=True))
    fila_inicio = max(1, fila_inicio)
    consolidado: dict[str, float] = {}
    errores: list[str] = []
    for n, f in enumerate(filas[fila_inicio - 1:], start=fila_inicio):
        if ic >= len(f) or iv >= len(f) or f[ic] is None or f[iv] is None:
            continue
        ced = normalizar_cedula(f[ic])
        if not ced or ced == "0000000000":
            continue
        try:
            val = round(float(str(f[iv]).replace("$", "").replace(",", "").strip()), 2)
        except (TypeError, ValueError):
            errores.append(f"Fila {n}: valor no numérico ({f[iv]!r}).")
            continue
        if not 0.01 <= val <= 100_000:
            continue
        consolidado[ced] = round(consolidado.get(ced, 0.0) + val, 2)
    return [{"cedula": c, "valor": v} for c, v in consolidado.items()], errores


def parse_biess_quirografarios(datos: bytes) -> tuple[list[dict], list[str]]:
    """Autodetecta columnas de cédula y valor por contenido (porta la lógica de
    `REGISTRAR_PRESTAMOS_UNIFICADO.pyw`). Devuelve [{'cedula','valor'}]."""
    ws = _hoja(datos)
    filas = [list(f) for f in ws.iter_rows(values_only=True) if any(c is not None for c in f)]
    if not filas:
        return [], ["Archivo vacío."]
    ncols = max(len(f) for f in filas)
    # elegir la columna con más cédulas y la que más montos
    col_ced = max(range(ncols), key=lambda c: sum(_es_cedula(f[c]) for f in filas if c < len(f)))
    col_val = max(
        (c for c in range(ncols) if c != col_ced),
        key=lambda c: sum(_es_monto(f[c]) for f in filas if c < len(f)),
        default=-1,
    )
    if col_val < 0:
        return [], ["No se detectó una columna de valores."]
    out: list[dict] = []
    errores: list[str] = []
    for n, f in enumerate(filas, 1):
        if col_ced >= len(f) or not _es_cedula(f[col_ced]):
            continue
        if col_val >= len(f) or not _es_monto(f[col_val]):
            errores.append(f"Fila {n}: cédula sin valor válido.")
            continue
        out.append(
            {
                "cedula": normalizar_cedula(f[col_ced]),
                "valor": round(float(str(f[col_val]).replace(",", "")), 2),
            }
        )
    return out, errores
