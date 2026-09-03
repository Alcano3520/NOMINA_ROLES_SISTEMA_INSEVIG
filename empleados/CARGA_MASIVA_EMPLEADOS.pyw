#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CARGA MASIVA DE EMPLEADOS - INSEVIG
Herramienta independiente para edición masiva de empleados y observaciones
vía plantilla Excel. Extraída de SISTEMA_GESTION_EMPLEADOS_10.pyw.
Compatible Linux/Windows.
"""

import os, sys, threading, logging
from datetime import datetime

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('carga_masiva_debug.log'),
        logging.StreamHandler(sys.stderr)
    ]
)
LOG = logging.getLogger(__name__)

_openssl = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'openssl_legacy.cnf')
if os.path.exists(_openssl):
    os.environ['OPENSSL_CONF'] = _openssl

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pyodbc

# ── Palette INSEVIG DARK MODE (igual a SISTEMA_GESTION_EMPLEADOS_10.pyw) ──
COL_BG       = '#1E1E1E'
COL_HEADER   = '#0D1B2A'
COL_ACCENT   = '#4A9EFF'
COL_PEND     = '#FF9F43'
COL_OK       = '#2ED573'
COL_DANGER   = '#FF6B6B'
COL_WHITE    = '#FFFFFF'
COL_GRAY     = '#A0A0A0'
COL_ENTRY_BG = '#2D2D2D'
COL_CARD     = '#2D2D2D'
COL_TEXT     = '#E0E0E0'
COL_BORDER   = '#404040'

FONT_DEFAULT = ('Segoe UI', 10)
FONT_SMALL   = ('Segoe UI', 9)
FONT_LABEL   = ('Segoe UI', 10, 'bold')
FONT_HEAD    = ('Segoe UI', 11, 'bold')
FONT_TITLE   = ('Segoe UI', 14, 'bold')

# ── SQL Server ──────────────────────────────────────────────────────
SQL_CFG = {
    'driver':   'ODBC Driver 17 for SQL Server',
    'server':   '192.168.2.115',
    'database': 'insevig',
    'uid':      'sa',
    'pwd':      'puntosoft123*',
}
SQL_FILTER = "CODEMP='10' AND CODSUC='10'"


def _get_sql_conn():
    """Intenta drivers en orden de prioridad."""
    drivers = [
        'ODBC Driver 17 for SQL Server',
        'ODBC Driver 18 for SQL Server',
        'ODBC Driver 13 for SQL Server',
        'ODBC Driver 11 for SQL Server',
        'SQL Server',
    ]
    cfg = SQL_CFG.copy()
    for drv in drivers:
        try:
            cfg['driver'] = drv
            cs = (f"DRIVER={{{drv}}};SERVER={cfg['server']};DATABASE={cfg['database']};"
                  f"UID={cfg['uid']};PWD={cfg['pwd']};Encrypt=No;TrustServerCertificate=yes")
            return pyodbc.connect(cs, timeout=10)
        except Exception:
            continue
    raise RuntimeError("No se pudo conectar a SQL Server con ningún driver ODBC.")


def _bind_scroll_rueda(widget, canvas):
    """Une `widget` (y todos sus descendientes) al scroll de `canvas` con
    la rueda del mouse. Bindear solo el canvas no basta: los widgets hijos
    (Entry, Label, Frame, etc.) reciben el evento de rueda antes que el
    canvas y lo consumen, así que hay que bindear recursivamente todo el
    árbol de contenido desplazable."""
    widget.bind('<MouseWheel>', lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units'), add='+')
    for child in widget.winfo_children():
        _bind_scroll_rueda(child, canvas)


# ═══════════════════════════════════════════════════════════════════
# Edición Masiva
# ═══════════════════════════════════════════════════════════════════
class EdicionMasivaFrame:
    def __init__(self, window, app):
        self.window = window
        self.app = app
        self.conn = app.conn
        self.datos_validados = None
        self._build()

    def _build(self):
        main = ttk.Frame(self.window)
        main.pack(fill='both', expand=True, padx=16, pady=16)

        tk.Label(main, text="📊 Edición Masiva por Plantilla Excel",
                 font=('Segoe UI', 14, 'bold'), fg=COL_HEADER, bg=COL_BG).pack(pady=(0, 8))

        # Instrucciones
        info = tk.Frame(main, bg=COL_CARD, relief='solid', borderwidth=1)
        info.pack(fill='x', padx=8, pady=(0, 8))
        tk.Label(info, text="ℹ️  INSTRUCCIONES:", font=FONT_LABEL, fg=COL_ACCENT, bg=COL_CARD).pack(anchor='w', padx=6, pady=(4, 2))
        txt_info = ("1. Selecciona los campos que quieres modificar en la pestaña 'Descargar'\n"
                   "2. Descarga la plantilla Excel con los datos actuales\n"
                   "3. Edita los valores en Excel (solo las celdas que cambiarán)\n"
                   "4. Sube el archivo en la pestaña 'Cargar y Aplicar'\n"
                   "5. Valida los cambios (muestra un resumen)\n"
                   "6. Aplica los cambios a la base de datos")
        tk.Label(info, text=txt_info, font=FONT_SMALL, fg=COL_TEXT, bg=COL_CARD,
                justify='left').pack(anchor='w', padx=6, pady=(2, 4))

        nb = ttk.Notebook(main)
        nb.pack(fill='both', expand=True)

        # Tab 1: Descargar
        t1 = ttk.Frame(nb)
        nb.add(t1, text="1. Descargar Plantilla")
        tk.Label(t1, text="Campos a incluir:", font=FONT_LABEL,
                 bg=COL_BG).pack(anchor='w', padx=8, pady=4)

        # Botones Seleccionar/Deseleccionar
        btn_bar = ttk.Frame(t1)
        btn_bar.pack(fill='x', padx=8, pady=4)
        ttk.Button(btn_bar, text="✓ Seleccionar Todo", command=self._seleccionar_todo).pack(side='left', padx=2)
        ttk.Button(btn_bar, text="✗ Deseleccionar Todo", command=self._deseleccionar_todo).pack(side='left', padx=2)

        self._campos_vars = {}
        cf = ttk.Frame(t1)
        cf.pack(fill='x', padx=8)
        campos = [
            ('NOMBRES','Nombres'),('APELLIDOS','Apellidos'),('CEDULA','Cédula'),
            ('SEXO','Sexo'),('ESTADO_CI','E.Civil'),('LUGAR_NAC','Lugar Nac.'),
            ('FECHA_NAC','F.Nacimiento'),('DIRECCION','Dirección'),('PROVINCIA','Provincia'),
            ('CANTON','Cantón'),('PARROQUIA','Parroquia'),('NACIONAL','Nacionalidad'),
            ('TELEFONO','Teléfono'),('RPCAM','2do Teléfono'),('emp_mail','Email'),
            ('FECHA_ING','F.Ingreso'),('FECHA_SAL','F.Salida'),('DEPTO','Depto'),('SECCION','Sección'),
            ('CARGO','Cargo'),('ESTADO','Estado'),('ACTIVIDAD','Actividad'),
            ('CONYUGUE','Cónyuge'),('TIPO_TRA','Tipo Trab.'),
            ('SUELDO','Sueldo'),('BONIFI','Bonif.'),('COMPEN','Compens.'),
            ('TRANSP','Transporte'),('LUNCH','Lunch'),('MOVILIZA','Moviliza'),
            ('HOR25','H.25%'),('HOR50','H.50%'),('HOR100','H.100%'),('DECIMO3','D3'),('DECIMO4','D4'),
            ('VACACION','Vacaciones'),('CARGAS','Cargas'),('ULTLIQ','Últ.Líq.'),('ULTDIATRA','Últ.Día Trab.'),
            ('DIAS_TRA','Días Trab.'),('TIP_SAN','T.Sangre'),('TIPO_PGO','T.Pago'),
            ('CODCTA','Cód.Cuenta'),('CTADPT','C.Depto'),('CTAAUX','C.Aux'),
            ('RUTA4','Ruta'),('CTA_CTE','C.Corriente'),('CTA_AHO','C.Ahorros'),
            ('INCL_ROL','Incl.Rol'),('INCL_BAN','Incl.Banco'),
            ('NOM_FAM','Familiares'),('DIR_FAM','Dir.Fam'),('TEL_FAM','Tel.Fam'),
            ('NOM_NO_FAM','No Familiares'),('DIR_NO_FAM','Dir.NoFam'),('TEL_NO_FAM','Tel.NoFam'),
            ('CED_MIL','C.Militar'),('EDAD','Edad'),('IDVOTA','C.Votación'),
            ('LICCOND','L.Conducir'),('CODIESS','C.IESS'),('ID_CONADIS','C.Conadis'),('TITULO','Título'),
            ('ANIO_EST','Años Est.'),('CERTVINF','Cert.Violencia'),('MANIOBRAS','Maniobras'),
            ('NUM_AFIL','Afil.IESS'),('FZA_PUB','Fuerza Pública'),('SER_MIL','Serv.Militar'),
            ('OBSERV','Visita Domic.')
        ]
        for i, (k, n) in enumerate(campos):
            v = tk.BooleanVar(value=(i < 10))
            self._campos_vars[k] = v
            tk.Checkbutton(cf, text=n, variable=v, font=FONT_SMALL,
                           bg=COL_BG, fg=COL_TEXT, selectcolor=COL_ACCENT).grid(row=i//6, column=i%6, sticky='w', padx=4, pady=1)

        btnf = ttk.Frame(t1)
        btnf.pack(pady=8)
        ttk.Button(btnf, text="⬇ Descargar Plantilla Excel",
                   command=self._descargar, style='Accent.TButton').pack(side='left', padx=4)

        # Tab 2: Cargar
        t2 = ttk.Frame(nb)
        nb.add(t2, text="2. Cargar y Aplicar")
        t2_content = ttk.Frame(t2)
        t2_content.pack(fill='both', expand=True, padx=8, pady=8)

        tk.Label(t2_content, text="Archivo Excel:", font=FONT_LABEL, bg=COL_BG).pack(anchor='w', pady=(0, 4))
        self._archivo_var = tk.StringVar()
        rf = ttk.Frame(t2_content)
        rf.pack(fill='x', pady=(0, 4))
        ttk.Entry(rf, textvariable=self._archivo_var, width=50).pack(side='left', padx=(0, 4))
        ttk.Button(rf, text="📁 Seleccionar", command=self._seleccionar).pack(side='left', padx=2)

        bf = ttk.Frame(t2_content)
        bf.pack(fill='x', pady=(0, 8))
        ttk.Button(bf, text="✓ Validar Cambios", command=self._validar,
                  style='Accent.TButton').pack(side='left', padx=2)
        self._btn_aplicar = ttk.Button(bf, text="⚡ Aplicar Cambios",
                                       command=self._aplicar, state='disabled',
                                       style='Accent.TButton')
        self._btn_aplicar.pack(side='left', padx=2)

        tk.Label(t2_content, text="Resumen de Cambios:", font=FONT_LABEL, bg=COL_BG).pack(anchor='w', pady=(8, 4))
        res_frame = ttk.Frame(t2_content)
        res_frame.pack(fill='both', expand=True)
        self._resultado = tk.Text(res_frame, font=('Consolas', 9),
                                  bg=COL_ENTRY_BG, fg=COL_TEXT, wrap='word', relief='solid', borderwidth=1)
        self._resultado.pack(side='left', fill='both', expand=True)
        vsb = ttk.Scrollbar(res_frame, orient='vertical', command=self._resultado.yview)
        vsb.pack(side='right', fill='y')
        self._resultado.configure(yscrollcommand=vsb.set)

    def _seleccionar_todo(self):
        for v in self._campos_vars.values():
            v.set(True)

    def _deseleccionar_todo(self):
        for v in self._campos_vars.values():
            v.set(False)

    def _descargar(self):
        if not self.conn:
            messagebox.showerror("Error", "Sin conexión a BD")
            return
        campos_sel = [k for k, v in self._campos_vars.items() if v.get()]
        if not campos_sel:
            messagebox.showwarning("Aviso", "Seleccione al menos un campo")
            return
        def tarea():
            try:
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cols = ['EMPLEADO'] + campos_sel
                    cur.execute(f"SELECT {', '.join(cols)} FROM RPEMPLEA WHERE {SQL_FILTER} ORDER BY EMPLEADO")
                    rows = cur.fetchall()
                from openpyxl import Workbook
                from openpyxl.styles import Font
                wb = Workbook()
                ws = wb.active
                ws.title = "EMPLEADOS"
                for ci, h in enumerate(cols, 1):
                    ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
                for ri, r in enumerate(rows, 2):
                    for ci, v in enumerate(r, 1):
                        ws.cell(row=ri, column=ci, value=v)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fn = f"PLANTILLA_EMPLEADOS_{ts}.xlsx"
                wb.save(fn)
                self.window.after(0, lambda: messagebox.showinfo("Éxito", f"Plantilla creada:\n{fn}"))
            except ImportError:
                self.window.after(0, lambda: messagebox.showerror("Error", "Requiere openpyxl:\npip install openpyxl"))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _seleccionar(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if f:
            self._archivo_var.set(f)
            self._btn_aplicar.config(state='disabled')

    def _validar(self):
        arch = self._archivo_var.get()
        if not arch:
            messagebox.showwarning("Aviso", "Seleccione un archivo")
            return
        def tarea():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(arch, data_only=True)
                ws = wb['EMPLEADOS']
                headers = [c.value for c in ws[1]]
                if 'EMPLEADO' not in headers:
                    self.window.after(0, lambda: messagebox.showerror("Error", "Columna EMPLEADO requerida"))
                    return
                datos = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        datos.append(row)
                if not datos:
                    self.window.after(0, lambda: messagebox.showerror("Error", "Sin datos"))
                    return
                ei = headers.index('EMPLEADO')
                cambios = []
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    for row in datos:
                        emp = row[ei]
                        cur.execute("SELECT COUNT(*) FROM RPEMPLEA WHERE EMPLEADO=?", (emp,))
                        if cur.fetchone()[0] == 0:
                            continue
                        cambios_emp = {}
                        for i, h in enumerate(headers):
                            if h != 'EMPLEADO' and i < len(row) and row[i] is not None and str(row[i]).strip():
                                cambios_emp[h] = row[i]
                        if cambios_emp:
                            cambios.append({'codigo': emp, 'cambios': cambios_emp})
                self.datos_validados = cambios
                self.window.after(0, lambda: self._mostrar_validacion(cambios))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_validacion(self, cambios):
        self._resultado.delete(1.0, 'end')
        t = sum(len(c['cambios']) for c in cambios)
        self._resultado.insert('end', f"Empleados con cambios: {len(cambios)}\n")
        self._resultado.insert('end', f"Total cambios: {t}\n\n")
        for c in cambios:
            self._resultado.insert('end', f"  {c['codigo']}: {', '.join(c['cambios'].keys())}\n")
        if cambios:
            self._btn_aplicar.config(state='normal')

    def _aplicar(self):
        if not self.datos_validados:
            return
        if not messagebox.askyesno("Confirmar", "¿Aplicar cambios masivos?", icon='warning'):
            return
        self._btn_aplicar.config(state='disabled')
        def tarea():
            try:
                ok, err = 0, 0
                fallidos = []
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    for emp in self.datos_validados:
                        try:
                            sets = '=?, '.join(emp['cambios'].keys()) + '=?'
                            vals = list(emp['cambios'].values()) + [emp['codigo']]
                            cur.execute(
                                f"UPDATE RPEMPLEA SET {sets} WHERE EMPLEADO=? AND {SQL_FILTER}", vals)
                            ok += 1
                        except Exception as e:
                            LOG.error("Error actualizando empleado %s: %s", emp['codigo'], e)
                            fallidos.append(str(emp['codigo']))
                            err += 1
                    self.conn.commit()
                detalle = f"\n\nEmpleados con error: {', '.join(fallidos[:15])}" if fallidos else ""
                if len(fallidos) > 15:
                    detalle += f" ... y {len(fallidos) - 15} más"
                self.window.after(0, lambda: messagebox.showinfo(
                    "Completado", f"Actualizados: {ok}, Errores: {err}{detalle}"))
                self.window.after(0, self.app._cargar_lista)
                self.datos_validados = None
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════
# Agregar Observaciones Masivas
# ═══════════════════════════════════════════════════════════════════
class ObservacionesMasivasFrame:
    def __init__(self, window, app):
        self.window = window
        self.app = app
        self.conn = app.conn
        self.datos_validados = None

        nb = ttk.Notebook(window)
        nb.pack(fill='both', expand=True, padx=8, pady=8)

        t1 = ttk.Frame(nb)
        nb.add(t1, text="Descargar Plantilla")
        self._build_tab1(t1)

        t2 = ttk.Frame(nb)
        nb.add(t2, text="Cargar Observaciones")
        self._build_tab2(t2)

    def _build_tab1(self, parent):
        inst = tk.Label(parent,
            text="1. Selecciona empleados\n2. Elige una fecha\n3. Descarga la plantilla\n4. Agrega observaciones en la columna 'texto_obs'",
            font=FONT_LABEL, bg=COL_BG, justify='left')
        inst.pack(anchor='w', padx=12, pady=10)

        row1 = tk.Frame(parent, bg=COL_BG)
        row1.pack(fill='x', padx=10, pady=8)
        tk.Label(row1, text="Empleados:", font=FONT_LABEL, bg=COL_BG).pack(side='left')
        self._filtro_emp_var = tk.StringVar(value="ACTIVOS")
        cb = ttk.Combobox(row1, textvariable=self._filtro_emp_var,
                         values=["ACTIVOS", "INACTIVOS", "TODOS"], width=15, state='readonly')
        cb.pack(side='left', padx=(6, 0))

        row2 = tk.Frame(parent, bg=COL_BG)
        row2.pack(fill='x', padx=10, pady=8)
        tk.Label(row2, text="Fecha Obs:", font=FONT_LABEL, bg=COL_BG).pack(side='left')
        self._fecha_obs_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        e = ttk.Entry(row2, textvariable=self._fecha_obs_var, width=15)
        e.pack(side='left', padx=(6, 0))
        tk.Label(row2, text="(YYYY-MM-DD)", font=FONT_SMALL, bg=COL_BG).pack(side='left', padx=(6, 0))

        ttk.Button(parent, text="⬇ Descargar Plantilla",
                  command=self._descargar_plantilla).pack(fill='x', padx=10, pady=10)

        tk.Label(parent, text="Estado:", font=FONT_LABEL, bg=COL_BG).pack(anchor='w', padx=10, pady=(10, 0))
        self._resultado_t1 = tk.Text(parent, height=10, width=80, font=FONT_SMALL, bg=COL_ENTRY_BG, fg=COL_WHITE)
        self._resultado_t1.pack(fill='both', expand=True, padx=10, pady=(6, 10))

    def _build_tab2(self, parent):
        t2_content = ttk.Frame(parent)
        t2_content.pack(fill='both', expand=True, padx=8, pady=8)

        tk.Label(t2_content, text="Archivo Excel:", font=FONT_LABEL, bg=COL_BG).pack(anchor='w', pady=(0, 4))

        row = tk.Frame(t2_content, bg=COL_BG)
        row.pack(fill='x', pady=(0, 8))
        self._archivo_obs_var = tk.StringVar()
        e = ttk.Entry(row, textvariable=self._archivo_obs_var, width=50)
        e.pack(side='left', fill='x', expand=True)
        ttk.Button(row, text="📁 Seleccionar", command=self._seleccionar_obs_archivo).pack(side='left', padx=(6, 0))

        btn_row = tk.Frame(t2_content, bg=COL_BG)
        btn_row.pack(fill='x', pady=(0, 12))
        ttk.Button(btn_row, text="✓ Validar Cambios", command=self._validar_obs).pack(side='left', padx=(0, 6))
        self._btn_aplicar_obs = ttk.Button(btn_row, text="⚡ Aplicar Observaciones",
                                          command=self._aplicar_obs, state='disabled')
        self._btn_aplicar_obs.pack(side='left')

        tk.Label(t2_content, text="Resumen:", font=FONT_LABEL, bg=COL_BG).pack(anchor='w', pady=(6, 4))
        self._resultado_t2 = tk.Text(t2_content, height=18, width=90, font=FONT_SMALL, bg=COL_ENTRY_BG, fg=COL_WHITE)
        self._resultado_t2.pack(fill='both', expand=True)

    def _descargar_plantilla(self):
        if not self.conn:
            messagebox.showerror("Error", "Sin conexión a BD")
            return

        filtro = self._filtro_emp_var.get()
        fecha_str = self._fecha_obs_var.get()

        try:
            fecha_obs = datetime.strptime(fecha_str, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Error", "Formato de fecha inválido. Use YYYY-MM-DD")
            return

        def tarea():
            try:
                if filtro == "ACTIVOS":
                    estado_filter = "ESTADO='ACT'"
                elif filtro == "INACTIVOS":
                    estado_filter = "ESTADO<>'ACT'"
                else:
                    estado_filter = "1=1"

                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT EMPLEADO, APELLIDOS, NOMBRES FROM RPEMPLEA WHERE {SQL_FILTER} AND {estado_filter} ORDER BY APELLIDOS")
                    rows = cur.fetchall()

                from openpyxl import Workbook
                from openpyxl.styles import Font
                wb = Workbook()
                ws = wb.active
                ws.title = "OBSERVACIONES"

                headers = ['empleado', 'apellidos', 'nombres', 'fecha_ven', 'texto_obs']
                for ci, h in enumerate(headers, 1):
                    ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

                for ri, r in enumerate(rows, 2):
                    ws.cell(row=ri, column=1, value=r[0])
                    ws.cell(row=ri, column=2, value=r[1])
                    ws.cell(row=ri, column=3, value=r[2])
                    ws.cell(row=ri, column=4, value=fecha_obs)
                    ws.cell(row=ri, column=5, value="")

                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fn = f"PLANTILLA_OBSERVACIONES_{ts}.xlsx"
                wb.save(fn)

                self.window.after(0, lambda: messagebox.showinfo("Éxito", f"Plantilla creada:\n{fn}"))
                self.window.after(0, lambda: self._resultado_t1.insert('end', f"✓ Plantilla descargada: {fn}\n  Empleados: {len(rows)}\n  Fecha: {fecha_str}\n"))

            except ImportError:
                self.window.after(0, lambda: messagebox.showerror("Error", "Requiere openpyxl:\npip install openpyxl"))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))

        threading.Thread(target=tarea, daemon=True).start()

    def _seleccionar_obs_archivo(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
        if f:
            self._archivo_obs_var.set(f)
            self._btn_aplicar_obs.config(state='disabled')

    def _validar_obs(self):
        arch = self._archivo_obs_var.get()
        if not arch:
            messagebox.showwarning("Aviso", "Seleccione un archivo")
            return

        def tarea():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(arch, data_only=True)
                ws = wb['OBSERVACIONES']
                headers = [c.value for c in ws[1]]

                if not all(h in headers for h in ['empleado', 'fecha_ven', 'texto_obs']):
                    self.window.after(0, lambda: messagebox.showerror("Error",
                        "Faltan columnas requeridas: empleado, fecha_ven, texto_obs"))
                    return

                ei = headers.index('empleado')
                efe = headers.index('fecha_ven')
                et = headers.index('texto_obs')

                datos = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[ei] and row[et] is not None and str(row[et]).strip():
                        datos.append(row)

                if not datos:
                    self.window.after(0, lambda: messagebox.showerror("Error", "Sin datos válidos"))
                    return

                cambios = []
                for row in datos:
                    emp = str(row[ei]).strip()
                    fecha = row[efe]
                    texto = str(row[et]).strip()
                    if emp and texto:
                        cambios.append({'empleado': emp, 'fecha_ven': fecha, 'texto_obs': texto})

                self.datos_validados = cambios
                self.window.after(0, lambda: self._mostrar_validacion_obs(cambios))

            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))

        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_validacion_obs(self, cambios):
        self._resultado_t2.delete(1.0, 'end')
        self._resultado_t2.insert('end', f"Observaciones a agregar: {len(cambios)}\n\n")
        for c in cambios[:20]:
            self._resultado_t2.insert('end', f"• {c['empleado']} ({c['fecha_ven']}): {c['texto_obs'][:60]}{'...' if len(c['texto_obs']) > 60 else ''}\n")
        if len(cambios) > 20:
            self._resultado_t2.insert('end', f"\n... y {len(cambios) - 20} más\n")
        self._btn_aplicar_obs.config(state='normal')

    def _aplicar_obs(self):
        if not self.datos_validados:
            return
        if not messagebox.askyesno("Confirmar", f"¿Agregar {len(self.datos_validados)} observaciones?", icon='warning'):
            return

        self._btn_aplicar_obs.config(state='disabled')

        def tarea():
            try:
                from agregar_observaciones_masivas import procesar_carga
                with self.app.conn_lock:
                    stats = procesar_carga(self.datos_validados, self.conn, force_new_row=True)

                self.window.after(0, lambda: self._mostrar_resultado_obs(stats))

            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))

        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_resultado_obs(self, stats):
        self._resultado_t2.delete(1.0, 'end')
        self._resultado_t2.insert('end', f"""
╔═══════════════════════════════════════════════════════╗
║             RESULTADO DE AGREGACIÓN                   ║
╚═══════════════════════════════════════════════════════╝

✓ Insertados:     {stats['insertados']}
⚠ Duplicados:     {stats['duplicados']}
❌ Sin espacio:    {stats['sin_espacio']}
🔴 Errores:       {stats['errores']}

Detalles:
───────────────────────────────────────────────────────
""")
        for d in stats['detalles'][:30]:
            if d['tipo'] == 'INSERTADO':
                self._resultado_t2.insert('end', f"✓ {d['empleado']} → {d['campo']}: {d['texto']}\n")
            elif d['tipo'] == 'DUPLICADO':
                self._resultado_t2.insert('end', f"⚠ {d['empleado']}: Duplicado (ya existe)\n")
            elif d['tipo'] == 'ERROR':
                self._resultado_t2.insert('end', f"❌ {d['empleado']}: {d.get('error', 'Error desconocido')}\n")

        if len(stats['detalles']) > 30:
            self._resultado_t2.insert('end', f"\n... y {len(stats['detalles']) - 30} más\n")

        messagebox.showinfo("Completado",
            f"Insertados: {stats['insertados']}\nDuplicados: {stats['duplicados']}\nErrores: {stats['errores']}")
        self.datos_validados = None


# ═══════════════════════════════════════════════════════════════════
# Editor de Observaciones — buscador con autocompletar + edición directa
# ═══════════════════════════════════════════════════════════════════
MESES_NUM = {'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
             'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12}
MESES_NOMBRES = list(MESES_NUM.keys())


class EditorObservacionesFrame:
    """Buscador de empleado (autocompletar en vivo) + edición de sus
    observaciones (RPEMPOBSERV) para el mes/año seleccionado. Mismo
    criterio mes+año y UPDATE TOP(1) que SISTEMA_GESTION_EMPLEADOS_10.pyw
    (_mostrar_obs/_guardar_obs), para que ambas herramientas vean y
    editen exactamente la misma fila."""

    def __init__(self, window, app):
        self.window = window
        self.app = app
        self.conn = app.conn
        self.empleados = []
        self.emp_actual = None
        self._obs_widgets = []
        self._sugerencia_map = {}
        self._sugerencias_visible = False
        self._build()
        threading.Thread(target=self._cargar_empleados, daemon=True).start()

    def _build(self):
        main = ttk.Frame(self.window)
        main.pack(fill='both', expand=True, padx=16, pady=16)

        tk.Label(main, text="🔍 Editor de Observaciones por Empleado",
                 font=FONT_TITLE, fg=COL_HEADER, bg=COL_BG).pack(pady=(0, 8), anchor='w')

        buscar_frame = ttk.Frame(main)
        buscar_frame.pack(fill='x', pady=(0, 2))
        tk.Label(buscar_frame, text="Buscar empleado:", font=FONT_LABEL, bg=COL_BG).pack(side='left', padx=(0, 8))
        self._search_var = tk.StringVar()
        self._search_entry = ttk.Entry(buscar_frame, textvariable=self._search_var, font=FONT_DEFAULT)
        self._search_entry.pack(side='left', fill='x', expand=True)
        self._search_var.trace_add('write', self._on_search_change)
        self._search_entry.bind('<Down>', self._focus_listbox)
        self._search_entry.bind('<Return>', self._select_first_suggestion)
        self._search_entry.bind('<Escape>', lambda e: self._ocultar_sugerencias())
        tk.Label(buscar_frame, text="por código, cédula, apellido o nombre",
                 font=FONT_SMALL, fg=COL_GRAY, bg=COL_BG).pack(side='left', padx=(8, 0))

        self._sugerencias_frame = tk.Frame(main, bg=COL_CARD, relief='solid', borderwidth=1)
        self._sugerencias_list = tk.Listbox(self._sugerencias_frame, font=FONT_DEFAULT,
                                             bg=COL_ENTRY_BG, fg=COL_TEXT, selectbackground=COL_ACCENT,
                                             selectforeground=COL_HEADER, height=8,
                                             borderwidth=0, highlightthickness=0, activestyle='none')
        self._sugerencias_list.pack(fill='both', expand=True, padx=2, pady=2)
        self._sugerencias_list.bind('<<ListboxSelect>>', self._on_sugerencia_click)
        self._sugerencias_list.bind('<Return>', self._on_sugerencia_click)

        self._lbl_empleado = tk.Label(main, text="Ningún empleado seleccionado", font=FONT_HEAD,
                                       fg=COL_GRAY, bg=COL_BG, anchor='w')
        self._lbl_empleado.pack(fill='x', pady=(10, 6))

        ctrl = ttk.Frame(main)
        ctrl.pack(fill='x', pady=(0, 8))
        tk.Label(ctrl, text='Período:', font=FONT_LABEL, bg=COL_BG).pack(side='left', padx=(0, 6))
        self._mes_var = tk.StringVar(value=MESES_NOMBRES[datetime.now().month - 1])
        ttk.Combobox(ctrl, textvariable=self._mes_var, values=MESES_NOMBRES, width=12,
                     state='readonly').pack(side='left', padx=6)
        self._anio_var = tk.StringVar(value=str(datetime.now().year))
        ttk.Combobox(ctrl, textvariable=self._anio_var, values=[str(y) for y in range(2020, 2031)],
                     width=8, state='readonly').pack(side='left', padx=6)
        self._btn_mostrar = ttk.Button(ctrl, text="Mostrar", command=self._mostrar_obs, state='disabled')
        self._btn_mostrar.pack(side='left', padx=8)
        self._btn_guardar = ttk.Button(ctrl, text="💾 Guardar Obs.", command=self._guardar_obs,
                                        style='Accent.TButton', state='disabled')
        self._btn_guardar.pack(side='left', padx=4)

        cont = ttk.Frame(main)
        cont.pack(fill='both', expand=True)
        self._obs_canvas = tk.Canvas(cont, bg=COL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(cont, orient='vertical', command=self._obs_canvas.yview)
        self._obs_frame = tk.Frame(self._obs_canvas, bg=COL_BG)
        self._obs_canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        self._obs_canvas.configure(yscrollcommand=vsb.set)
        self._obs_canvas.create_window(0, 0, window=self._obs_frame, anchor='nw')
        self._obs_frame.bind('<Configure>',
                              lambda e: self._obs_canvas.configure(scrollregion=self._obs_canvas.bbox('all')))
        self._obs_canvas.bind('<MouseWheel>', lambda e: self._obs_canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units'))

        tk.Label(self._obs_frame, text="Busque y seleccione un empleado para ver sus observaciones",
                 font=FONT_LABEL, bg=COL_BG, fg=COL_GRAY).pack(pady=20)

    def _cargar_empleados(self):
        try:
            with self.app.conn_lock:
                cur = self.conn.cursor()
                cur.execute(f"SELECT EMPLEADO, CEDULA, APELLIDOS, NOMBRES FROM RPEMPLEA "
                            f"WHERE {SQL_FILTER} ORDER BY APELLIDOS")
                self.empleados = cur.fetchall()
        except Exception as e:
            LOG.error("Error cargando empleados para autocompletar: %s", e)
            self.empleados = []

    def _on_search_change(self, *_args):
        texto = self._search_var.get().strip().lower()
        if not texto:
            self._ocultar_sugerencias()
            return
        matches = []
        for emp in self.empleados:
            cod, ced, ape, nom = emp
            hay = f"{cod} {ced} {ape} {nom}".lower()
            if texto in hay:
                matches.append(emp)
                if len(matches) >= 30:
                    break
        if not matches:
            self._ocultar_sugerencias()
            return
        self._sugerencias_list.delete(0, 'end')
        self._sugerencia_map = {}
        for i, (cod, ced, ape, nom) in enumerate(matches):
            self._sugerencias_list.insert('end', f"{cod}  |  {ape} {nom}  |  C.I. {ced}")
            self._sugerencia_map[i] = cod
        self._mostrar_sugerencias()

    def _mostrar_sugerencias(self):
        if not self._sugerencias_visible:
            self._sugerencias_frame.pack(fill='x', pady=(0, 8), before=self._lbl_empleado)
            self._sugerencias_visible = True

    def _ocultar_sugerencias(self):
        if self._sugerencias_visible:
            self._sugerencias_frame.pack_forget()
            self._sugerencias_visible = False

    def _focus_listbox(self, _event=None):
        if self._sugerencias_visible and self._sugerencias_list.size() > 0:
            self._sugerencias_list.focus_set()
            self._sugerencias_list.selection_set(0)

    def _select_first_suggestion(self, _event=None):
        if self._sugerencias_visible and self._sugerencias_list.size() > 0:
            self._sugerencias_list.selection_clear(0, 'end')
            self._sugerencias_list.selection_set(0)
            self._on_sugerencia_click()

    def _on_sugerencia_click(self, _event=None):
        sel = self._sugerencias_list.curselection()
        if not sel:
            return
        cod = self._sugerencia_map.get(sel[0])
        if cod is None:
            return
        emp = next((e for e in self.empleados if str(e[0]) == str(cod)), None)
        if not emp:
            return
        self.emp_actual = emp
        cod, ced, ape, nom = emp
        self._lbl_empleado.config(text=f"👤 {ape} {nom}   |   Código: {cod}   |   C.I. {ced}", fg=COL_ACCENT)
        self._search_var.set("")
        self._ocultar_sugerencias()
        self._btn_mostrar.config(state='normal')
        self._btn_guardar.config(state='disabled')
        for w in self._obs_frame.winfo_children():
            w.destroy()
        tk.Label(self._obs_frame, text="Presione 'Mostrar' para ver las observaciones del período",
                 font=FONT_LABEL, bg=COL_BG, fg=COL_GRAY).pack(pady=20)

    def _mostrar_obs(self):
        if not self.emp_actual:
            return
        emp_cod = str(self.emp_actual[0])
        mes_num = MESES_NUM.get(self._mes_var.get(), datetime.now().month)
        ano_num = int(self._anio_var.get()) if self._anio_var.get().isdigit() else datetime.now().year

        for w in self._obs_frame.winfo_children():
            w.destroy()
        self._obs_widgets = []
        self._btn_guardar.config(state='disabled')
        mes_txt, anio_txt = self._mes_var.get(), self._anio_var.get()

        def tarea():
            try:
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"""
                        SELECT refer1, refer2, refer3, refer4, refer5, refer6, refer7, fecha_ven
                        FROM RPEMPOBSERV
                        WHERE empleado = ? AND MONTH(fecha_ven) = ? AND YEAR(fecha_ven) = ? AND {SQL_FILTER}
                        ORDER BY fecha_ven DESC
                    """, (emp_cod, mes_num, ano_num))
                    row = cur.fetchone()
                self.window.after(0, lambda: self._render_obs(row, mes_txt, anio_txt))
            except Exception as e:
                LOG.error("Error cargando observaciones: %s", e)
                self.window.after(0, lambda msg=str(e): tk.Label(
                    self._obs_frame, text=f"Error: {msg}", font=FONT_LABEL, bg=COL_BG, fg=COL_DANGER).pack(pady=20))
        threading.Thread(target=tarea, daemon=True).start()

    def _render_obs(self, row, mes_txt, anio_txt):
        try:
            if not row:
                tk.Label(self._obs_frame, text=f"Sin observaciones para {mes_txt} {anio_txt}",
                         font=FONT_LABEL, bg=COL_BG, fg=COL_GRAY).pack(pady=20)
                return

            fecha_ven = row[7]
            header = tk.Frame(self._obs_frame, bg=COL_HEADER, height=50)
            header.pack(fill='x', pady=(0, 10))
            tk.Label(header, text=f"📅 {mes_txt} {anio_txt}  |  📆 {fecha_ven.strftime('%d/%m/%Y')}",
                     font=FONT_HEAD, bg=COL_HEADER, fg=COL_WHITE).pack(pady=8)

            # Se muestran los 7 campos SIEMPRE (llenos y vacíos), para que nunca
            # haga falta escribir una observación nueva encima de una ya existente.
            campos_llenos = 0
            for i in range(7):
                lleno = bool(row[i])
                if lleno:
                    campos_llenos += 1
                card = tk.Frame(self._obs_frame, bg=COL_CARD, relief='solid', borderwidth=2)
                card.pack(fill='both', expand=False, pady=(0, 8))
                card_header = tk.Frame(card, bg=COL_ACCENT if lleno else COL_GRAY, height=28)
                card_header.pack(fill='x')
                titulo = f"[CAMPO {i + 1}]" if lleno else f"[CAMPO {i + 1} — VACÍO, disponible para nueva observación]"
                tk.Label(card_header, text=titulo, font=FONT_LABEL,
                         bg=COL_ACCENT if lleno else COL_GRAY,
                         fg=COL_WHITE if lleno else COL_HEADER).pack(pady=4)
                content = tk.Text(card, font=FONT_DEFAULT, bg=COL_CARD, fg=COL_TEXT,
                                   height=3, wrap='word', relief='flat', borderwidth=0)
                content.pack(fill='both', expand=True, padx=10, pady=10)
                if lleno:
                    content.insert('1.0', row[i])
                self._obs_widgets.append({'campo': i + 1, 'widget': content, 'valor_original': row[i] or ''})

            if campos_llenos >= 7:
                nota = "Los 7 campos están llenos. Para agregar más use la pestaña 'Agregar Observaciones Masivas' (crea una fila adicional)."
            else:
                nota = f"Total: {campos_llenos}/7 campos llenos"
            tk.Label(self._obs_frame, text=nota, font=FONT_SMALL, bg=COL_BG, fg=COL_GRAY,
                     wraplength=600, justify='left').pack(pady=8)
            self._btn_guardar.config(state='normal')
            _bind_scroll_rueda(self._obs_frame, self._obs_canvas)
        except Exception as e:
            tk.Label(self._obs_frame, text=f"Error: {e}", font=FONT_LABEL, bg=COL_BG, fg=COL_DANGER).pack(pady=20)

    def _guardar_obs(self):
        if not self.emp_actual or not self._obs_widgets:
            return
        emp_cod = str(self.emp_actual[0])
        mes_num = MESES_NUM.get(self._mes_var.get(), datetime.now().month)
        ano_num = int(self._anio_var.get()) if self._anio_var.get().isdigit() else datetime.now().year

        # Leer el contenido de los widgets en el hilo principal (Tkinter no es thread-safe)
        cambios_pendientes = []
        for widget_info in self._obs_widgets:
            texto_nuevo = widget_info['widget'].get('1.0', 'end').strip()
            if texto_nuevo != widget_info['valor_original']:
                cambios_pendientes.append((widget_info['campo'], texto_nuevo))

        if not cambios_pendientes:
            messagebox.showinfo("Sin cambios", "No hay cambios para guardar")
            return

        def tarea():
            try:
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"""
                        SELECT fecha_ven FROM RPEMPOBSERV
                        WHERE empleado = ? AND MONTH(fecha_ven) = ? AND YEAR(fecha_ven) = ? AND {SQL_FILTER}
                        ORDER BY fecha_ven DESC
                    """, (emp_cod, mes_num, ano_num))
                    row = cur.fetchone()
                    if not row:
                        self.window.after(0, lambda: messagebox.showerror("Error", "No se encontró la observación"))
                        return
                    fecha_ven = row[0]

                    cambios = 0
                    for campo_num, texto_nuevo in cambios_pendientes:
                        columna = f"refer{campo_num}"
                        cur.execute(f"""
                            UPDATE TOP (1) RPEMPOBSERV SET {columna} = ?
                            WHERE empleado = ? AND fecha_ven = ? AND {SQL_FILTER}
                        """, (texto_nuevo, emp_cod, fecha_ven))
                        cambios += 1
                    self.conn.commit()
                self.window.after(0, lambda: self._guardar_obs_ok(cambios))
            except Exception as e:
                LOG.error("Error guardando observaciones: %s", e)
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", f"No se pudo guardar: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()

    def _guardar_obs_ok(self, cambios):
        messagebox.showinfo("✅ Guardado", f"Se guardaron {cambios} cambios en observaciones")
        self._mostrar_obs()


# ═══════════════════════════════════════════════════════════════════
# Exportador de Catálogos (DBTABLAS) — adaptado de SACAR_TABLAS_SEC.pyw
# ═══════════════════════════════════════════════════════════════════
class ExportadorCatalogosFrame:
    """Exporta tipos de DBTABLAS (CODEMP='10') a Excel, una hoja por tipo
    más una hoja 'TODOS'. Reutiliza la conexión ya abierta de la app
    (app.conn) en vez de crear una nueva — el script original tenía un
    connection string roto (SERVER=SERVER\\server, placeholder sin
    reemplazar)."""

    def __init__(self, window, app):
        self.window = window
        self.app = app
        self.conn = app.conn
        self.check_vars = {}
        self._build()
        self._cargar_tipos()

    def _build(self):
        frame = ttk.Frame(self.window)
        frame.pack(fill='both', expand=True, padx=16, pady=16)

        tk.Label(frame, text="📤 Exportar Catálogos", font=('Segoe UI', 14, 'bold'),
                 fg=COL_HEADER, bg=COL_BG).pack(anchor='w')
        tk.Label(frame, text="Tipos de DBTABLAS filtrados por CODEMP='10'",
                 font=FONT_SMALL, fg=COL_GRAY, bg=COL_BG).pack(anchor='w', pady=(0, 10))

        tipos_frame = ttk.LabelFrame(frame, text="Seleccione los TIPOS a exportar", padding=8)
        tipos_frame.pack(fill='both', expand=True, pady=(0, 10))

        self._catalogos_canvas = canvas = tk.Canvas(tipos_frame, bg=COL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(tipos_frame, orient='vertical', command=canvas.yview)
        self._frame_checks = tk.Frame(canvas, bg=COL_BG)
        self._frame_checks.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=self._frame_checks, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        _bind_scroll_rueda(canvas, canvas)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=(0, 8))
        ttk.Button(btn_frame, text="Solo SEC y DPT", command=self._solo_sec_dpt).pack(side='left', padx=(0, 4))
        ttk.Button(btn_frame, text="Todos", command=self._seleccionar_todos).pack(side='left', padx=4)
        ttk.Button(btn_frame, text="Ninguno", command=self._deseleccionar_todos).pack(side='left', padx=4)

        ttk.Button(frame, text="⬇ EXPORTAR A EXCEL", command=self._exportar,
                   style='Accent.TButton').pack(fill='x', pady=(0, 8))

        self.status = tk.StringVar(value="Cargando tipos...")
        tk.Label(frame, textvariable=self.status, font=FONT_SMALL,
                 fg=COL_GRAY, bg=COL_BG, anchor='w').pack(fill='x')

    def _cargar_tipos(self):
        def tarea():
            try:
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute("SELECT DISTINCT TIPO FROM DBTABLAS WHERE CODEMP = '10' ORDER BY TIPO")
                    tipos = [r[0] for r in cur.fetchall()]
                self.window.after(0, lambda: self._mostrar_tipos(tipos))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): self.status.set(f"Error: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_tipos(self, tipos):
        tipos_defecto = ('SEC', 'DPT')
        for i, tipo in enumerate(tipos):
            var = tk.BooleanVar(value=(tipo in tipos_defecto))
            self.check_vars[tipo] = var
            cb = tk.Checkbutton(self._frame_checks, text=tipo, font=FONT_DEFAULT, variable=var,
                                 bg=COL_BG, fg=COL_TEXT, selectcolor=COL_ACCENT,
                                 activebackground=COL_BG, activeforeground=COL_TEXT)
            cb.grid(row=i // 3, column=i % 3, sticky='w', padx=10, pady=2)
        self.status.set(f"Tipos cargados: {len(tipos)}")
        _bind_scroll_rueda(self._frame_checks, self._catalogos_canvas)

    def _solo_sec_dpt(self):
        for tipo, var in self.check_vars.items():
            var.set(tipo in ('SEC', 'DPT'))

    def _seleccionar_todos(self):
        for var in self.check_vars.values():
            var.set(True)

    def _deseleccionar_todos(self):
        for var in self.check_vars.values():
            var.set(False)

    def _exportar(self):
        tipos_sel = [t for t, v in self.check_vars.items() if v.get()]
        if not tipos_sel:
            messagebox.showwarning("Aviso", "Seleccione al menos un TIPO")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        tipos_txt = "_".join(tipos_sel[:3]) + ("_etc" if len(tipos_sel) > 3 else "")
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"DBTABLAS_{tipos_txt}_{ts}.xlsx")
        if not archivo:
            return

        self.status.set("Exportando...")

        def tarea():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                tipos_str = ",".join(f"'{t}'" for t in tipos_sel)
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT TIPO, CODIGO, NOMBRE, FACTOR, T_C, T_P FROM DBTABLAS "
                                f"WHERE CODEMP='10' AND TIPO IN ({tipos_str}) ORDER BY TIPO, CODIGO")
                    cols = [c[0] for c in cur.description]
                    rows = cur.fetchall()

                if not rows:
                    self.window.after(0, lambda: messagebox.showwarning("Aviso", "No se encontraron datos"))
                    self.window.after(0, lambda: self.status.set("Sin datos"))
                    return

                wb = Workbook()
                wb.remove(wb.active)
                por_tipo = {}
                for r in rows:
                    por_tipo.setdefault(r[0], []).append(r)

                def escribir_hoja(ws, filas):
                    for ci, h in enumerate(cols, 1):
                        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
                    for ri, r in enumerate(filas, 2):
                        for ci, v in enumerate(r, 1):
                            ws.cell(row=ri, column=ci, value=v)

                for tipo in tipos_sel:
                    filas_tipo = por_tipo.get(tipo)
                    if filas_tipo:
                        escribir_hoja(wb.create_sheet(tipo[:31]), filas_tipo)
                escribir_hoja(wb.create_sheet('TODOS'), rows)
                wb.save(archivo)

                n = len(rows)
                self.window.after(0, lambda: self.status.set(f"Exportado: {n} registros"))
                self.window.after(0, lambda: messagebox.showinfo(
                    "Éxito", f"Archivo creado:\n{archivo}\n\nRegistros: {n}\nTipos: {', '.join(tipos_sel)}"))
                self.window.after(0, lambda: os.startfile(os.path.dirname(archivo) or '.'))
            except ImportError:
                self.window.after(0, lambda: messagebox.showerror("Error", "Requiere openpyxl:\npip install openpyxl"))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", f"Error al exportar: {msg}"))
                self.window.after(0, lambda msg=str(e): self.status.set(f"Error: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════
# Aplicación principal — solo carga masiva
# ═══════════════════════════════════════════════════════════════════
class CargaMasivaApp:
    def __init__(self, root):
        self.root = root
        self.conn = None
        self.conn_lock = threading.Lock()  # pyodbc no soporta un mismo Connection desde varios hilos a la vez
        self._configurar_estilo()
        self._crear_interfaz()
        self._conectar_bd()

    def _configurar_estilo(self):
        self.root.title("Carga Masiva de Empleados — INSEVIG")
        self.root.geometry("1150x800")
        self.root.configure(bg=COL_BG)
        self.root.option_add('*Background', COL_BG)
        self.root.option_add('*Foreground', COL_TEXT)
        self.root.option_add('*Font', FONT_DEFAULT)

        style = ttk.Style()
        try:
            style.theme_use('clam')
        except Exception:
            pass
        style.configure('TFrame', background=COL_BG)
        style.configure('TLabelframe', background=COL_BG, foreground=COL_TEXT)
        style.configure('TLabelframe.Label', background=COL_BG, foreground=COL_ACCENT, font=FONT_LABEL)
        style.configure('TNotebook', background=COL_BG, borderwidth=0)
        style.configure('TNotebook.Tab', background=COL_CARD, foreground=COL_TEXT, padding=(14, 6), font=FONT_LABEL)
        style.map('TNotebook.Tab', background=[('selected', COL_ACCENT)], foreground=[('selected', COL_HEADER)])
        style.configure('TButton', background=COL_CARD, foreground=COL_TEXT, font=FONT_DEFAULT)
        style.configure('Accent.TButton', background=COL_ACCENT, foreground=COL_HEADER, font=FONT_LABEL)
        style.configure('TEntry', fieldbackground=COL_ENTRY_BG, foreground=COL_TEXT)
        style.configure('TCombobox', fieldbackground=COL_ENTRY_BG, foreground=COL_TEXT)

    def _crear_interfaz(self):
        header = tk.Frame(self.root, bg=COL_HEADER, height=54)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="📊 CARGA MASIVA DE EMPLEADOS — INSEVIG",
                 font=FONT_TITLE, fg=COL_WHITE, bg=COL_HEADER).pack(side='left', padx=16)
        self._status_var = tk.StringVar(value="Conectando…")
        tk.Label(header, textvariable=self._status_var, font=FONT_SMALL,
                 fg=COL_ACCENT, bg=COL_HEADER).pack(side='right', padx=16)

        self._contenedor = tk.Frame(self.root, bg=COL_BG)
        self._contenedor.pack(fill='both', expand=True, padx=10, pady=10)
        tk.Label(self._contenedor, text="Conectando a la base de datos…",
                 font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT).pack(pady=40)

    def _status(self, msg):
        self._status_var.set(msg)

    def _conectar_bd(self):
        def tarea():
            try:
                conn = _get_sql_conn()
                self.root.after(0, lambda: self._on_conectado(conn))
            except Exception as e:
                self.root.after(0, lambda msg=str(e): self._on_error_conexion(msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _on_conectado(self, conn):
        self.conn = conn
        self._status("✓ Conectado")
        for w in self._contenedor.winfo_children():
            w.destroy()
        nb = ttk.Notebook(self._contenedor)
        nb.pack(fill='both', expand=True)

        t1 = ttk.Frame(nb)
        nb.add(t1, text="Edición Masiva (Excel)")
        EdicionMasivaFrame(t1, self)

        t2 = ttk.Frame(nb)
        nb.add(t2, text="Agregar Observaciones Masivas")
        ObservacionesMasivasFrame(t2, self)

        t3 = ttk.Frame(nb)
        nb.add(t3, text="Editor de Observaciones")
        EditorObservacionesFrame(t3, self)

        t4 = ttk.Frame(nb)
        nb.add(t4, text="Exportar Catálogos")
        ExportadorCatalogosFrame(t4, self)

    def _on_error_conexion(self, msg):
        LOG.error("Fallo de conexion: %s", msg)
        self._status("✗ Sin conexión")
        for w in self._contenedor.winfo_children():
            w.destroy()
        tk.Label(self._contenedor,
                 text="No se pudo conectar a la base de datos.\n"
                      "Verifique la red o contacte al administrador.",
                 font=FONT_LABEL, bg=COL_BG, fg=COL_DANGER, justify='center').pack(pady=40)
        ttk.Button(self._contenedor, text="Reintentar", command=self._reintentar,
                   style='Accent.TButton').pack()

    def _reintentar(self):
        for w in self._contenedor.winfo_children():
            w.destroy()
        tk.Label(self._contenedor, text="Conectando a la base de datos…",
                 font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT).pack(pady=40)
        self._conectar_bd()

    def _on_close(self):
        if self.conn:
            try:
                self.conn.close()
            except Exception:
                pass
        self.root.destroy()

    def _cargar_lista(self):
        """No hay grilla de empleados en esta herramienta; solo confirma en el status bar."""
        self._status("✓ Cambios aplicados")


# ═══════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    try:
        base_dir = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(os.path.abspath(__file__))
        ico = os.path.join(base_dir, 'carga_masiva.ico')
        if os.path.exists(ico):
            root.iconbitmap(default=ico)
    except Exception:
        pass
    app = CargaMasivaApp(root)
    root.protocol("WM_DELETE_WINDOW", app._on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
