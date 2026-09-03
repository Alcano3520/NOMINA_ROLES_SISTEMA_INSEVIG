#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SISTEMA GESTION EMPLEADOS v10.0 - INSEVIG
GUI Mejorada siguiendo estándares RRHH documentados.
Compatible Linux/Windows.
"""

import os, sys, threading, webbrowser, shutil, tempfile, calendar, logging, html
from pathlib import Path
from datetime import datetime

# Setup logging (archivo + terminal)
import sys
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('app_debug.log'),
        logging.StreamHandler(sys.stderr)  # También a terminal
    ]
)
LOG = logging.getLogger(__name__)

_openssl = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'openssl_legacy.cnf')
if os.path.exists(_openssl):
    os.environ['OPENSSL_CONF'] = _openssl

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import pyodbc

try:
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('insevig.empleados.10.0')
except Exception:
    pass

# ── Palette INSEVIG DARK MODE ──────────────────────────────────────
COL_BG       = '#1E1E1E'      # Fondo principal oscuro
COL_HEADER   = '#0D1B2A'      # Header muy oscuro
COL_ACCENT   = '#4A9EFF'      # Azul claro para acentos
COL_PEND     = '#FF9F43'      # Naranja claro
COL_OK       = '#2ED573'      # Verde claro
COL_DANGER   = '#FF6B6B'      # Rojo claro
COL_WHITE    = '#FFFFFF'      # Blanco para texto
COL_GRAY     = '#A0A0A0'      # Gris claro
COL_ENTRY_BG = '#2D2D2D'      # Fondo entrada oscuro
COL_CARD     = '#2D2D2D'      # Tarjetas oscuras
COL_TEXT     = '#E0E0E0'      # Texto claro
COL_BORDER   = '#404040'      # Borde oscuro

FONT_DEFAULT = ('Segoe UI', 13)
FONT_SMALL   = ('Segoe UI', 10)
FONT_LABEL   = ('Segoe UI', 11, 'bold')
FONT_HEAD    = ('Segoe UI', 12, 'bold')
FONT_TITLE   = ('Segoe UI', 15, 'bold')

# ── SQL Server ──────────────────────────────────────────────────────
SQL_CFG = {
    'driver':   'ODBC Driver 17 for SQL Server',
    'server':   '192.168.2.115',
    'database': 'insevig',
    'uid':      'sa',
    'pwd':      'puntosoft123*',
}
SQL_FILTER = "CODEMP='10' AND CODSUC='10'"


def _get_sql_conn():
    """Intenta drivers en orden de prioridad."""
    drivers = [
        'ODBC Driver 17 for SQL Server',
        'ODBC Driver 18 for SQL Server',
        'ODBC Driver 13 for SQL Server',
        'ODBC Driver 11 for SQL Server',
        'SQL Server',
    ]
    cfg = SQL_CFG.copy()
    for drv in drivers:
        try:
            cfg['driver'] = drv
            cs = (f"DRIVER={{{drv}}};SERVER={cfg['server']};DATABASE={cfg['database']};"
                  f"UID={cfg['uid']};PWD={cfg['pwd']};Encrypt=No;TrustServerCertificate=yes")
            return pyodbc.connect(cs, timeout=10)
        except Exception:
            continue
    raise RuntimeError("No se pudo conectar a SQL Server con ningún driver ODBC.")


# ── ToolTip ──────────────────────────────────────────────────────────
class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind('<Enter>', self._show)
        widget.bind('<Leave>', self._hide)

    def _show(self, ev=None):
        if self.tip_window:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + 25
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        lbl = tk.Label(tw, text=self.text, font=FONT_SMALL, bg=COL_CARD,
                       fg=COL_ACCENT, wraplength=300, justify='left',
                       padx=8, pady=4, relief='solid', borderwidth=1)
        lbl.pack()

    def _hide(self, ev=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


def _bind_scroll_rueda(widget, canvas):
    """Une `widget` (y todos sus descendientes) al scroll de `canvas` con
    la rueda del mouse. Bindear solo el canvas no basta: los widgets hijos
    (Entry, Label, Frame, etc.) reciben el evento de rueda antes que el
    canvas y lo consumen, así que hay que bindear recursivamente todo el
    árbol de contenido desplazable."""
    widget.bind('<MouseWheel>', lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units'), add='+')
    for child in widget.winfo_children():
        _bind_scroll_rueda(child, canvas)


class SistemaGestionEmpleados10:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gestión de Empleados - INSEVIG v10")
        self.root.geometry("1280x800")
        try:
            self.root.state('zoomed')
        except tk.TclError:
            self.root.attributes('-zoomed', True)
        self.root.configure(bg=COL_BG)

        self._running = True
        self._status_var = tk.StringVar(value="Iniciando...")
        self._tabs_dirty = set()
        self._debounce_id = None

        self.conn = None
        self.empleado_actual = None
        self.datos_originales = None
        self.datos_modificados = False
        self.modo_edicion = False

        self._catalogos_cargados = False
        self.cargos = {}
        self.secciones = {}
        self.departamentos = {}
        self.sexos = {'1': 'MASCULINO', '2': 'FEMENINO'}  # sin catálogo propio en DBTABLAS (TIPO='SEX' no existe); valores fijos
        # Sin catálogo propio en DBTABLAS (TIPO='ECS' no existe; se buscó por
        # nombre "CIVIL"/"SOLTERO"/"CASADO" en las 47 TIPO de DBTABLAS y no
        # aparece en ninguna). NO sigue el orden "estándar" del registro civil:
        # se asumió 1=Soltero/2=Casado por convención y estaba AL REVÉS —
        # confirmado con datos reales: empleado 5220 tiene ESTADO_CI='1' y su
        # campo CONYUGUE literalmente dice "CASADO-GISSELA...", y en general
        # ESTADO_CI='1' correlaciona 84.8% con tener CONYUGUE lleno vs. solo
        # 36.8% para ESTADO_CI='2' (el código mayoritario, 91% de activos —
        # coincide con plantilla joven de guardias, mayoría soltera). Solo 1 y
        # 2 están confirmados así; 3/4/5 siguen siendo suposición (muy pocos
        # empleados: 14, 1 y 10 respectivamente) — no tratarlos como certeza.
        self.estados_civiles = {'1': 'CASADO', '2': 'SOLTERO', '3': 'DIVORCIADO', '4': 'VIUDO', '5': 'UNION LIBRE'}
        # Sin catálogo propio en DBTABLAS (TIPO='TTR' no existe). No es "tipo
        # de contrato" (Fijo/Temporal/Contrato) como se asumía antes: en
        # SP_RP_IMPRIME_ROL el parámetro que filtra por esta columna se llama
        # literalmente @TIPO_EMPLEADO, y en SP_RP_PREPARA_ROL hay un comentario
        # "FONDO DE RESERVA ESLIVE OBREROS Y SEMANAL" junto a
        # "if @mtipo_tra = '3' and @mforma_liq = '1'" — confirma 3=OBRERO. El
        # 100% de los empleados activos usa TIPO_TRA='1' (=EMPLEADO). El
        # código '2' solo aparece en 407 empleados LIQUIDADOS, ninguno activo
        # — sin confirmar qué significa (posible código retirado), no
        # inventarle una etiqueta.
        self.tipos_trabajo = {'1': 'EMPLEADO', '2': 'CÓDIGO 2 (sin confirmar)', '3': 'OBRERO'}
        # RPEMPLEA.TIPO_PGO es de hecho el PERÍODO DE PAGO (semanal/quincenal/
        # mensual), no "forma de pago" — el catálogo DBTABLAS TIPO='FPF' que
        # se usaba antes aquí es para forma de pago de facturación (cheque,
        # transferencia...), con códigos de 2 dígitos que nunca calzan con el
        # dígito único que usa TIPO_PGO (100% de los 2452 empleados activos,
        # en las 3 sucursales, tiene TIPO_PGO='3'). Sin catálogo propio en
        # DBTABLAS para el período de pago; convención estándar ecuatoriana.
        self.periodos_pago = {'1': 'SEMANAL', '2': 'QUINCENAL', '3': 'MENSUAL'}
        self.bancos = {}
        self._auditoria_data = {}
        self._combos_widgets = {}
        self._form_widgets = {}
        self._check_widgets = {}
        self._readonly_descs = set()
        self._lista_rows_cache = []
        self._combo_full_values = {}  # lista completa "codigo - descripcion" por combo buscable, para filtrar en vivo
        self.conn_lock = threading.Lock()  # pyodbc no soporta un mismo Connection desde varios hilos a la vez

        self._configurar_estilo()
        self._build_layout()
        self._set_form_state('view')
        self._status("Conectando a BD...")
        self.root.after(100, self._conectar_bd)

        self.root.protocol('WM_DELETE_WINDOW', self._on_close)

    # ── Estilo ──────────────────────────────────────────────────────
    def _configurar_estilo(self):
        s = ttk.Style()
        s.theme_use('alt')  # Tema alt soporta mejor dark mode
        s.configure('.', font=FONT_DEFAULT, background=COL_BG, foreground=COL_TEXT)

        # Configurar tk widgets (no ttk) - GLOBAL
        self.root.option_add('*background', COL_BG)
        self.root.option_add('*foreground', COL_TEXT)
        self.root.option_add('*Entry.background', COL_ENTRY_BG)
        self.root.option_add('*Entry.foreground', COL_TEXT)
        self.root.option_add('*Text.background', COL_ENTRY_BG)
        self.root.option_add('*Text.foreground', COL_TEXT)
        self.root.option_add('*Frame.background', COL_BG)
        self.root.option_add('*Label.background', COL_BG)
        self.root.option_add('*Label.foreground', COL_TEXT)
        self.root.option_add('*Button.background', COL_CARD)
        self.root.option_add('*Button.foreground', COL_TEXT)

        s.configure('Treeview', background=COL_CARD, fieldbackground=COL_CARD,
                    foreground=COL_TEXT, rowheight=30, font=FONT_DEFAULT)
        s.configure('Treeview.Heading', background=COL_HEADER, foreground=COL_WHITE,
                    font=FONT_HEAD)
        s.map('Treeview', background=[('selected', COL_ACCENT)])

        s.configure('TNotebook', background=COL_BG, borderwidth=1)
        s.configure('TNotebook.Tab', font=FONT_HEAD,
                    padding=[18, 8], background=COL_CARD, foreground=COL_TEXT)
        s.map('TNotebook.Tab', background=[('selected', COL_ACCENT)],
              foreground=[('selected', COL_WHITE)])

        s.configure('TButton', font=FONT_LABEL, padding=[14, 6], background=COL_CARD, foreground=COL_TEXT)
        s.map('TButton', background=[('active', '#404040')])

        s.configure('Accent.TButton', font=FONT_LABEL, padding=[16, 8],
                    background=COL_ACCENT, foreground=COL_WHITE)
        s.map('Accent.TButton', background=[('active', '#357ABD')])

        s.configure('TLabelframe', background=COL_BG, borderwidth=2, relief='solid')
        s.configure('TLabelframe.Label', font=FONT_LABEL,
                    foreground=COL_ACCENT, background=COL_BG)

        s.configure('TEntry', fieldbackground=COL_ENTRY_BG, foreground=COL_WHITE, font=FONT_DEFAULT)
        s.map('TEntry', fieldbackground=[('disabled', COL_ENTRY_BG), ('readonly', COL_ENTRY_BG)],
              foreground=[('disabled', COL_TEXT), ('readonly', COL_TEXT)])
        s.configure('TCombobox', fieldbackground=COL_ENTRY_BG, foreground=COL_WHITE, font=FONT_DEFAULT)
        s.map('TCombobox', fieldbackground=[('readonly', COL_ENTRY_BG), ('disabled', COL_ENTRY_BG)],
              foreground=[('readonly', COL_WHITE), ('disabled', COL_TEXT)])

    # ── Status ──────────────────────────────────────────────────────
    def _status(self, msg):
        self._status_var.set(msg)
        self.root.update_idletasks()

    # ── Layout ──────────────────────────────────────────────────────
    def _build_layout(self):
        # Barra superior
        top = tk.Frame(self.root, bg=COL_HEADER, height=54)
        top.pack(fill='x')
        top.pack_propagate(False)
        tk.Label(top, text="SISTEMA DE GESTIÓN DE EMPLEADOS — INSEVIG",
                 font=FONT_TITLE, fg=COL_WHITE, bg=COL_HEADER).pack(side='left', padx=20, pady=12)

        # Cuerpo
        body = tk.Frame(self.root, bg=COL_BG)
        body.pack(fill='both', expand=True)

        # Izquierda (320px fijo)
        left = tk.Frame(body, bg=COL_BG, width=430)
        left.pack(side='left', fill='y', padx=(8, 0), pady=8)
        left.pack_propagate(False)

        # Separador
        ttk.Separator(body, orient='vertical').pack(side='left', fill='y', padx=4)

        # Derecha (expansible)
        right = tk.Frame(body, bg=COL_BG)
        right.pack(side='left', fill='both', expand=True, padx=8, pady=8)

        self._build_left_panel(left)
        self._build_notebook(right)

        # Barra de botones inferior
        btn_bar = tk.Frame(self.root, bg=COL_CARD, height=44)
        btn_bar.pack(fill='x')
        btn_bar.pack_propagate(False)

        ttk.Button(btn_bar, text="💾 GUARDAR", command=self._guardar_cambios,
                   style='Accent.TButton').pack(side='left', padx=(10, 4), pady=5)
        ttk.Button(btn_bar, text="✖ CANCELAR", command=self._cancelar_cambios
                   ).pack(side='left', padx=4, pady=5)
        ttk.Button(btn_bar, text="📄 IMPRIMIR", command=self._imprimir_empleado
                   ).pack(side='left', padx=4, pady=5)
        ttk.Separator(btn_bar, orient='vertical').pack(side='left', fill='y', padx=10)
        self._lbl_empleado_actual = tk.Label(btn_bar, text="", font=FONT_LABEL,
                                              fg=COL_TEXT, bg=COL_CARD, anchor='w')
        self._lbl_empleado_actual.pack(side='left', fill='x', expand=True, padx=6)
        ttk.Button(btn_bar, text="SALIR", command=self._on_close).pack(side='right', padx=10, pady=5)

        # Barra de estado inferior
        bar = tk.Frame(self.root, bg=COL_HEADER, height=36)
        bar.pack(fill='x')
        bar.pack_propagate(False)
        tk.Label(bar, textvariable=self._status_var,
                 font=FONT_LABEL, fg=COL_WHITE, bg=COL_HEADER,
                 anchor='w').pack(side='left', padx=12, pady=5)

    # ── Panel izquierdo ─────────────────────────────────────────────
    def _build_left_panel(self, parent):
        # Búsqueda
        g = ttk.LabelFrame(parent, text="BÚSQUEDA", padding=10)
        g.pack(fill='x', pady=(0, 8))

        row = tk.Frame(g, bg=COL_BG)
        row.pack(fill='x', pady=3)
        tk.Label(row, text="Cédula:", font=FONT_LABEL,
                 bg=COL_BG).pack(side='left')
        self._cedula_var = tk.StringVar()
        e = ttk.Entry(row, textvariable=self._cedula_var, width=12, font=FONT_DEFAULT)
        e.pack(side='left', padx=(6, 0))
        e.bind('<Return>', lambda ev: self._buscar_por_cedula())
        ttk.Button(row, text="Buscar", command=self._buscar_por_cedula,
                   style='Accent.TButton').pack(side='left', padx=(6, 0))

        row2 = tk.Frame(g, bg=COL_BG)
        row2.pack(fill='x', pady=3)
        tk.Label(row2, text="Código:", font=FONT_LABEL,
                 bg=COL_BG).pack(side='left')
        self._codigo_var = tk.StringVar()
        e2 = ttk.Entry(row2, textvariable=self._codigo_var, width=12, font=FONT_DEFAULT)
        e2.pack(side='left', padx=(6, 0))
        e2.bind('<Return>', lambda ev: self._buscar_por_codigo())
        ttk.Button(row2, text="Buscar", command=self._buscar_por_codigo,
                   style='Accent.TButton').pack(side='left', padx=(6, 0))

        row3 = tk.Frame(g, bg=COL_BG)
        row3.pack(fill='x', pady=(8, 0))
        tk.Label(row3, text="Autocompletar:", font=FONT_LABEL, bg=COL_BG).pack(anchor='w')
        self._autocomplete_var = tk.StringVar()
        self._autocomplete_entry = ttk.Entry(g, textvariable=self._autocomplete_var, font=FONT_DEFAULT)
        self._autocomplete_entry.pack(fill='x', pady=(2, 0))
        self._autocomplete_var.trace_add('write', self._debounce_autocomplete_change)
        self._autocomplete_entry.bind('<Down>', self._autocomplete_focus_listbox)
        self._autocomplete_entry.bind('<Return>', self._autocomplete_select_first)
        self._autocomplete_entry.bind('<Escape>', lambda ev: self._autocomplete_ocultar())
        ToolTip(self._autocomplete_entry, "Escriba código, cédula, apellido o nombre y elija de la lista")

        self._autocomplete_sug_frame = tk.Frame(g, bg=COL_CARD, relief='solid', borderwidth=1)
        self._autocomplete_sug_list = tk.Listbox(self._autocomplete_sug_frame, font=FONT_SMALL,
                                                  bg=COL_ENTRY_BG, fg=COL_TEXT, selectbackground=COL_ACCENT,
                                                  selectforeground=COL_HEADER, height=6,
                                                  borderwidth=0, highlightthickness=0, activestyle='none')
        self._autocomplete_sug_list.pack(fill='both', expand=True, padx=2, pady=2)
        self._autocomplete_sug_list.bind('<<ListboxSelect>>', self._autocomplete_on_click)
        self._autocomplete_sug_list.bind('<Return>', self._autocomplete_on_click)
        self._autocomplete_sug_visible = False
        self._autocomplete_map = {}
        self._autocomplete_empleados = []

        ttk.Button(g, text="🔍 Búsqueda Avanzada", command=self._abrir_buscador,
                   style='Accent.TButton').pack(fill='x', pady=(6, 0))

        # Botones de acción: agrupados en un menú desplegable compacto
        # (antes eran 5 botones apilados que empujaban la lista de
        # EMPLEADOS fuera de la pantalla visible).
        act = tk.Frame(parent, bg=COL_BG)
        act.pack(fill='x', pady=(8, 6))

        self._acciones_menu = tk.Menu(
            parent, tearoff=0, bg=COL_CARD, fg=COL_TEXT,
            activebackground=COL_ACCENT, activeforeground=COL_HEADER,
            font=FONT_DEFAULT, borderwidth=0, relief='flat')
        self._acciones_menu.add_command(label="🆕  Nuevo", command=self._nuevo_empleado)
        self._acciones_menu.add_command(label="✏️  Modificar", command=self._modificar_empleado)
        self._acciones_menu.add_command(label="🗑️  Eliminar", command=self._eliminar_empleado,
                                         foreground=COL_DANGER, activeforeground=COL_DANGER)
        self._acciones_menu.add_separator()
        self._acciones_menu.add_command(label="📋  Vista Completa", command=self._abrir_vista_completa)
        self._acciones_menu.add_command(label="📤  Exportar Catálogos", command=self._abrir_exportador_catalogos)

        def _mostrar_menu_acciones():
            x = btn_acciones.winfo_rootx()
            y = btn_acciones.winfo_rooty() + btn_acciones.winfo_height()
            self._acciones_menu.post(x, y)

        btn_acciones = ttk.Button(act, text="⚙  Acciones  ▾", command=_mostrar_menu_acciones,
                                   style='Accent.TButton')
        btn_acciones.pack(fill='x', pady=2)

        lf = ttk.LabelFrame(parent, text="EMPLEADOS", padding=6)
        lf.pack(fill='both', expand=True)

        buscar_row = tk.Frame(lf, bg=COL_BG)
        buscar_row.pack(fill='x', pady=(0, 6))
        tk.Label(buscar_row, text="🔍", font=FONT_LABEL, bg=COL_BG).pack(side='left')
        self._filtro_texto_var = tk.StringVar()
        e_filtro = ttk.Entry(buscar_row, textvariable=self._filtro_texto_var, font=FONT_DEFAULT)
        e_filtro.pack(side='left', fill='x', expand=True, padx=(4, 0))
        self._filtro_texto_var.trace_add('write', lambda *a: self._debounce_renderizar_lista())
        ToolTip(e_filtro, "Filtra la lista en vivo por código, cédula, apellido o nombre")

        ctrl = tk.Frame(lf, bg=COL_BG)
        ctrl.pack(fill='x', pady=(0, 4))
        tk.Label(ctrl, text="Mostrar:", font=FONT_LABEL,
                 bg=COL_BG).pack(side='left')
        self._filtro_var = tk.StringVar(value="ACTIVOS")
        cb = ttk.Combobox(ctrl, textvariable=self._filtro_var,
                          values=["ACTIVOS", "INACTIVOS", "TODOS"], width=10, state='readonly')
        cb.pack(side='left', padx=(6, 0))
        cb.bind('<<ComboboxSelected>>', lambda ev: self._cargar_lista())

        nav = tk.Frame(ctrl, bg=COL_BG)
        nav.pack(side='right')
        for txt, cmd in [('◀◀', self._primer_emp), ('◀', self._anterior_emp),
                         ('▶', self._siguiente_emp), ('▶▶', self._ultimo_emp)]:
            tk.Button(nav, text=txt, font=FONT_HEAD,
                      command=cmd, bg=COL_CARD, fg=COL_TEXT, relief='flat',
                      padx=6, pady=1).pack(side='left')

        cols = ('cod', 'ape', 'nom')
        self._tree = ttk.Treeview(lf, columns=cols, show='headings', height=14)
        self._tree.heading('cod', text='Cód.')
        self._tree.heading('ape', text='Apellidos')
        self._tree.heading('nom', text='Nombres')
        self._tree.column('cod', width=55, anchor='center')
        self._tree.column('ape', width=180)
        self._tree.column('nom', width=180)

        vsb = ttk.Scrollbar(lf, orient='vertical', command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        self._tree.bind('<<TreeviewSelect>>', self._on_tree_select)

        bot = tk.Frame(lf, bg=COL_BG)
        bot.pack(fill='x', pady=(6, 0))
        self._orden_var = tk.StringVar(value="alfabetico")
        for txt, val in [("A-Z", "alfabetico"), ("Depto.", "departamento")]:
            tk.Radiobutton(bot, text=txt, variable=self._orden_var, value=val,
                           command=self._cargar_lista, font=FONT_LABEL,
                           bg=COL_BG, fg=COL_TEXT,
                           selectcolor=COL_ACCENT, activeforeground=COL_ACCENT,
                           activebackground=COL_BG).pack(side='left', padx=(0, 8))

        ttk.Button(bot, text="Actualizar", command=self._cargar_lista).pack(side='right')

    # ── Notebook ────────────────────────────────────────────────────
    def _build_notebook(self, parent):
        self._nb = ttk.Notebook(parent)
        self._nb.pack(fill='both', expand=True)

        self._tab_datos = ttk.Frame(self._nb)
        self._tab_ingresos = ttk.Frame(self._nb)
        self._tab_observaciones = ttk.Frame(self._nb)
        self._tab_otros = ttk.Frame(self._nb)
        self._tab_certificados = ttk.Frame(self._nb)
        self._tab_referencias = ttk.Frame(self._nb)

        self._nb.add(self._tab_datos, text="Datos Generales")
        self._nb.add(self._tab_ingresos, text="Ingresos / Dctos.")
        self._nb.add(self._tab_observaciones, text="Observaciones")
        self._nb.add(self._tab_otros, text="Otros Datos")
        self._nb.add(self._tab_certificados, text="Certificados")
        self._nb.add(self._tab_referencias, text="Referencias")

        self._build_tab_datos()
        self._build_tab_ingresos()
        self._build_tab_observaciones()
        self._build_tab_otros()
        self._build_tab_certificados()
        self._build_tab_referencias()

        self._nb.bind('<<NotebookTabChanged>>', self._on_tab_change)

    # ── Pestaña: Datos Generales ────────────────────────────────────
    def _build_tab_datos(self):
        self._dg_vars = {}
        canvas = tk.Canvas(self._tab_datos, bg=COL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(self._tab_datos, orient='vertical', command=canvas.yview)
        sf = ttk.Frame(canvas)
        sf.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=sf, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        def field_set(container, row, label, col_key, width=20, is_combo=False, values=None, col=0):
            c = col * 2
            tk.Label(container, text=label, font=FONT_LABEL,
                     bg=COL_BG).grid(row=row, column=c, sticky='w', padx=4, pady=2)
            var = tk.StringVar()
            self._dg_vars[col_key] = var
            if is_combo:
                w = ttk.Combobox(container, textvariable=var, values=values, width=width, state='readonly', font=FONT_DEFAULT)
                self._combos_widgets[col_key] = w
            else:
                w = ttk.Entry(container, textvariable=var, width=width, font=FONT_DEFAULT)
            w.grid(row=row, column=c + 1, sticky='w', padx=4, pady=2)
            self._form_widgets[col_key] = w
            return w

        g1 = ttk.LabelFrame(sf, text="Información Personal", padding=8)
        g1.grid(row=0, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        field_set(g1, 0, 'Código:', 'EMPLEADO', 14, col=0)
        field_set(g1, 0, 'Cédula:', 'CEDULA', 18, col=1)
        field_set(g1, 0, 'Cód.Suc:', 'CODSUC', 10, col=2)
        field_set(g1, 0, 'Cód.Emp:', 'CODEMP', 10, col=3)
        field_set(g1, 1, 'Nombres:', 'NOMBRES', 48, col=0)
        field_set(g1, 2, 'Apellidos:', 'APELLIDOS', 48, col=0)
        w_sexo = field_set(g1, 3, 'Sexo:', 'SEXO', 14, is_combo=True, values=['1', '2'], col=0)
        ToolTip(w_sexo, "1=Masculino, 2=Femenino")
        w_ec = field_set(g1, 3, 'Estado Civil:', 'ESTADO_CI', 20, is_combo=True, values=['1', '2', '3', '4', '5'], col=1)
        ToolTip(w_ec, "Sin catálogo propio en DBTABLAS. Confirmado con datos reales: "
                       "1=Casado, 2=Soltero. Sin confirmar (pocos empleados): "
                       "3=Divorciado, 4=Viudo, 5=Unión Libre")
        w_nac = field_set(g1, 4, 'Lugar Nac.:', 'LUGAR_NAC', 32, col=0)
        field_set(g1, 4, 'Fecha Nac.:', 'FECHA_NAC', 18, col=1)

        g2 = ttk.LabelFrame(sf, text="Ubicación", padding=8)
        g2.grid(row=1, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        field_set(g2, 0, 'Dirección:', 'DIRECCION', 64, col=0)
        field_set(g2, 1, 'Provincia:', 'PROVINCIA', 22, col=0)
        field_set(g2, 1, 'Cantón:', 'CANTON', 22, col=1)
        field_set(g2, 2, 'Parroquia:', 'PARROQUIA', 22, col=0)
        field_set(g2, 2, 'Nacionalidad:', 'NACIONAL', 22, col=1)

        g3 = ttk.LabelFrame(sf, text="Información Laboral", padding=8)
        g3.grid(row=2, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        field_set(g3, 0, 'Fecha Ingreso:', 'FECHA_ING', 12, col=0)
        field_set(g3, 0, 'Fecha Salida:', 'FECHA_SAL', 12, col=1)
        self._readonly_descs = set()

        self._combo_depto = self._crear_selector_catalogo(g3, 1, 'Departamento:', 'DEPTO', "Escriba para buscar el departamento")

        self._combo_cargo = self._crear_selector_catalogo(g3, 2, 'Cargo:', 'CARGO', "Escriba para buscar el cargo")

        self._combo_seccion = self._crear_selector_catalogo(g3, 3, 'Sección:', 'SECCION', "Escriba para buscar la sección")

        field_set(g3, 4, 'Estado:', 'ESTADO', 22, is_combo=True,
                  values=['ACT - ACTIVO', 'LIQ - LIQUIDADO', 'SUS - SUSPENDIDO'], col=0)
        w_tel = field_set(g3, 4, 'Teléfono:', 'TELEFONO', 20, col=1)
        ToolTip(w_tel, "Teléfono principal del empleado")
        field_set(g3, 5, 'Email:', 'emp_mail', 44, col=0)
        w_tel2 = field_set(g3, 6, '2do Teléfono:', 'RPCAM', 20, col=0)
        ToolTip(w_tel2, "Teléfono alternativo o celular")
        w_tt = field_set(g3, 6, 'Tipo de Empleado:', 'TIPO_TRA', 20, is_combo=True,
                          values=['1 - EMPLEADO', '2 - CÓDIGO 2 (sin confirmar)', '3 - OBRERO'], col=1)
        ToolTip(w_tt, "Sin catálogo propio en DBTABLAS. Confirmado con el código real: en "
                       "SP_RP_IMPRIME_ROL el parámetro se llama @TIPO_EMPLEADO, y "
                       "SP_RP_PREPARA_ROL trata explícitamente 3=OBRERO. El 100% de los "
                       "empleados activos usa 1=EMPLEADO. El código 2 solo aparece en "
                       "empleados liquidados (ninguno activo) — sin confirmar qué significa.")
        field_set(g3, 7, 'Actividad:', 'ACTIVIDAD', 34, col=0)
        field_set(g3, 8, 'Cónyugue:', 'CONYUGUE', 44, col=0)

        g4 = ttk.LabelFrame(sf, text="Auditoría", padding=6)
        g4.grid(row=3, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        self._lbl_audit = tk.Label(g4, text="", font=FONT_SMALL,
                                    fg=COL_GRAY, bg=COL_BG, anchor='w')
        self._lbl_audit.pack(fill='x', padx=4, pady=2)

        _bind_scroll_rueda(canvas, canvas)
        _bind_scroll_rueda(sf, canvas)

    # ── Pestaña: Ingresos / Descuentos ───────────────────────────────
    def _build_tab_ingresos(self):
        self._ing_vars = {}
        f = self._tab_ingresos

        def field(frame, row, col, label, key, width=16):
            tk.Label(frame, text=label, font=FONT_LABEL,
                     bg=COL_BG).grid(row=row, column=col * 2, sticky='w', padx=4, pady=2)
            var = tk.StringVar()
            self._ing_vars[key] = var
            w = ttk.Entry(frame, textvariable=var, width=width, font=FONT_DEFAULT)
            w.grid(row=row, column=col * 2 + 1, sticky='w', padx=4, pady=2)
            self._form_widgets[f'ing_{key}'] = w
            return w, var

        g1 = ttk.LabelFrame(f, text="Sueldo y Beneficios de Ley", padding=8)
        g1.pack(fill='x', padx=6, pady=4)
        field(g1, 0, 0, 'Sueldo:', 'SUELDO')
        field(g1, 0, 1, 'Bonificación:', 'BONIFI')
        field(g1, 0, 2, 'Compensación:', 'COMPEN')
        field(g1, 1, 0, 'Transporte:', 'TRANSP')
        field(g1, 1, 1, 'Horas 25%:', 'HOR25')
        field(g1, 1, 2, 'Horas 50%:', 'HOR50')
        field(g1, 2, 0, 'Horas 100%:', 'HOR100')

        g2 = ttk.LabelFrame(f, text="Acumulados de Beneficios Sociales Históricos", padding=8)
        g2.pack(fill='x', padx=6, pady=4)
        field(g2, 0, 0, 'Décimo 3ro:', 'DECIMO3')
        field(g2, 0, 1, 'Décimo 4to:', 'DECIMO4')
        field(g2, 0, 2, 'Vacaciones:', 'VACACION')
        field(g2, 1, 0, 'Fdo. Reserva:', 'FONRESER')

        g3 = ttk.LabelFrame(f, text="Rol Extra", padding=8)
        g3.pack(fill='x', padx=6, pady=4)
        field(g3, 0, 0, 'Moviliza:', 'MOVILIZA')
        field(g3, 0, 1, 'Lunch:', 'LUNCH')
        field(g3, 0, 2, 'Anticipo (%):', 'ANTICIPO')
        field(g3, 1, 0, 'Descuento:', 'DESCUENTO')
        field(g3, 1, 1, 'Ing. Extra:', 'ING_EXTRA')
        field(g3, 1, 2, 'Dct. Extra:', 'DCT_EXTRA')
        field(g3, 2, 0, 'Concepto:', 'CONCEPTO', 34)

        self._flag_vars = {}

        def flag(frame, row, col, text, key, tooltip=None):
            var = tk.BooleanVar()
            self._flag_vars[key] = var
            chk = tk.Checkbutton(frame, text=text, variable=var, font=FONT_DEFAULT,
                                  bg=COL_BG, fg=COL_TEXT, selectcolor=COL_ACCENT,
                                  activebackground=COL_BG, activeforeground=COL_TEXT,
                                  command=self._marcar_modificado)
            chk.grid(row=row, column=col, sticky='w', padx=4, pady=4)
            self._check_widgets[key] = chk
            if tooltip:
                ToolTip(chk, tooltip)
            return chk

        g4 = ttk.LabelFrame(f, text="Parámetros de Nómina (usados por Procesar Rol)", padding=8)
        g4.pack(fill='x', padx=6, pady=4)

        self._afil_iess_var = tk.BooleanVar()

        def _num_afil_es_real(valor):
            """True si NUM_AFIL ya tiene un número real (no 0/vacío/9999999999)."""
            try:
                v = float(valor)
            except (TypeError, ValueError):
                return False
            return v not in (0, 9999999999)

        def _on_toggle_fondo_reserva():
            marcado = self._afil_iess_var.get()
            actual = self._ref_vars.get('NUM_AFIL')
            actual = actual.get() if actual else ''
            if not marcado and _num_afil_es_real(actual):
                # No se permite desmarcar si ya hay un número real cargado:
                # se restaura el check y se pide borrarlo en Referencias,
                # para no perder por accidente un número real de afiliación.
                self._afil_iess_var.set(True)
                messagebox.showwarning(
                    "No se puede desmarcar",
                    "═══════════════════════════════════════════════\n"
                    "  Este empleado ya tiene un número de afiliación\n"
                    f"  IESS real cargado: {actual}\n\n"
                    "  Para marcarlo como NO afiliado, borre ese número\n"
                    "  directamente en la pestaña Referencias.\n"
                    "═══════════════════════════════════════════════"
                )
                return
            if 'NUM_AFIL' in self._ref_vars:
                self._ref_vars['NUM_AFIL'].set('0' if marcado else '9999999999')
            self._marcar_modificado()

        chk_afil = tk.Checkbutton(g4, text='Fondo de Reserva',
                                   variable=self._afil_iess_var, font=FONT_DEFAULT,
                                   bg=COL_BG, fg=COL_TEXT, selectcolor=COL_ACCENT,
                                   activebackground=COL_BG, activeforeground=COL_TEXT,
                                   command=_on_toggle_fondo_reserva)
        chk_afil.grid(row=0, column=0, sticky='w', padx=4, pady=4)
        self._check_widgets['_AFIL_IESS'] = chk_afil
        ToolTip(chk_afil, "Escribe directo en NUM_AFIL (Referencias → No. Afiliación IESS). "
                           "SP_RP_PREPARA_ROL calcula el Fondo de Reserva cuando NUM_AFIL <> 9999999999. "
                           "Marcar=0 (afiliado genérico), desmarcar=9999999999 (no afiliado). "
                           "Si ya hay un número real cargado, bórrelo en Referencias para desmarcar.")

        flag(g4, 1, 0, 'Décimo Tercero se paga aparte (excluir de Procesar Rol)', 'CAT_PROYECT_7',
             "Columna genérica CAT_PROYECT_7. Marcado = este empleado recibe el décimo tercero por una vía distinta al rol automático.")
        flag(g4, 2, 0, 'Décimo Cuarto se paga aparte (excluir de Procesar Rol)', 'CAT_PROYECT_8',
             "Columna genérica CAT_PROYECT_8. Marcado = este empleado recibe el décimo cuarto por una vía distinta al rol automático.")
        flag(g4, 3, 0, 'Aporta IESS Cónyuge (descuento 3.41% del sueldo)', 'RPCAM2',
             "Columna genérica RPCAM2. Activa el rubro 218 (Aporte IESS Cónyuge) en Procesar Rol.")

    # ── Pestaña: Observaciones ──────────────────────────────────────
    def _build_tab_observaciones(self):
        f = self._tab_observaciones
        top = ttk.Frame(f)
        top.pack(fill='x', padx=8, pady=8)

        tk.Label(top, text='Período:', font=FONT_LABEL,
                 bg=COL_BG).pack(side='left', padx=(0, 6))
        meses = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
                 'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
        self._mes_var = tk.StringVar(value=meses[datetime.now().month - 1])
        ttk.Combobox(top, textvariable=self._mes_var, values=meses, width=12, state='readonly').pack(side='left', padx=6)
        self._anio_var = tk.StringVar(value=str(datetime.now().year))
        ttk.Combobox(top, textvariable=self._anio_var,
                     values=[str(y) for y in range(2020, 2031)],
                     width=8, state='readonly').pack(side='left', padx=6)
        ttk.Button(top, text="Mostrar", command=self._mostrar_obs).pack(side='left', padx=8)
        ttk.Button(top, text="💾 Guardar Obs.", command=self._guardar_obs,
                  style='Accent.TButton').pack(side='left', padx=4)
        ttk.Button(top, text="🖨 Imprimir Historial", command=self._imprimir_observaciones).pack(side='left', padx=4)

        self._lbl_fecha_fin = tk.Label(top, text='', font=FONT_DEFAULT,
                                       fg=COL_GRAY, bg=COL_BG)

        # Para almacenar referencias a los Text widgets de observaciones
        self._obs_widgets = []
        self._lbl_fecha_fin.pack(side='right')

        # Canvas con scroll para mostrar recuadros de observaciones
        self._obs_canvas = tk.Canvas(f, bg=COL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(f, orient='vertical', command=self._obs_canvas.yview)
        self._obs_frame = tk.Frame(self._obs_canvas, bg=COL_BG)

        self._obs_canvas.pack(side='left', fill='both', expand=True, padx=8, pady=8)
        vsb.pack(side='right', fill='y', padx=(0, 8), pady=8)

        self._obs_canvas.configure(yscrollcommand=vsb.set)
        self._obs_canvas_window = self._obs_canvas.create_window(0, 0, window=self._obs_frame, anchor='nw')

        # Bind para resize
        self._obs_frame.bind('<Configure>', lambda e: self._obs_canvas.configure(scrollregion=self._obs_canvas.bbox('all')))
        self._obs_canvas.bind('<MouseWheel>', lambda e: self._obs_canvas.yview_scroll(int(-1*(e.delta/120)), 'units'))
        self._obs_canvas.bind('<Button-4>', lambda e: self._obs_canvas.yview_scroll(-1, 'units'))
        self._obs_canvas.bind('<Button-5>', lambda e: self._obs_canvas.yview_scroll(1, 'units'))

        # Placeholder
        lbl = tk.Label(self._obs_frame, text="Selecciona un período y presiona 'Mostrar'",
                      font=FONT_LABEL, bg=COL_BG, fg=COL_GRAY)
        lbl.pack(pady=20)

        self._form_widgets['OBSERV'] = self._obs_frame

    def _mostrar_obs(self):
        """Mostrar observaciones en recuadros"""
        if not self.empleado_actual:
            messagebox.showwarning("ATENCIÓN", "Seleccione un empleado primero")
            return

        m = self._mes_var.get()
        a = self._anio_var.get()

        meses = {'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
                'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12}
        mes_num = meses.get(m, 6)
        ano_num = int(a) if a.isdigit() else 2026
        emp_cod = self.empleado_actual.get('EMPLEADO', '')

        # Limpiar frame anterior
        for widget in self._obs_frame.winfo_children():
            widget.destroy()
        self._obs_widgets = []  # Limpiar referencias anteriores
        self._status("Cargando observaciones...")

        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"""
                        SELECT refer1, refer2, refer3, refer4, refer5, refer6, refer7, fecha_ven
                        FROM RPEMPOBSERV
                        WHERE empleado = ? AND MONTH(fecha_ven) = ? AND YEAR(fecha_ven) = ?
                          AND CODEMP='10' AND CODSUC='10'
                        ORDER BY fecha_ven DESC
                    """, (emp_cod, mes_num, ano_num))
                    row = cur.fetchone()
                self.root.after(0, lambda: self._render_obs(row, m, a))
            except Exception as e:
                LOG.error("Error cargando observaciones: %s", e)
                self.root.after(0, lambda msg=str(e): self._obs_error(msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _obs_error(self, msg):
        lbl = tk.Label(self._obs_frame, text=f"Error: {msg}", font=FONT_LABEL,
                     bg=COL_BG, fg=COL_DANGER)
        lbl.pack(pady=20)
        messagebox.showerror("Error", msg)
        self._status("Error al cargar observaciones")

    def _render_obs(self, row, m, a):
        for widget in self._obs_frame.winfo_children():
            widget.destroy()
        self._obs_widgets = []

        if not row:
            lbl = tk.Label(self._obs_frame, text=f"Sin observaciones para {m} {a}",
                         font=FONT_LABEL, bg=COL_BG, fg=COL_GRAY)
            lbl.pack(pady=20)
            self._lbl_fecha_fin.config(text='')
            self._status("Sin observaciones para el período")
            return

        try:
            fecha_ven = row[7]

            # Encabezado
            header = tk.Frame(self._obs_frame, bg=COL_HEADER, height=60)
            header.pack(fill='x', padx=0, pady=(0, 10))

            tk.Label(header, text=f"📅 {m} {a}  |  📆 {fecha_ven.strftime('%d/%m/%Y')}",
                    font=FONT_HEAD, bg=COL_HEADER, fg=COL_WHITE).pack(pady=10)

            # Se muestran los 7 campos SIEMPRE (llenos y vacíos), para que nunca
            # haga falta escribir una observación nueva encima de una ya existente.
            campos_llenos = 0
            for i in range(7):
                lleno = bool(row[i])
                if lleno:
                    campos_llenos += 1

                card = tk.Frame(self._obs_frame, bg=COL_CARD, relief='solid', borderwidth=2)
                card.pack(fill='both', expand=False, padx=0, pady=(0, 8))

                card_header = tk.Frame(card, bg=COL_ACCENT if lleno else COL_GRAY, height=30)
                card_header.pack(fill='x')

                titulo = f"[CAMPO {i+1}]" if lleno else f"[CAMPO {i+1} — VACÍO, disponible para nueva observación]"
                tk.Label(card_header, text=titulo, font=FONT_LABEL,
                        bg=COL_ACCENT if lleno else COL_GRAY,
                        fg=COL_WHITE if lleno else COL_HEADER).pack(pady=6)

                content = tk.Text(card, font=FONT_DEFAULT,
                                bg=COL_CARD, fg=COL_TEXT,
                                height=3, width=70, wrap='word',
                                relief='flat', borderwidth=0)
                content.pack(fill='both', expand=True, padx=12, pady=12)
                if lleno:
                    content.insert(1.0, row[i])

                self._obs_widgets.append({
                    'campo': i + 1,
                    'widget': content,
                    'valor_original': row[i] or ''
                })

            # Footer
            if campos_llenos >= 7:
                nota = "Los 7 campos están llenos. Para agregar más use Carga Masiva de Observaciones (crea una fila adicional)."
            else:
                nota = f"Total: {campos_llenos}/7 campos llenos"
            footer = tk.Label(self._obs_frame, text=nota,
                            font=FONT_SMALL, bg=COL_BG, fg=COL_GRAY, wraplength=600, justify='left')
            footer.pack(pady=10)

            self._lbl_fecha_fin.config(text=f'Fecha: {fecha_ven.strftime("%d/%m/%Y")}')
            self._status("Observaciones cargadas correctamente")
            _bind_scroll_rueda(self._obs_frame, self._obs_canvas)

        except Exception as e:
            lbl = tk.Label(self._obs_frame, text=f"Error: {str(e)}", font=FONT_LABEL,
                         bg=COL_BG, fg=COL_DANGER)
            lbl.pack(pady=20)
            messagebox.showerror("Error", str(e))

    def _guardar_obs(self):
        """Guardar observaciones modificadas en RPEMPOBSERV"""
        if not self.empleado_actual or not self._obs_widgets:
            messagebox.showwarning("ATENCIÓN", "Cargue observaciones primero con 'Mostrar'")
            return

        m = self._mes_var.get()
        a = self._anio_var.get()
        meses = {'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
                'JULIO': 7, 'AGOSTO': 8, 'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12}
        mes_num = meses.get(m, 6)
        ano_num = int(a) if a.isdigit() else 2026
        emp_cod = self.empleado_actual.get('EMPLEADO', '')

        # Leer el contenido de los widgets en el hilo principal (Tkinter no es thread-safe)
        cambios_pendientes = []
        for widget_info in self._obs_widgets:
            texto_nuevo = widget_info['widget'].get(1.0, 'end').strip()
            if texto_nuevo != widget_info['valor_original']:
                cambios_pendientes.append((widget_info['campo'], texto_nuevo))

        if not cambios_pendientes:
            messagebox.showinfo("Sin cambios", "No hay cambios para guardar")
            return

        self._status("Guardando observaciones...")

        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()

                    # Obtener la fecha_ven de la observación actual
                    cur.execute(f"""
                        SELECT fecha_ven FROM RPEMPOBSERV
                        WHERE empleado = ? AND MONTH(fecha_ven) = ? AND YEAR(fecha_ven) = ?
                          AND CODEMP='10' AND CODSUC='10'
                        ORDER BY fecha_ven DESC
                    """, (emp_cod, mes_num, ano_num))

                    row = cur.fetchone()
                    if not row:
                        self.root.after(0, lambda: messagebox.showerror("Error", "No se encontró la observación"))
                        self.root.after(0, lambda: self._status("Error al guardar"))
                        return

                    fecha_ven = row[0]

                    cambios = 0
                    for campo_num, texto_nuevo in cambios_pendientes:
                        columna = f'refer{campo_num}'
                        cur.execute(f"""
                            UPDATE TOP (1) RPEMPOBSERV
                            SET {columna} = ?
                            WHERE empleado = ? AND fecha_ven = ? AND CODEMP='10' AND CODSUC='10'
                        """, (texto_nuevo, emp_cod, fecha_ven))
                        cambios += 1

                    self.conn.commit()
                self.root.after(0, lambda: self._guardar_obs_ok(cambios))
            except Exception as e:
                LOG.error("Error guardando observaciones: %s", e)
                self.root.after(0, lambda msg=str(e): messagebox.showerror("Error", f"No se pudo guardar: {msg}"))
                self.root.after(0, lambda: self._status("Error al guardar"))
        threading.Thread(target=tarea, daemon=True).start()

    def _guardar_obs_ok(self, cambios):
        messagebox.showinfo("✅ Guardado", f"Se guardaron {cambios} cambios en observaciones")
        self._mostrar_obs()  # Recargar para confirmar

    def _imprimir_observaciones(self):
        """Genera un HTML con TODO el historial de observaciones del empleado
        (todas las fechas, no solo el período mostrado) y lo abre en el
        navegador para imprimir desde ahí (Ctrl+P)."""
        if not self.empleado_actual:
            messagebox.showwarning("ATENCIÓN", "Seleccione un empleado primero")
            return
        emp_cod = self.empleado_actual.get('EMPLEADO', '')
        nom = f"{self.empleado_actual.get('NOMBRES', '')} {self.empleado_actual.get('APELLIDOS', '')}".strip()
        ced = self.empleado_actual.get('CEDULA', '')

        self._status("Generando historial de observaciones...")

        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"""
                        SELECT refer1, refer2, refer3, refer4, refer5, refer6, refer7, fecha_ven
                        FROM RPEMPOBSERV
                        WHERE empleado = ? AND CODEMP='10' AND CODSUC='10'
                        ORDER BY fecha_ven DESC
                    """, (emp_cod,))
                    rows = cur.fetchall()
                self.root.after(0, lambda: self._generar_imprimir_obs(rows, emp_cod, nom, ced))
                self.root.after(0, lambda: self._status("Listo"))
            except Exception as e:
                LOG.error("Error obteniendo historial para imprimir: %s", e)
                self.root.after(0, lambda msg=str(e): messagebox.showerror("Error", f"No se pudo obtener el historial: {msg}"))
                self.root.after(0, lambda: self._status("Error"))
        threading.Thread(target=tarea, daemon=True).start()

    def _generar_imprimir_obs(self, rows, emp_cod, nom, ced):
        if not rows:
            messagebox.showinfo("Sin datos", "Este empleado no tiene observaciones registradas")
            return

        secciones = []
        for row in rows:
            fecha_ven = row[7]
            campos = [row[i] for i in range(7) if row[i]]
            if not campos:
                continue
            items = "".join(f"<li>{html.escape(str(c))}</li>" for c in campos)
            secciones.append(
                f'<div class="fecha">{fecha_ven.strftime("%d/%m/%Y")}</div>'
                f'<ul>{items}</ul>'
            )

        contenido = "\n".join(secciones) if secciones else "<p>Sin observaciones con contenido.</p>"

        html_doc = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8">
<title>Historial de Observaciones — {html.escape(nom)}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 32px; color: #222; }}
  h1 {{ font-size: 18px; margin-bottom: 2px; }}
  .info {{ color: #555; margin-bottom: 20px; font-size: 13px; }}
  .fecha {{ background: #0D1B2A; color: #fff; padding: 6px 10px; margin-top: 16px;
            font-weight: bold; border-radius: 3px; }}
  ul {{ margin: 6px 0 0 0; padding-left: 24px; }}
  li {{ margin-bottom: 6px; line-height: 1.4; }}
  @media print {{ .fecha {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }} }}
</style></head>
<body>
  <h1>INSEVIG — Historial de Observaciones</h1>
  <div class="info">
    Empleado: <b>{html.escape(nom)}</b> &nbsp;|&nbsp; Código: {html.escape(str(emp_cod))}
    &nbsp;|&nbsp; C.I.: {html.escape(str(ced))}<br>
    Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}
  </div>
  {contenido}
</body></html>"""

        try:
            tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8')
            tmp.write(html_doc)
            tmp.close()
            webbrowser.open(f'file:///{tmp.name}')
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar la vista de impresión: {e}")

    # ── Pestaña: Otros Datos ────────────────────────────────────────
    def _build_tab_otros(self):
        self._ot_vars = {}
        self._check_states = {'INCL_ROL': False, 'INCL_BAN': False}
        f = self._tab_otros

        g1 = ttk.LabelFrame(f, text="Datos Generales", padding=8)
        g1.pack(fill='x', padx=6, pady=4)

        self._ot_vars['INCL_ROL'] = tk.StringVar(value='N')
        self._ot_vars['INCL_BAN'] = tk.StringVar(value='N')
        check_frame = ttk.Frame(g1)
        check_frame.grid(row=0, column=0, columnspan=2, sticky='w', padx=4, pady=2)
        self._chk_rol = tk.Checkbutton(check_frame, text="Incluir en el Rol",
                                       variable=tk.BooleanVar(),
                                       font=FONT_LABEL,
                                       bg=COL_BG, fg=COL_TEXT,
                                       selectcolor=COL_ACCENT, activebackground=COL_BG,
                                       command=lambda: self._toggle_check('INCL_ROL'))
        self._chk_rol.pack(side='left', padx=(0, 24))
        self._check_widgets['INCL_ROL'] = self._chk_rol
        ToolTip(self._chk_rol, "Columna RPEMPLEA.INCL_ROL ('S'/'N'). No aparece como condición "
                                "en SP_RP_PREPARA_ROL (el que corre Procesar Rol) — solo se exporta "
                                "como columna en reportes generales de empleados (SP_RP_EMPLEADOS, "
                                "SP_RP_EMPLEADOS_X_DEPTO), no filtra el cálculo de nómina en sí.")
        self._chk_ban = tk.Checkbutton(check_frame, text="Acreditar",
                                       variable=tk.BooleanVar(),
                                       font=FONT_LABEL,
                                       bg=COL_BG, fg=COL_TEXT,
                                       selectcolor=COL_ACCENT, activebackground=COL_BG,
                                       command=lambda: self._toggle_check('INCL_BAN'))
        self._chk_ban.pack(side='left')
        self._check_widgets['INCL_BAN'] = self._chk_ban
        ToolTip(self._chk_ban, "Columna RPEMPLEA.INCL_BAN ('S'/'N'). Confirmado en "
                                "SP_RP_ANEXO_BANCO_INSEVIG: WHERE INCL_BAN='S' — solo estos "
                                "empleados entran en el archivo/anexo que se envía al banco para "
                                "acreditar el sueldo.")

        def field(container, row, col, label, key, width=14, is_combo=False, values=None):
            tk.Label(container, text=label, font=FONT_LABEL,
                     bg=COL_BG).grid(row=row, column=col * 2, sticky='w', padx=4, pady=2)
            var = tk.StringVar()
            self._ot_vars[key] = var
            if is_combo:
                w = ttk.Combobox(container, textvariable=var, values=values, width=width, state='readonly', font=FONT_DEFAULT)
                self._combos_widgets[key] = w
            else:
                w = ttk.Entry(container, textvariable=var, width=width, font=FONT_DEFAULT)
            w.grid(row=row, column=col * 2 + 1, sticky='w', padx=4, pady=2)
            self._form_widgets[f'ot_{key}'] = w

        field(g1, 1, 0, 'Cargas:', 'CARGAS', 10)
        field(g1, 2, 0, 'Últ. Liquidación:', 'ULTLIQ', 16)
        field(g1, 2, 1, 'Últ. Día Trabajado:', 'ULTDIATRA', 16)
        field(g1, 3, 0, 'Días Trab.:', 'DIAS_TRA', 10)
        field(g1, 4, 0, 'Grupo Sang.:', 'TIP_SAN', 12, is_combo=True,
              values=['O+','O-','A+','A-','B+','B-','AB+','AB-'])
        field(g1, 5, 0, 'Período de Pago:', 'TIPO_PGO', 12, is_combo=True, values=['1', '2', '3'])

        g2 = ttk.LabelFrame(f, text="Cuentas Contables", padding=8)
        g2.pack(fill='x', padx=6, pady=4)
        field(g2, 0, 0, 'Código Cta.:', 'CODCTA')
        field(g2, 1, 0, 'Cta. Depto.:', 'CTADPT')
        field(g2, 2, 0, 'Cta. Auxiliar:', 'CTAAUX')

        g3 = ttk.LabelFrame(f, text="Información Bancaria", padding=8)
        g3.pack(fill='x', padx=6, pady=4)
        field(g3, 0, 0, 'Banco:', 'RUTA4', 18, is_combo=True,
              values=['PRODUBANCO','PICHINCHA','GUAYAQUIL','PACIFICO'])
        field(g3, 1, 0, 'Cta. Cte.:', 'CTA_CTE', 24)
        field(g3, 2, 0, 'Cta. Ahorros:', 'CTA_AHO', 24)

    def _toggle_check(self, field):
        cur = self._check_states.get(field, False)
        self._check_states[field] = not cur
        self._ot_vars[field].set('S' if not cur else 'N')
        chk = self._chk_rol if field == 'INCL_ROL' else self._chk_ban
        if not cur:
            chk.select()
        else:
            chk.deselect()
        self._marcar_modificado()

    def _actualizar_check_visual(self):
        for f, var in [('INCL_ROL', self._chk_rol), ('INCL_BAN', self._chk_ban)]:
            if self._ot_vars[f].get() == 'S':
                var.select()
                self._check_states[f] = True
            else:
                var.deselect()
                self._check_states[f] = False

    # ── Pestaña: Certificados ───────────────────────────────────────
    def _build_tab_certificados(self):
        self._cert_vars = {}
        f = self._tab_certificados

        cf = ttk.Frame(f)
        cf.pack(fill='x', padx=8, pady=8)
        for i, nombre in enumerate(['Cédula Identidad', 'Cert. Votación',
                                    'Record Policial', 'Libreta Militar']):
            bx = ttk.LabelFrame(cf, text=nombre, padding=10)
            bx.grid(row=0, column=i, padx=6, pady=6, sticky='n')
            cv = tk.Canvas(bx, width=120, height=90, bg=COL_CARD,
                           highlightbackground=COL_BORDER, highlightthickness=2)
            cv.pack(pady=6)
            cv.create_text(60, 45, text="Archivo", font=FONT_DEFAULT, fill=COL_GRAY)

        g1 = ttk.LabelFrame(f, text="Familiares", padding=8)
        g1.pack(fill='x', padx=6, pady=4)
        tk.Label(g1, text='Nombres:', font=FONT_LABEL, bg=COL_BG).grid(row=0, column=0, sticky='w', padx=4, pady=2)
        self._cert_vars['NOM_FAM'] = tk.StringVar()
        w = ttk.Entry(g1, textvariable=self._cert_vars['NOM_FAM'], width=44); w.grid(row=0, column=1, sticky='w', padx=4, pady=2)
        self._form_widgets['cert_NOM_FAM'] = w
        tk.Label(g1, text='Dirección:', font=FONT_LABEL, bg=COL_BG).grid(row=1, column=0, sticky='w', padx=4, pady=2)
        self._cert_vars['DIR_FAM'] = tk.StringVar()
        w = ttk.Entry(g1, textvariable=self._cert_vars['DIR_FAM'], width=44); w.grid(row=1, column=1, sticky='w', padx=4, pady=2)
        self._form_widgets['cert_DIR_FAM'] = w
        tk.Label(g1, text='Teléfonos:', font=FONT_LABEL, bg=COL_BG).grid(row=2, column=0, sticky='w', padx=4, pady=2)
        self._cert_vars['TEL_FAM'] = tk.StringVar()
        w = ttk.Entry(g1, textvariable=self._cert_vars['TEL_FAM'], width=22); w.grid(row=2, column=1, sticky='w', padx=4, pady=2)
        self._form_widgets['cert_TEL_FAM'] = w

        g2 = ttk.LabelFrame(f, text="No Familiares", padding=8)
        g2.pack(fill='x', padx=6, pady=4)
        tk.Label(g2, text='Nombres:', font=FONT_LABEL, bg=COL_BG).grid(row=0, column=0, sticky='w', padx=4, pady=2)
        self._cert_vars['NOM_NO_FAM'] = tk.StringVar()
        w = ttk.Entry(g2, textvariable=self._cert_vars['NOM_NO_FAM'], width=44); w.grid(row=0, column=1, sticky='w', padx=4, pady=2)
        self._form_widgets['cert_NOM_NO_FAM'] = w
        tk.Label(g2, text='Dirección:', font=FONT_LABEL, bg=COL_BG).grid(row=1, column=0, sticky='w', padx=4, pady=2)
        self._cert_vars['DIR_NO_FAM'] = tk.StringVar()
        w = ttk.Entry(g2, textvariable=self._cert_vars['DIR_NO_FAM'], width=44); w.grid(row=1, column=1, sticky='w', padx=4, pady=2)
        self._form_widgets['cert_DIR_NO_FAM'] = w
        tk.Label(g2, text='Teléfonos:', font=FONT_LABEL, bg=COL_BG).grid(row=2, column=0, sticky='w', padx=4, pady=2)
        self._cert_vars['TEL_NO_FAM'] = tk.StringVar()
        w = ttk.Entry(g2, textvariable=self._cert_vars['TEL_NO_FAM'], width=22); w.grid(row=2, column=1, sticky='w', padx=4, pady=2)
        self._form_widgets['cert_TEL_NO_FAM'] = w

    # ── Pestaña: Referencias ────────────────────────────────────────
    def _build_tab_referencias(self):
        self._ref_vars = {}
        f = self._tab_referencias
        canvas = tk.Canvas(f, bg=COL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(f, orient='vertical', command=canvas.yview)
        sf = ttk.Frame(canvas)
        sf.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=sf, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        def field(container, row, label, key, width=22):
            tk.Label(container, text=label, font=FONT_LABEL,
                     bg=COL_BG).grid(row=row, column=0, sticky='w', padx=4, pady=2)
            var = tk.StringVar()
            self._ref_vars[key] = var
            w = ttk.Entry(container, textvariable=var, width=width, font=FONT_DEFAULT)
            w.grid(row=row, column=1, sticky='w', padx=4, pady=2)
            self._form_widgets[f'ref_{key}'] = w

        g1 = ttk.LabelFrame(sf, text="Datos Referenciales", padding=8)
        g1.grid(row=0, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        field(g1, 0, 'Cédula Militar:', 'CED_MIL', 20)
        tk.Label(g1, text='Edad:', font=FONT_LABEL, bg=COL_BG).grid(row=0, column=2, sticky='w', padx=4, pady=2)
        self._ref_vars['EDAD'] = tk.StringVar()
        w = ttk.Entry(g1, textvariable=self._ref_vars['EDAD'], width=10, font=FONT_DEFAULT); w.grid(row=0, column=3, sticky='w', padx=4, pady=2)
        self._form_widgets['ref_EDAD'] = w
        field(g1, 1, 'Tipo Sangre:', 'TIP_SAN', 12)
        field(g1, 2, 'Nro Cert. Votación:', 'IDVOTA', 20)
        field(g1, 3, 'Licencia Conducir:', 'LICCOND', 16)
        field(g1, 4, 'Código IESS:', 'CODIESS', 20)
        field(g1, 5, 'Carnet Conadis:', 'ID_CONADIS', 20)
        field(g1, 6, 'Visita Domiciliaria:', 'OBSERV', 60)

        g2 = ttk.LabelFrame(sf, text="Estudios", padding=8)
        g2.grid(row=1, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        self._ref_vars['PRIMARIA'] = tk.BooleanVar()
        chk = tk.Checkbutton(g2, text='Primaria', variable=self._ref_vars['PRIMARIA'],
                       font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT,
                       selectcolor=COL_ACCENT, activebackground=COL_BG)
        chk.grid(row=0, column=0, sticky='w', padx=4, pady=2)
        self._check_widgets['PRIMARIA'] = chk
        self._ref_vars['SECUNDARIA'] = tk.BooleanVar()
        chk = tk.Checkbutton(g2, text='Secundaria', variable=self._ref_vars['SECUNDARIA'],
                       font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT,
                       selectcolor=COL_ACCENT, activebackground=COL_BG)
        chk.grid(row=0, column=1, sticky='w', padx=4, pady=2)
        self._check_widgets['SECUNDARIA'] = chk
        self._ref_vars['EST_SUP'] = tk.BooleanVar()
        chk = tk.Checkbutton(g2, text='Universidad', variable=self._ref_vars['EST_SUP'],
                       font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT,
                       selectcolor=COL_ACCENT, activebackground=COL_BG)
        chk.grid(row=0, column=2, sticky='w', padx=4, pady=2)
        self._check_widgets['EST_SUP'] = chk
        field(g2, 1, 'Título:', 'TITULO', 34)
        field(g2, 2, 'Años Estudio:', 'ANIO_EST', 10)

        g3 = ttk.LabelFrame(sf, text="Servicios", padding=8)
        g3.grid(row=2, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        field(g3, 0, 'Tipo:', 'RPCAM5', 30)
        field(g3, 1, 'Contrato Inspectoría:', 'CONTINS', 30)
        field(g3, 2, 'GIPASE:', 'RPCAM3', 38)
        field(g3, 3, 'AFIS:', 'RPCAM4', 38)
        field(g3, 4, 'Certificados:', 'certificados', 38)
        field(g3, 5, 'Reentrenamiento:', 'reentrenamiento', 38)
        field(g3, 6, 'Vacuna:', 'vacuna', 38)

        self._ref_vars['FZA_PUB'] = tk.BooleanVar()
        chk = tk.Checkbutton(g3, text='Miembro activo de la Fuerza Pública', variable=self._ref_vars['FZA_PUB'],
                       font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT,
                       selectcolor=COL_ACCENT, activebackground=COL_BG)
        chk.grid(row=7, column=0, columnspan=2, sticky='w', padx=4, pady=2)
        self._check_widgets['FZA_PUB'] = chk

        self._ref_vars['SER_MIL'] = tk.BooleanVar()
        chk = tk.Checkbutton(g3, text='Realizó el Servicio Militar', variable=self._ref_vars['SER_MIL'],
                       font=FONT_LABEL, bg=COL_BG, fg=COL_TEXT,
                       selectcolor=COL_ACCENT, activebackground=COL_BG)
        chk.grid(row=8, column=0, columnspan=2, sticky='w', padx=4, pady=2)
        self._check_widgets['SER_MIL'] = chk

        g4 = ttk.LabelFrame(sf, text="Información Adicional", padding=8)
        g4.grid(row=3, column=0, columnspan=2, sticky='ew', padx=4, pady=3)
        field(g4, 0, 'Cert. Violencia Intraf.:', 'CERTVINF', 50)
        field(g4, 1, 'Maniobras:', 'MANIOBRAS', 50)
        field(g4, 2, 'No. Afiliación IESS:', 'NUM_AFIL', 20)

        _bind_scroll_rueda(canvas, canvas)
        _bind_scroll_rueda(sf, canvas)

    # ── Lazy Refresh ────────────────────────────────────────────────
    def _on_tab_change(self, ev=None):
        if not self._running:
            return
        tab = self._nb.select()
        mapping = {}
        for i, (name, key) in enumerate([
            ('Datos Generales', 'datos'), ('Ingresos / Dctos.', 'ingresos'),
            ('Observaciones', 'obs'), ('Otros Datos', 'otros'),
            ('Certificados', 'cert'), ('Referencias', 'ref')
        ]):
            mapping[str(self._nb.children.get(f'!frame{i+1 if i>0 else ""}', ''))] = key
        # Simpler: map by tab text
        txt = self._nb.tab(tab, 'text')
        txt_to_key = {
            'Datos Generales': 'datos', 'Ingresos / Dctos.': 'ingresos',
            'Observaciones': 'obs', 'Otros Datos': 'otros',
            'Certificados': 'cert', 'Referencias': 'ref'
        }
        key = txt_to_key.get(txt)
        if key and key in self._tabs_dirty:
            self._tabs_dirty.discard(key)

    def _marcar_dirty(self, *keys):
        for k in keys:
            self._tabs_dirty.add(k)

    # ── BD: Conexión ────────────────────────────────────────────────
    def _conectar_bd(self):
        def tarea():
            try:
                conn = _get_sql_conn()
                self.conn = conn
                self._cargar_catalogos()
                self.root.after(100, self._actualizar_combos_catalogos)
                self.root.after(200, self._cargar_lista)
                self.root.after(250, self._cargar_autocomplete_empleados)
                self.root.after(0, lambda: self._status("Conectado a SQL Server"))
            except Exception as e:
                self.root.after(0, lambda msg=f"Error BD: {str(e)}": self._status(msg))
                self.root.after(0, lambda msg=str(e): messagebox.showerror("Error BD", msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _cargar_catalogos(self):
        if not self.conn or self._catalogos_cargados:
            return
        try:
            with self.conn_lock:
                cur = self.conn.cursor()
                # ECS, TTR y FPF no se cargan aquí: no existe catálogo real para
                # Estado Civil/Tipo Trabajo en DBTABLAS, y FPF es un catálogo
                # de forma de pago de facturación que no corresponde a
                # RPEMPLEA.TIPO_PGO (ver self.estados_civiles/tipos_trabajo/
                # periodos_pago en __init__ para el detalle).
                for tipo, dest in [('FNC', self.cargos), ('SEC', self.secciones), ('DPT', self.departamentos),
                                   ('BAN', self.bancos)]:
                    try:
                        cur.execute("SELECT CODIGO, NOMBRE FROM DBTABLAS WHERE TIPO = ? AND CODEMP = '10' ORDER BY NOMBRE", (tipo,))
                        dest.clear()
                        for r in cur.fetchall():
                            dest[str(r[0]).strip()] = r[1]
                    except Exception as e:
                        LOG.error("Error cargando catalogo TIPO='%s': %s", tipo, e)
                        dest.clear()
            self._catalogos_cargados = True
        except Exception as e:
            LOG.error("Error general cargando catalogos: %s", e)

    def _obtener_nombre(self, codigo, catalogo):
        if not codigo:
            return ""
        return catalogo.get(str(codigo).strip(), "")

    def _extraer_codigo(self, val):
        if val and ' - ' in val:
            return val.split(' - ')[0].strip()
        return val

    # ── Flags de nómina (CAT_PROYECT_7/8, RPCAM2) ─────────────────────
    # Columnas varchar genéricas reutilizadas que usan '1'/'0' como texto
    # (ver 14_FONDO_RESERVA_DECIMOS_APORTE_CONYUGE.txt). NUM_AFIL (afiliado
    # IESS / aplica Fondo de Reserva) NO vive aquí: es un número real, no
    # un booleano, así que se muestra aparte como casillero de solo lectura
    # calculado directo de NUM_AFIL, y se edita en Referencias.
    def _cargar_flag_vars(self, datos):
        for k, v in self._flag_vars.items():
            v.set(str(datos.get(k)) == '1')
        num_afil = datos.get('NUM_AFIL')
        try:
            # Misma condición exacta de SP_RP_PREPARA_ROL (IF @MNUMAFI <> 9999999999):
            # NUM_AFIL=0 SÍ cuenta como afiliado aquí (es el valor por defecto en la
            # gran mayoría de empleados activos) — solo 9999999999 es "no afiliado".
            afiliado = num_afil is not None and float(num_afil) != 9999999999
        except (TypeError, ValueError):
            afiliado = False
        self._afil_iess_var.set(afiliado)

    def _flag_vars_a_datos(self):
        return {k: ('1' if v.get() else '0') for k, v in self._flag_vars.items()}

    def _match_combo_val(self, codigo, items):
        if not codigo:
            return ""
        cod = str(codigo).strip()
        for item in items:
            if item.startswith(cod + ' - ') or item == cod:
                return item
        return cod

    def _actualizar_combos_catalogos(self):
        for key, cat in [('DEPTO', self.departamentos), ('CARGO', self.cargos), ('SECCION', self.secciones)]:
            if cat:
                self._combo_full_values[key] = sorted([f"{k} - {v}" for k, v in cat.items()])
        for key, cat in [('SEXO', self.sexos), ('ESTADO_CI', self.estados_civiles),
                          ('TIPO_TRA', self.tipos_trabajo), ('TIPO_PGO', self.periodos_pago),
                          ('RUTA4', self.bancos)]:
            combo = self._combos_widgets.get(key)
            if combo and cat:
                items = sorted([f"{k} - {v}" for k, v in cat.items()])
                combo['values'] = items

    def _crear_selector_catalogo(self, parent, row, label, key, tooltip):
        """Campo de texto ancho + lista desplegable propia (no ttk.Combobox
        nativo: su lista interna no se refresca de forma confiable mientras
        está abierta y se está escribiendo). Al escribir filtra en vivo
        self._combo_full_values[key] ('código - descripción'); al elegir un
        resultado actualiza self._dg_vars[key] igual que antes."""
        tk.Label(parent, text=label, font=FONT_LABEL,
                 bg=COL_BG).grid(row=row, column=0, sticky='w', padx=4, pady=2)
        self._dg_vars[key] = tk.StringVar()
        entry = ttk.Entry(parent, textvariable=self._dg_vars[key], width=68, font=FONT_DEFAULT)
        entry.grid(row=row, column=1, columnspan=2, sticky='w', padx=4, pady=2)
        self._form_widgets[key] = entry
        ToolTip(entry, tooltip)

        sug_frame = tk.Frame(parent, bg=COL_CARD, relief='solid', borderwidth=1)
        sug_list = tk.Listbox(sug_frame, font=FONT_DEFAULT, bg=COL_ENTRY_BG, fg=COL_TEXT,
                               selectbackground=COL_ACCENT, selectforeground=COL_HEADER,
                               height=8, borderwidth=0, highlightthickness=0, activestyle='none')
        sug_list.pack(fill='both', expand=True, padx=2, pady=2)
        estado = {'visible': False}

        def mostrar():
            if not estado['visible']:
                x = entry.winfo_x()
                y = entry.winfo_y() + entry.winfo_height()
                ancho = max(420, entry.winfo_width())
                sug_frame.place(in_=parent, x=x, y=y, width=ancho)
                sug_frame.lift()
                estado['visible'] = True

        def ocultar():
            if estado['visible']:
                sug_frame.place_forget()
                estado['visible'] = False

        def on_key(event):
            if not self.modo_edicion:
                return
            if event.keysym in ('Up', 'Down', 'Return', 'Escape', 'Tab'):
                return
            texto = entry.get().strip().lower()
            full = self._combo_full_values.get(key, [])
            if not texto:
                ocultar()
                return
            filtrado = [v for v in full if texto in v.lower()][:30]
            if not filtrado:
                ocultar()
                return
            sug_list.delete(0, 'end')
            for v in filtrado:
                sug_list.insert('end', v)
            mostrar()

        def elegir(_event=None):
            sel = sug_list.curselection()
            if not sel:
                return
            self._dg_vars[key].set(sug_list.get(sel[0]))
            ocultar()
            self._marcar_modificado()

        def on_focus_out(_event):
            entry.after(150, _validar)

        def _validar():
            ocultar()
            texto = entry.get()
            full = self._combo_full_values.get(key, [])
            if texto and texto not in full:
                cod = self._extraer_codigo(texto)
                match = self._match_combo_val(cod, full)
                self._dg_vars[key].set(match if match else '')

        entry.bind('<KeyRelease>', on_key)
        entry.bind('<FocusOut>', on_focus_out)
        entry.bind('<Escape>', lambda e: ocultar())
        entry.bind('<Down>', lambda e: sug_list.focus_set() if estado['visible'] else None)
        sug_list.bind('<<ListboxSelect>>', elegir)
        sug_list.bind('<Return>', elegir)
        return entry

    # ── Autocompletar ───────────────────────────────────────────────
    def _cargar_autocomplete_empleados(self):
        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT EMPLEADO, CEDULA, APELLIDOS, NOMBRES FROM RPEMPLEA "
                                f"WHERE {SQL_FILTER} ORDER BY APELLIDOS")
                    self._autocomplete_empleados = cur.fetchall()
            except Exception as e:
                LOG.error("Error cargando lista para autocompletar: %s", e)
                self._autocomplete_empleados = []
        threading.Thread(target=tarea, daemon=True).start()

    def _debounce_autocomplete_change(self, *_args):
        # Filtrar sobre ~2450 empleados en cada tecla presionada bloqueaba
        # el hilo principal y hacía sentir la escritura "trabada"/con
        # letras perdidas. Se espera una pausa breve antes de filtrar.
        if getattr(self, '_autocomplete_debounce_id', None):
            self.root.after_cancel(self._autocomplete_debounce_id)
        self._autocomplete_debounce_id = self.root.after(180, self._on_autocomplete_change)

    def _on_autocomplete_change(self, *_args):
        texto = self._autocomplete_var.get().strip().lower()
        if not texto:
            self._autocomplete_ocultar()
            return
        matches = []
        for emp in self._autocomplete_empleados:
            cod, ced, ape, nom = emp
            hay = f"{cod} {ced} {ape} {nom}".lower()
            if texto in hay:
                matches.append(emp)
                if len(matches) >= 30:
                    break
        if not matches:
            self._autocomplete_ocultar()
            return
        self._autocomplete_sug_list.delete(0, 'end')
        self._autocomplete_map = {}
        for i, (cod, ced, ape, nom) in enumerate(matches):
            self._autocomplete_sug_list.insert('end', f"{cod}  |  {ape} {nom}  |  C.I. {ced}")
            self._autocomplete_map[i] = cod
        self._autocomplete_mostrar()

    def _autocomplete_mostrar(self):
        if not self._autocomplete_sug_visible:
            self._autocomplete_sug_frame.pack(fill='x', pady=(2, 0))
            self._autocomplete_sug_visible = True

    def _autocomplete_ocultar(self):
        if self._autocomplete_sug_visible:
            self._autocomplete_sug_frame.pack_forget()
            self._autocomplete_sug_visible = False

    def _autocomplete_focus_listbox(self, _event=None):
        if self._autocomplete_sug_visible and self._autocomplete_sug_list.size() > 0:
            self._autocomplete_sug_list.focus_set()
            self._autocomplete_sug_list.selection_set(0)

    def _autocomplete_select_first(self, _event=None):
        if self._autocomplete_sug_visible and self._autocomplete_sug_list.size() > 0:
            self._autocomplete_sug_list.selection_clear(0, 'end')
            self._autocomplete_sug_list.selection_set(0)
            self._autocomplete_on_click()

    def _autocomplete_on_click(self, _event=None):
        sel = self._autocomplete_sug_list.curselection()
        if not sel:
            return
        cod = self._autocomplete_map.get(sel[0])
        if cod is None:
            return
        self._autocomplete_var.set("")
        self._autocomplete_ocultar()
        self._codigo_var.set(str(cod))
        self._buscar_por_codigo()

    # ── BD: Búsquedas ───────────────────────────────────────────────
    def _buscar_por_cedula(self):
        ced = self._cedula_var.get().strip()
        if not ced:
            messagebox.showwarning("Aviso", "Ingrese una cédula")
            return
        self._status("Buscando...")
        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT * FROM RPEMPLEA WHERE CEDULA=? AND {SQL_FILTER}", (ced,))
                    emp = cur.fetchone()
                    cols = [c[0] for c in cur.description]
                self.root.after(0, lambda: self._procesar_busqueda(emp, cols, "cédula", ced))
            except Exception as e:
                self.root.after(0, lambda msg=str(e): self._status(f"Error: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()

    def _buscar_por_codigo(self):
        cod = self._codigo_var.get().strip()
        if not cod:
            messagebox.showwarning("Aviso", "Ingrese un código")
            return
        self._status("Buscando...")
        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT * FROM RPEMPLEA WHERE EMPLEADO=? AND {SQL_FILTER}", (cod,))
                    emp = cur.fetchone()
                    cols = [c[0] for c in cur.description]
                self.root.after(0, lambda: self._procesar_busqueda(emp, cols, "código", cod))
            except Exception as e:
                self.root.after(0, lambda msg=str(e): self._status(f"Error: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()

    def _procesar_busqueda(self, emp, cols, tipo, valor):
        if emp:
            self._cargar_datos_empleado(emp, cols)
            self._status(f"Empleado encontrado por {tipo}: {valor}")
        else:
            messagebox.showinfo("No encontrado", f"No se encontró empleado con {tipo}: {valor}")
            self._status(f"No encontrado por {tipo}")

    # ── Cargar datos a la UI ────────────────────────────────────────
    def _cargar_datos_empleado(self, empleado, descripcion):
        cols = [c[0] for c in descripcion] if not isinstance(descripcion[0], str) else descripcion
        datos = dict(zip(cols, empleado))

        # Datos generales
        for k, v in self._dg_vars.items():
            val = datos.get(k)
            if val is not None:
                texto = str(val) if not isinstance(val, datetime) else val.strftime("%d/%m/%Y")
                combo = self._combos_widgets.get(k)
                items = combo.cget('values') if combo else self._combo_full_values.get(k)
                if items:
                    match = self._match_combo_val(texto, items)
                    if match:
                        texto = match
                v.set(texto)
            else:
                v.set("")

        # Ingresos
        for k, v in self._ing_vars.items():
            val = datos.get(k)
            v.set(str(val) if val is not None else "")
        self._cargar_flag_vars(datos)

        # Otros
        for k, v in self._ot_vars.items():
            val = datos.get(k)
            if k in ['INCL_ROL', 'INCL_BAN']:
                sv = str(val) if val is not None else 'N'
                v.set(sv)
                self._check_states[k] = (sv == 'S')
            else:
                if val is None:
                    texto = ""
                elif isinstance(val, datetime):
                    texto = val.strftime("%d/%m/%Y")
                else:
                    texto = str(val)
                combo = self._combos_widgets.get(k)
                if combo and texto:
                    items = combo.cget('values')
                    match = self._match_combo_val(texto, items)
                    if match:
                        texto = match
                v.set(texto)
        self._actualizar_check_visual()

        # Certificados
        for k, v in self._cert_vars.items():
            val = datos.get(k)
            v.set(str(val) if val is not None else "")

        # Referencias
        for k, v in self._ref_vars.items():
            val = datos.get(k)
            if k in ['PRIMARIA', 'SECUNDARIA', 'EST_SUP', 'FZA_PUB', 'SER_MIL']:
                if isinstance(v, tk.BooleanVar):
                    v.set(bool(val) if val else False)
            else:
                v.set(str(val) if val is not None else "")

        # Observaciones (se cargan con el botón "Mostrar" en la pestaña)
        # Las observaciones se muestran en RPEMPOBSERV mediante _mostrar_obs()

        self.empleado_actual = datos
        self.datos_originales = datos.copy()
        self.datos_modificados = False
        self.modo_edicion = False

        # Auditoría
        self._auditoria_data = {
            'creado_por': datos.get('creado_por', ''),
            'fecha_crea': datos.get('fecha_crea', ''),
            'mod_por': datos.get('mod_por', ''),
            'fecha_mod': datos.get('fecha_mod', ''),
        }
        self._actualizar_label_auditoria()

        self._set_form_state('view')
        self.modo_edicion = False
        self._actualizar_label_empleado()
        self._marcar_dirty('datos', 'ingresos', 'obs', 'otros', 'cert', 'ref')

    # ── BD: Lista de empleados ──────────────────────────────────────
    def _cargar_lista(self):
        if not self.conn:
            return
        self._status("Cargando lista...")
        def tarea():
            try:
                f = self._filtro_var.get()
                where = ""
                if f == "ACTIVOS":
                    where = "AND ESTADO = 'ACT'"
                elif f == "INACTIVOS":
                    where = "AND ESTADO != 'ACT'"
                order = "ORDER BY APELLIDOS, NOMBRES" if self._orden_var.get() == "alfabetico" else "ORDER BY DEPTO, APELLIDOS, NOMBRES"
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT EMPLEADO, APELLIDOS, NOMBRES, CEDULA FROM RPEMPLEA WHERE {SQL_FILTER} {where} {order}")
                    rows = cur.fetchall()
                self.root.after(0, lambda: self._mostrar_lista(rows))
            except Exception as e:
                self.root.after(0, lambda msg=str(e): self._status(f"Error lista: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_lista(self, rows):
        self._lista_rows_cache = rows
        self._renderizar_lista()

    def _debounce_renderizar_lista(self):
        # Reconstruir el Treeview completo (hasta ~2450 filas) en cada
        # tecla presionada bloqueaba el hilo principal y hacía sentir la
        # escritura "trabada"/con letras perdidas. Se espera una pausa
        # breve antes de re-filtrar.
        if getattr(self, '_filtro_debounce_id', None):
            self.root.after_cancel(self._filtro_debounce_id)
        self._filtro_debounce_id = self.root.after(180, self._renderizar_lista)

    def _renderizar_lista(self):
        texto = self._filtro_texto_var.get().strip().lower()
        self._tree.delete(*self._tree.get_children())
        n = 0
        for r in getattr(self, '_lista_rows_cache', []):
            if texto:
                hay = f"{r[0]} {r[1] or ''} {r[2] or ''} {r[3] or ''}".lower()
                if texto not in hay:
                    continue
            self._tree.insert('', 'end', values=(r[0], (r[1] or '').upper(), (r[2] or '').upper()))
            n += 1
        total = len(getattr(self, '_lista_rows_cache', []))
        self._status(f"Lista: {n} de {total} empleados" if texto else f"Lista: {n} empleados")

    # ── Navegación lista ────────────────────────────────────────────
    def _on_tree_select(self, ev=None):
        sel = self._tree.selection()
        if not sel:
            return
        if self.datos_modificados:
            r = messagebox.askyesnocancel("Cambios sin guardar",
                                          "¿Guardar cambios antes de continuar?")
            if r is None:
                return
            elif r:
                if not self._guardar_cambios():
                    return
        item = self._tree.item(sel[0])
        self._codigo_var.set(str(item['values'][0]))
        self._buscar_por_codigo()

    def _primer_emp(self):
        ch = self._tree.get_children()
        if ch:
            self._tree.selection_set(ch[0])
            self._on_tree_select()

    def _ultimo_emp(self):
        ch = self._tree.get_children()
        if ch:
            self._tree.selection_set(ch[-1])
            self._on_tree_select()

    def _anterior_emp(self):
        sel = self._tree.selection()
        if not sel:
            return
        ch = self._tree.get_children()
        try:
            i = ch.index(sel[0])
            if i > 0:
                self._tree.selection_set(ch[i - 1])
                self._on_tree_select()
        except ValueError:
            pass

    def _siguiente_emp(self):
        sel = self._tree.selection()
        if not sel:
            return
        ch = self._tree.get_children()
        try:
            i = ch.index(sel[0])
            if i < len(ch) - 1:
                self._tree.selection_set(ch[i + 1])
                self._on_tree_select()
        except ValueError:
            pass

    # ── Nuevo ───────────────────────────────────────────────────────
    def _nuevo_empleado(self):
        for v in list(self._dg_vars.values()) + list(self._ing_vars.values()):
            v.set("")
        for k, v in self._ot_vars.items():
            if k in ['INCL_ROL', 'INCL_BAN']:
                v.set('N')
                self._check_states[k] = False
            else:
                v.set("")
        self._actualizar_check_visual()
        for v in list(self._cert_vars.values()):
            v.set("")
        for k, v in self._ref_vars.items():
            if k in ['PRIMARIA', 'SECUNDARIA', 'EST_SUP', 'FZA_PUB', 'SER_MIL']:
                v.set(False)
            else:
                v.set("")
        for v in self._flag_vars.values():
            v.set(False)
        # Limpiar frame de observaciones
        for widget in self._obs_frame.winfo_children():
            widget.destroy()
        self.empleado_actual = None
        self.datos_originales = None
        self.datos_modificados = False
        self.modo_edicion = True
        self._set_form_state('edit')
        self._status("Nuevo empleado — modo edición activo")

    # ── Modificar ───────────────────────────────────────────────────
    def _modificar_empleado(self):
        if not self.empleado_actual:
            messagebox.showwarning(
                "ATENCIÓN — Sin empleado seleccionado",
                "═══════════════════════════════════════════════\n"
                "  NO HAY UN EMPLEADO SELECCIONADO\n\n"
                "  Por favor, primero seleccione un empleado\n"
                "  de la lista o búsquelo por cédula/código\n"
                "  antes de intentar modificarlo.\n"
                "═══════════════════════════════════════════════"
            )
            return
        cod = self.empleado_actual.get('EMPLEADO', '')
        nom = f"{self.empleado_actual.get('NOMBRES', '')} {self.empleado_actual.get('APELLIDOS', '')}"
        ced = self.empleado_actual.get('CEDULA', '')
        respuesta = messagebox.askyesno(
            "CONFIRMAR ACTIVACIÓN DE MODO EDICIÓN",
            "═══════════════════════════════════════════════════════════════\n"
            "  ¿DESEA ACTIVAR EL MODO EDICIÓN PARA ESTE EMPLEADO?\n"
            "═══════════════════════════════════════════════════════════════\n\n"
            f"  • EMPLEADO:  {nom}\n"
            f"  • CÓDIGO:    {cod}\n"
            f"  • CÉDULA:    {ced}\n\n"
            "  Al activar el modo edición podrá:\n"
            "    ✓ Modificar todos los datos del empleado\n"
            "    ✓ Los cambios se aplicarán al presionar GUARDAR\n"
            "    ✓ Use CANCELAR para descartar los cambios\n\n"
            "═══════════════════════════════════════════════════════════════\n"
            "  ¿DESEA CONTINUAR CON LA ACTIVACIÓN DEL MODO EDICIÓN?\n"
            "═══════════════════════════════════════════════════════════════",
            icon='question'
        )
        if respuesta:
            self.modo_edicion = True
            self._set_form_state('edit')
            self._status(f"⚠ MODO EDICIÓN ACTIVADO — {nom}")

    # ── Eliminar ────────────────────────────────────────────────────
    def _eliminar_empleado(self):
        if not self.empleado_actual:
            messagebox.showwarning(
                "ATENCIÓN — Sin empleado seleccionado",
                "═══════════════════════════════════════════════\n"
                "  NO HAY UN EMPLEADO SELECCIONADO\n\n"
                "  Por favor, seleccione un empleado de la\n"
                "  lista o búsquelo antes de eliminarlo.\n"
                "═══════════════════════════════════════════════"
            )
            return
        cod = self.empleado_actual.get('EMPLEADO', '')
        nom = f"{self.empleado_actual.get('NOMBRES', '')} {self.empleado_actual.get('APELLIDOS', '')}"
        ced = self.empleado_actual.get('CEDULA', '')
        cargo = self.empleado_actual.get('CARGO', '')
        depto = self.empleado_actual.get('DEPTO', '')
        estado = self.empleado_actual.get('ESTADO', '')
        fecha_ing = self.empleado_actual.get('FECHA_ING', '')
        sueldo = self.empleado_actual.get('SUELDO', 0)

        # ═══ PRIMERA CONFIRMACIÓN ═══
        r1 = messagebox.askyesno(
            "⚠ ELIMINAR EMPLEADO — ADVERTENCIA INICIAL",
            "═══════════════════════════════════════════════════════════════\n"
            "  ¡ESTÁ A PUNTO DE ELIMINAR UN EMPLEADO!\n"
            "═══════════════════════════════════════════════════════════════\n\n"
            f"  EMPLEADO:  {nom}\n"
            f"  CÓDIGO:    {cod}\n"
            f"  CÉDULA:    {ced}\n\n"
            "  CONSECUENCIAS:\n"
            "  ⚠ Esta acción ELIMINARÁ PERMANENTEMENTE al empleado\n"
            "  ⚠ Se perderán TODOS sus datos e historial\n"
            "  ⚠ Esta operación NO SE PUEDE DESHACER\n\n"
            "═══════════════════════════════════════════════════════════════\n"
            "  ¿ESTÁ SEGURO DE QUE DESEA CONTINUAR?\n"
            "═══════════════════════════════════════════════════════════════",
            icon='warning'
        )
        if not r1:
            self._status("Eliminación cancelada por el usuario")
            return

        # ═══ SEGUNDA CONFIRMACIÓN — datos críticos ═══
        advertencias = []
        try:
            if sueldo and float(str(sueldo)) > 0:
                advertencias.append(f"  • SUELDO ASIGNADO: ${float(str(sueldo)):,.2f}")
        except Exception:
            pass
        if estado and str(estado) in ('ACT', 'ACTIVO'):
            advertencias.append("  • EL EMPLEADO ESTÁ ACTIVO")
        if fecha_ing:
            advertencias.append(f"  • FECHA DE INGRESO: {fecha_ing}")
        if cargo:
            advertencias.append(f"  • CARGO ASIGNADO: {cargo}")
        if depto:
            advertencias.append(f"  • DEPARTAMENTO: {depto}")

        txt_advertencias = "\n".join(advertencias) if advertencias else "  • No se detectaron datos críticos adicionales"

        r2 = messagebox.askyesno(
            "⚠ VERIFICACIÓN DE SEGURIDAD — DATOS CRÍTICOS",
            "═══════════════════════════════════════════════════════════════\n"
            "  DATOS CRÍTICOS ASOCIADOS A ESTE EMPLEADO:\n"
            "═══════════════════════════════════════════════════════════════\n\n"
            f"{txt_advertencias}\n\n"
            "  RECOMENDACIONES ANTES DE ELIMINAR:\n"
            "  • ¿Tiene un respaldo actualizado de la base de datos?\n"
            "  • ¿Ha verificado que no hay nóminas pendientes?\n"
            "  • ¿Ha consultado con el departamento de RRHH?\n"
            "  • ¿Está seguro de que no necesitará estos datos?\n\n"
            "  💡 SUGERENCIA: Considere marcar como LIQUIDADO\n"
            "     en lugar de eliminar permanentemente.\n\n"
            "═══════════════════════════════════════════════════════════════\n"
            "  ¿DESEA CONTINUAR CON LA ELIMINACIÓN?\n"
            "═══════════════════════════════════════════════════════════════",
            icon='error'
        )
        if not r2:
            self._status("Eliminación cancelada en verificación de seguridad")
            return

        # ═══ TERCERA CONFIRMACIÓN — escribir código ═══
        conf = simpledialog.askstring(
            "⚠ CONFIRMACIÓN FINAL — ESCRIBA EL CÓDIGO",
            "═══════════════════════════════════════════════════════════════\n"
            "  CONFIRMACIÓN FINAL DE ELIMINACIÓN\n"
            "═══════════════════════════════════════════════════════════════\n\n"
            f"  Para confirmar que realmente desea ELIMINAR\n"
            f"  permanentemente a:\n\n"
            f"    {nom}\n"
            f"    Código: {cod}\n\n"
            "  Escríba exactamente el código del empleado\n"
            "  en el campo de abajo para confirmar:\n\n"
            "  ⚠ Esta es su ÚLTIMA OPORTUNIDAD de cancelar.\n"
            "  Después de esto NO HABRÁ VUELTA ATRÁS.\n"
            "═══════════════════════════════════════════════════════════════",
            parent=self.root
        )
        if not conf or conf.strip() != str(cod):
            if conf:
                messagebox.showwarning(
                    "CÓDIGO INCORRECTO",
                    "═══════════════════════════════════════════════\n"
                    f"  El código ingresado no coincide.\n\n"
                    f"  Se esperaba: {cod}\n"
                    f"  Se recibió:  {conf.strip()}\n\n"
                    "  La eliminación ha sido cancelada por seguridad.\n"
                    "═══════════════════════════════════════════════"
                )
            else:
                messagebox.showinfo(
                    "ELIMINACIÓN CANCELADA",
                    "═══════════════════════════════════════════════\n"
                    "  Eliminación cancelada por el usuario.\n"
                    "═══════════════════════════════════════════════"
                )
            self._status("Eliminación cancelada — código incorrecto")
            return

        # ═══ EJECUTAR ═══
        self._status(f"Eliminando empleado {cod}...")
        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute("DELETE FROM RPEMPLEA WHERE EMPLEADO=? AND " + SQL_FILTER, (cod,))
                    if cur.rowcount == 0:
                        self.root.after(0, lambda: messagebox.showerror(
                            "ERROR AL ELIMINAR",
                            "═══════════════════════════════════════════════\n"
                            "  No se pudo eliminar el empleado.\n\n"
                            "  Posibles causas:\n"
                            "  • El empleado ya fue eliminado\n"
                            "    por otro usuario\n"
                            "  • Error de conexión a la BD\n"
                            "═══════════════════════════════════════════════"
                        ))
                        return
                    self.conn.commit()
                self.root.after(0, lambda: messagebox.showinfo(
                    "ELIMINACIÓN EXITOSA",
                    "═══════════════════════════════════════════════\n"
                    f"  Empleado ELIMINADO correctamente:\n\n"
                    f"  {nom}\n"
                    f"  Código: {cod}\n\n"
                    "  El registro ha sido eliminado\n"
                    "  permanentemente de la base de datos.\n"
                    "═══════════════════════════════════════════════"
                ))
                self.root.after(0, lambda: self._status(f"Empleado {cod} eliminado"))
                self.root.after(0, self._cargar_lista)
                self.root.after(0, self._nuevo_empleado)
            except Exception as e:
                self.root.after(0, lambda msg=str(e): messagebox.showerror(
                    "ERROR CRÍTICO",
                    "═══════════════════════════════════════════════\n"
                    f"  Error al eliminar empleado:\n\n"
                    f"  {msg}\n\n"
                    "  El empleado NO ha sido eliminado.\n"
                    "  Contacte al administrador.\n"
                    "═══════════════════════════════════════════════"
                ))
        threading.Thread(target=tarea, daemon=True).start()

    # ── Guardar ─────────────────────────────────────────────────────
    def _guardar_cambios(self):
        if not self.conn:
            messagebox.showerror(
                "ERROR — Sin conexión",
                "═══════════════════════════════════════════════\n"
                "  No hay conexión a la base de datos.\n\n"
                "  Verifique que SQL Server esté disponible\n"
                "  en 192.168.2.115 e intente nuevamente.\n"
                "═══════════════════════════════════════════════"
            )
            return False
        if not self.modo_edicion and self.empleado_actual:
            messagebox.showwarning(
                "ATENCIÓN — Modo edición desactivado",
                "═══════════════════════════════════════════════════════════════\n"
                "  No se puede guardar porque el modo edición\n"
                "  no está activo.\n\n"
                "  Presione el botón MODIFICAR para habilitar\n"
                "  la edición de datos, luego vuelva a pulsar\n"
                "  GUARDAR.\n"
                "═══════════════════════════════════════════════════════════════"
            )
            return False
        cod = self._dg_vars['EMPLEADO'].get()
        nom = self._dg_vars['NOMBRES'].get()
        ape = self._dg_vars['APELLIDOS'].get()
        ced = self._dg_vars['CEDULA'].get()
        if not cod or not ced or not nom or not ape:
            messagebox.showerror(
                "ERROR — Campos obligatorios faltantes",
                "═══════════════════════════════════════════════════════════════\n"
                "  Los siguientes campos son OBLIGATORIOS:\n"
                "═══════════════════════════════════════════════════════════════\n\n"
                f"  {'✓' if cod else '✗'} CÓDIGO DE EMPLEADO\n"
                f"  {'✓' if ced else '✗'} CÉDULA\n"
                f"  {'✓' if nom else '✗'} NOMBRES\n"
                f"  {'✓' if ape else '✗'} APELLIDOS\n\n"
                "  Complete todos los campos obligatorios\n"
                "  antes de guardar.\n"
                "═══════════════════════════════════════════════════════════════"
            )
            return False

        tipo_op = "ACTUALIZAR" if self.empleado_actual else "CREAR NUEVO"
        conf = messagebox.askyesno(
            f"⚠ CONFIRMAR — {tipo_op}",
            "═══════════════════════════════════════════════════════════════\n"
            f"  ¿ESTÁ SEGURO DE {tipo_op}?\n"
            "═══════════════════════════════════════════════════════════════\n\n"
            f"  EMPLEADO:  {nom} {ape}\n"
            f"  CÓDIGO:    {cod}\n"
            f"  CÉDULA:    {ced}\n\n"
            "  Esta acción:\n"
            f"  {'✓ Modificará los datos del empleado existente' if self.empleado_actual else '✓ Creará un nuevo registro de empleado'}\n"
            "  ✓ Los cambios son PERMANENTES en la base de datos\n"
            "  ✓ No se puede deshacer automáticamente\n\n"
            "═══════════════════════════════════════════════════════════════\n"
            "  ¿DESEA CONTINUAR?\n"
            "═══════════════════════════════════════════════════════════════",
            icon='warning'
        )
        if not conf:
            return False

        self._status("Guardando...")
        def tarea():
            try:
                datos = {}
                for k, v in self._dg_vars.items():
                    raw = v.get()
                    datos[k] = self._extraer_codigo(raw) if raw else None
                for k, v in self._ing_vars.items():
                    val = v.get()
                    if val:
                        try:
                            datos[k] = float(val)
                        except ValueError:
                            datos[k] = val
                    else:
                        datos[k] = None
                datos.update(self._flag_vars_a_datos())
                for k, v in self._ot_vars.items():
                    if k in ['INCL_ROL', 'INCL_BAN']:
                        datos[k] = v.get() if v.get() in ['S', 'N'] else 'N'
                    else:
                        raw = v.get()
                        datos[k] = self._extraer_codigo(raw) if raw else None
                for k, v in self._cert_vars.items():
                    datos[k] = v.get() if v.get() else None
                for k, v in self._ref_vars.items():
                    if k in ['PRIMARIA', 'SECUNDARIA', 'EST_SUP', 'FZA_PUB', 'SER_MIL']:
                        datos[k] = 1 if v.get() else 0
                    else:
                        datos[k] = v.get() if v.get() else None

                with self.conn_lock:
                    cur = self.conn.cursor()
                    if self.empleado_actual:
                        campos = [f"{k}=?" for k in datos if k != 'EMPLEADO']
                        vals = [datos[k] for k in datos if k != 'EMPLEADO'] + [cod]
                        cur.execute(f"UPDATE RPEMPLEA SET {', '.join(campos)} WHERE EMPLEADO=? AND {SQL_FILTER}", vals)
                    else:
                        # Verificar duplicados antes de INSERT
                        cur.execute(f"SELECT COUNT(*) FROM RPEMPLEA WHERE EMPLEADO=? AND {SQL_FILTER}", (cod,))
                        if cur.fetchone()[0] > 0:
                            self.root.after(0, lambda: messagebox.showerror("ERROR",
                                f"═══════════════════════════════════════════════\n"
                                f"  Ya existe un empleado con código {cod}.\n\n"
                                f"  Verifique que el código sea único.\n"
                                f"═══════════════════════════════════════════════"))
                            return
                        cur.execute(f"SELECT COUNT(*) FROM RPEMPLEA WHERE CEDULA=? AND {SQL_FILTER}", (ced,))
                        if cur.fetchone()[0] > 0:
                            self.root.after(0, lambda: messagebox.showerror("ERROR",
                                f"═══════════════════════════════════════════════\n"
                                f"  Ya existe un empleado con cédula {ced}.\n\n"
                                f"  Verifique que la cédula sea única.\n"
                                f"═══════════════════════════════════════════════"))
                            return
                        cols = list(datos.keys())
                        ph = ', '.join(['?'] * len(cols))
                        vals = [datos[k] for k in cols]
                        cur.execute(f"INSERT INTO RPEMPLEA ({', '.join(cols)}) VALUES ({ph})", vals)
                    self.conn.commit()

                self.root.after(0, lambda: messagebox.showinfo(
                    "OPERACIÓN EXITOSA",
                    "═══════════════════════════════════════════════\n"
                    f"  Datos guardados correctamente.\n\n"
                    f"  EMPLEADO:  {nom} {ape}\n"
                    f"  CÓDIGO:    {cod}\n"
                    "═══════════════════════════════════════════════"
                ))
                self.root.after(0, lambda: self._status(f"Datos guardados: {cod}"))
                self.root.after(0, self._cargar_lista)
                self.root.after(0, lambda: setattr(self, 'datos_modificados', False))
                self.root.after(0, lambda: setattr(self, 'modo_edicion', False))
                self.root.after(0, lambda: self._set_form_state('view'))
                self.root.after(0, self._actualizar_label_empleado)
            except Exception as e:
                self.root.after(0, lambda msg=str(e): messagebox.showerror(
                    "ERROR AL GUARDAR",
                    "═══════════════════════════════════════════════\n"
                    f"  Ocurrió un error al guardar:\n\n"
                    f"  {msg}\n\n"
                    "  Los cambios NO han sido aplicados.\n"
                    "  Intente nuevamente.\n"
                    "═══════════════════════════════════════════════"
                ))
        threading.Thread(target=tarea, daemon=True).start()
        return True

    def _cancelar_cambios(self):
        if not self.modo_edicion:
            messagebox.showinfo(
                "INFORMACIÓN",
                "═══════════════════════════════════════════════\n"
                "  No hay cambios que cancelar.\n"
                "  El modo edición no está activo.\n"
                "═══════════════════════════════════════════════"
            )
            return
        if self.datos_modificados:
            r = messagebox.askyesno(
                "⚠ CANCELAR CAMBIOS",
                "═══════════════════════════════════════════════════════════════\n"
                "  ¿ESTÁ SEGURO DE CANCELAR LOS CAMBIOS?\n"
                "═══════════════════════════════════════════════════════════════\n\n"
                "  ADVERTENCIA:\n"
                "  • Se perderán TODAS las modificaciones\n"
                "    realizadas desde la última vez que guardó\n"
                "  • Los datos volverán a su estado original\n"
                "  • Esta acción NO se puede deshacer\n\n"
                "═══════════════════════════════════════════════════════════════\n"
                "  ¿DESCARTAR TODOS LOS CAMBIOS?\n"
                "═══════════════════════════════════════════════════════════════",
                icon='warning'
            )
            if not r:
                return
        if self.empleado_actual and self.datos_originales:
            self._cargar_datos_desde_dict(self.datos_originales)
            self.datos_modificados = False
            self.modo_edicion = False
            self._set_form_state('view')
            self._status("Cambios cancelados — datos originales restaurados")
        else:
            self._nuevo_empleado()
            self.modo_edicion = False
            self._set_form_state('view')
            self._status("Cancelado")

    def _cargar_datos_desde_dict(self, datos):
        for k, v in self._dg_vars.items():
            val = datos.get(k)
            texto = str(val) if val is not None else ""
            combo = self._combos_widgets.get(k)
            items = combo.cget('values') if combo else self._combo_full_values.get(k)
            if items and texto:
                match = self._match_combo_val(texto, items)
                if match:
                    texto = match
            v.set(texto)
        for k, v in self._ing_vars.items():
            val = datos.get(k)
            v.set(str(val) if val is not None else "")
        self._cargar_flag_vars(datos)
        for k, v in self._ot_vars.items():
            val = datos.get(k)
            if k in ['INCL_ROL', 'INCL_BAN']:
                sv = str(val) if val is not None else 'N'
                v.set(sv)
                self._check_states[k] = (sv == 'S')
            else:
                texto = str(val) if val is not None else ""
                combo = self._combos_widgets.get(k)
                if combo and texto:
                    match = self._match_combo_val(texto, combo.cget('values'))
                    if match:
                        texto = match
                v.set(texto)
        self._actualizar_check_visual()
        for k, v in self._cert_vars.items():
            val = datos.get(k)
            v.set(str(val) if val is not None else "")
        for k, v in self._ref_vars.items():
            val = datos.get(k)
            if k in ['PRIMARIA', 'SECUNDARIA', 'EST_SUP', 'FZA_PUB', 'SER_MIL']:
                if isinstance(v, tk.BooleanVar):
                    v.set(bool(val) if val else False)
            else:
                v.set(str(val) if val is not None else "")
        # Las observaciones se cargan con el botón "Mostrar" en la pestaña
        # No se cargan automáticamente desde el campo OBSERV

    def _imprimir_empleado(self):
        if not self.empleado_actual:
            messagebox.showwarning(
                "ATENCIÓN",
                "═══════════════════════════════════════════════\n"
                "  No hay empleado seleccionado para imprimir.\n"
                "  Seleccione o busque un empleado primero.\n"
                "═══════════════════════════════════════════════"
            )
            return
        nom = f"{self.empleado_actual.get('NOMBRES', '')} {self.empleado_actual.get('APELLIDOS', '')}".strip()
        messagebox.showinfo(
            "IMPRESIÓN",
            "═══════════════════════════════════════════════\n"
            f"  Función de impresión no implementada.\n\n"
            f"  Empleado: {nom}\n"
            "═══════════════════════════════════════════════"
        )

    def _actualizar_label_empleado(self):
        if self.empleado_actual:
            nom = f"{self.empleado_actual.get('NOMBRES', '')} {self.empleado_actual.get('APELLIDOS', '')}".strip()
            cod = self.empleado_actual.get('EMPLEADO', '')
            self._lbl_empleado_actual.config(text=f"Empleado actual: {nom} (cód. {cod})")
        else:
            self._lbl_empleado_actual.config(text="")

    def _actualizar_label_auditoria(self):
        ad = self._auditoria_data
        parts = []
        if ad.get('creado_por'):
            parts.append(f"Creado por: {ad['creado_por']}")
        if ad.get('fecha_crea'):
            try:
                d = ad['fecha_crea']
                if isinstance(d, datetime):
                    parts.append(f"Fecha: {d.strftime('%d/%m/%Y %H:%M')}")
                else:
                    parts.append(f"Fecha: {d}")
            except Exception:
                parts.append(f"Fecha: {ad['fecha_crea']}")
        if ad.get('mod_por'):
            parts.append(f"Modificado por: {ad['mod_por']}")
        if ad.get('fecha_mod'):
            try:
                d = ad['fecha_mod']
                if isinstance(d, datetime):
                    parts.append(f"Últ. mod: {d.strftime('%d/%m/%Y %H:%M')}")
                else:
                    parts.append(f"Últ. mod: {d}")
            except Exception:
                parts.append(f"Últ. mod: {ad['fecha_mod']}")
        txt = " | ".join(parts) if parts else "Sin datos de auditoría"
        if hasattr(self, '_lbl_audit'):
            self._lbl_audit.config(text=txt)

    def _set_form_state(self, mode='view'):
        editable = (mode == 'edit')
        for key, w in self._form_widgets.items():
            if key in getattr(self, '_readonly_descs', set()):
                w.configure(state='readonly', foreground=COL_TEXT)
            elif isinstance(w, ttk.Combobox):
                w.configure(state='readonly', foreground=COL_TEXT)
            elif isinstance(w, ttk.Entry):
                w.configure(state='normal' if editable else 'readonly', foreground=COL_TEXT)
            elif isinstance(w, tk.Text):
                w.configure(state='normal' if editable else 'disabled', foreground=COL_TEXT)
            elif isinstance(w, tk.Entry):
                w.configure(state='normal' if editable else 'readonly', foreground=COL_TEXT)
        for key, w in self._check_widgets.items():
            w.configure(state='normal' if editable else 'disabled')
        if editable:
            self._status("✎ MODO EDICIÓN ACTIVO — Modifique los datos y presione GUARDAR")
        else:
            self._status("📖 MODO VISTA — Use MODIFICAR para editar los datos")

    def _marcar_modificado(self, *args):
        if self.modo_edicion and not self.datos_modificados:
            self.datos_modificados = True
            self._status("⚠ DATOS MODIFICADOS — Presione GUARDAR para aplicar")

    # ── Vista Completa ──────────────────────────────────────────────
    def _abrir_vista_completa(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Vista Completa — Todos los Empleados")
        dlg.geometry("1100x650")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.configure(bg=COL_BG)
        VistaCompletaWindow(dlg, self.conn, self.conn_lock)

    def _abrir_buscador(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Búsqueda Avanzada")
        dlg.geometry("800x550")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.configure(bg=COL_BG)
        BuscadorAvanzadoFrame(dlg, self)

    def _abrir_exportador_catalogos(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Exportar Catálogos — DBTABLAS")
        dlg.geometry("520x560")
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.configure(bg=COL_BG)
        ExportadorCatalogosWindow(dlg, self.conn, self.conn_lock)

    # ── Cierre ──────────────────────────────────────────────────────
    def _on_close(self):
        if self.datos_modificados:
            r = messagebox.askyesnocancel(
                "⚠ SALIR DEL SISTEMA",
                "═══════════════════════════════════════════════════════════════\n"
                "  Hay cambios sin guardar.\n\n"
                "  • SÍ:  Guardar cambios y salir\n"
                "  • NO:  Salir sin guardar (se perderán)\n"
                "  • CANCELAR:  Volver al sistema\n"
                "═══════════════════════════════════════════════════════════════"
            )
            if r is None:
                return
            elif r:
                self._guardar_cambios()
                self.root.update_idletasks()
        self._running = False
        if self.conn:
            try:
                self.conn.close()
            except Exception:
                pass
        try:
            self.root.destroy()
        except Exception:
            pass




# ═══════════════════════════════════════════════════════════════════
# Búsqueda Avanzada
# ═══════════════════════════════════════════════════════════════════
class BuscadorAvanzadoFrame:
    def __init__(self, window, app):
        self.window = window
        self.app = app
        self.conn = app.conn
        self._build()

    def _build(self):
        main = ttk.Frame(self.window)
        main.pack(fill='both', expand=True, padx=20, pady=20)

        tk.Label(main, text="Búsqueda Avanzada de Empleados",
                 font=FONT_TITLE, fg=COL_HEADER, bg=COL_BG).pack(pady=(0, 14))

        g = ttk.LabelFrame(main, text="CRITERIOS DE BÚSQUEDA", padding=12)
        g.pack(fill='x', pady=(0, 12))

        self._apellido_var = tk.StringVar()
        self._nombre_var = tk.StringVar()
        self._cedula_var_b = tk.StringVar()
        self._estado_var_b = tk.StringVar(value="TODOS")
        self._depto_var_b = tk.StringVar()
        self._cargo_var_b = tk.StringVar()

        tk.Label(g, text="Apellidos:", font=FONT_LABEL, bg=COL_BG).grid(row=0, column=0, sticky='e', padx=6, pady=4)
        e1 = ttk.Entry(g, textvariable=self._apellido_var, width=30)
        e1.grid(row=0, column=1, sticky='w', padx=6, pady=4)
        e1.bind('<Return>', lambda ev: self._buscar())
        tk.Label(g, text="Nombres:", font=FONT_LABEL, bg=COL_BG).grid(row=0, column=2, sticky='e', padx=(14,6), pady=4)
        e2 = ttk.Entry(g, textvariable=self._nombre_var, width=30)
        e2.grid(row=0, column=3, sticky='w', padx=6, pady=4)
        e2.bind('<Return>', lambda ev: self._buscar())

        tk.Label(g, text="Cédula:", font=FONT_LABEL, bg=COL_BG).grid(row=1, column=0, sticky='e', padx=6, pady=4)
        ttk.Entry(g, textvariable=self._cedula_var_b, width=20).grid(row=1, column=1, sticky='w', padx=6, pady=4)
        tk.Label(g, text="Estado:", font=FONT_LABEL, bg=COL_BG).grid(row=1, column=2, sticky='e', padx=(14,6), pady=4)
        ttk.Combobox(g, textvariable=self._estado_var_b, values=["TODOS", "ACTIVO", "LIQUIDADO", "SUSPENDIDO"],
                     width=12, state='readonly').grid(row=1, column=3, sticky='w', padx=6, pady=4)

        tk.Label(g, text="Departamento:", font=FONT_LABEL, bg=COL_BG).grid(row=2, column=0, sticky='e', padx=6, pady=4)
        ttk.Entry(g, textvariable=self._depto_var_b, width=12).grid(row=2, column=1, sticky='w', padx=6, pady=4)
        tk.Label(g, text="Cargo:", font=FONT_LABEL, bg=COL_BG).grid(row=2, column=2, sticky='e', padx=(14,6), pady=4)
        ttk.Entry(g, textvariable=self._cargo_var_b, width=12).grid(row=2, column=3, sticky='w', padx=6, pady=4)

        bf = ttk.Frame(g)
        bf.grid(row=3, column=0, columnspan=4, pady=(10, 0))
        ttk.Button(bf, text="🔍 BUSCAR", command=self._buscar, style='Accent.TButton').pack(side='left', padx=6)
        ttk.Button(bf, text="📋 MOSTRAR TODOS", command=self._mostrar_todos).pack(side='left', padx=6)
        ttk.Button(bf, text="✖ LIMPIAR", command=self._limpiar).pack(side='left', padx=6)
        ttk.Button(bf, text="📊 EXPORTAR EXCEL", command=self._exportar_excel).pack(side='left', padx=6)
        ttk.Separator(bf, orient='vertical').pack(side='left', fill='y', padx=10)
        self._info_label = tk.Label(bf, text="", font=FONT_LABEL,
                                     fg=COL_ACCENT, bg=COL_BG)
        self._info_label.pack(side='left', padx=10)

        res = ttk.LabelFrame(main, text="RESULTADOS", padding=6)
        res.pack(fill='both', expand=True)

        cols = ('cod', 'ape', 'nom', 'ced', 'cargo', 'cargo_nom', 'depto', 'depto_nom', 'sueldo', 'telefono', 'email', 'est')
        self._tree = ttk.Treeview(res, columns=cols, show='headings', height=14)
        heads = [
            ('cod', 'Cód.', 60, 'center'),
            ('ape', 'Apellidos', 140, 'w'),
            ('nom', 'Nombres', 140, 'w'),
            ('ced', 'Cédula', 110, 'center'),
            ('cargo', 'Cgo.', 50, 'center'),
            ('cargo_nom', 'Nombre Cargo', 110, 'w'),
            ('depto', 'Dpto.', 50, 'center'),
            ('depto_nom', 'Nombre Depto', 110, 'w'),
            ('sueldo', 'Sueldo', 100, 'e'),
            ('telefono', 'Teléfono', 110, 'w'),
            ('email', 'Email', 160, 'w'),
            ('est', 'Estado', 80, 'center'),
        ]
        for k, t, w, a in heads:
            self._tree.heading(k, text=t)
            self._tree.column(k, width=w, anchor=a)

        vsb = ttk.Scrollbar(res, orient='vertical', command=self._tree.yview)
        hsb = ttk.Scrollbar(res, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        res.grid_rowconfigure(0, weight=1)
        res.grid_columnconfigure(0, weight=1)

        self._tree.bind('<Double-1>', lambda ev: self._seleccionar())
        self._tree.bind('<Return>', lambda ev: self._seleccionar())

        pie = tk.Frame(main, bg=COL_BG)
        pie.pack(fill='x', pady=(6, 0))
        tk.Label(pie, text="Doble clic o Enter para cargar en el formulario principal",
                 font=FONT_SMALL, fg=COL_GRAY, bg=COL_BG).pack(side='left')
        ttk.Button(pie, text="CERRAR", command=self.window.destroy).pack(side='right')

    def _buscar(self):
        ap = self._apellido_var.get().strip()
        nom = self._nombre_var.get().strip()
        ced = self._cedula_var_b.get().strip()
        est = self._estado_var_b.get()
        dep = self._depto_var_b.get().strip()
        car = self._cargo_var_b.get().strip()
        if not any([ap, nom, ced, dep, car]):
            messagebox.showwarning(
                "CRITERIOS DE BÚSQUEDA",
                "═══════════════════════════════════════════════════════════════\n"
                "  Ingrese al menos un criterio de búsqueda.\n\n"
                "  Puede buscar por:\n"
                "  • Apellidos (parcial)\n"
                "  • Nombres (parcial)\n"
                "  • Cédula exacta\n"
                "  • Departamento (código)\n"
                "  • Cargo (código)\n\n"
                "  O use 'MOSTRAR TODOS' para ver el listado completo.\n"
                "═══════════════════════════════════════════════════════════════"
            )
            return
        self._info_label.config(text="Buscando...")
        self.window.update_idletasks()
        def tarea():
            try:
                q = ("SELECT EMPLEADO, APELLIDOS, NOMBRES, CEDULA, CARGO, "
                     "'' as CARGO_NOM, DEPTO, '' as DEPTO_NOM, SUELDO, "
                     "TELEFONO, emp_mail, ESTADO "
                     "FROM RPEMPLEA WHERE " + SQL_FILTER)
                params = []
                if ap:
                    q += " AND UPPER(APELLIDOS) LIKE UPPER(?)"
                    params.append(f"%{ap}%")
                if nom:
                    for p in nom.split():
                        if p.strip():
                            q += " AND UPPER(NOMBRES) LIKE UPPER(?)"
                            params.append(f"%{p.strip()}%")
                if ced:
                    q += " AND CEDULA = ?"
                    params.append(ced)
                if est == "ACTIVO":
                    q += " AND ESTADO = 'ACT'"
                elif est == "LIQUIDADO":
                    q += " AND ESTADO = 'LIQ'"
                elif est == "SUSPENDIDO":
                    q += " AND ESTADO = 'SUS'"
                if dep:
                    q += " AND DEPTO = ?"
                    params.append(dep)
                if car:
                    q += " AND CARGO = ?"
                    params.append(car)
                q += " ORDER BY APELLIDOS, NOMBRES"

                nombres_cargo = {}
                nombres_depto = {}
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(q, params)
                    rows = cur.fetchall()

                    # Obtener nombres descriptivos para cargos y deptos
                    try:
                        cur2 = self.conn.cursor()
                        cur2.execute("SELECT CODIGO, NOMBRE FROM DBTABLAS WHERE TIPO='FNC' AND CODEMP='10'")
                        for r in cur2.fetchall():
                            nombres_cargo[str(r[0]).strip()] = r[1]
                        cur2.execute("SELECT CODIGO, NOMBRE FROM DBTABLAS WHERE TIPO='DPT' AND CODEMP='10'")
                        for r in cur2.fetchall():
                            nombres_depto[str(r[0]).strip()] = r[1]
                    except Exception as e:
                        LOG.error("Error cargando nombres de cargo/depto en busqueda avanzada: %s", e)

                resultados = []
                for r in rows:
                    r = list(r)
                    idx_cargo = 4
                    idx_depto = 6
                    r[5] = nombres_cargo.get(str(r[idx_cargo]).strip(), '') if r[idx_cargo] else ''
                    r[7] = nombres_depto.get(str(r[idx_depto]).strip(), '') if r[idx_depto] else ''
                    try:
                        if r[8]:
                            r[8] = f"${float(r[8]):,.2f}"
                    except Exception:
                        pass
                    resultados.append(r)

                self.window.after(0, lambda: self._mostrar_resultados(resultados))
            except Exception as e:
                self.window.after(0, lambda: self._info_label.config(text="Error en búsqueda"))
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_resultados(self, rows):
        self._tree.delete(*self._tree.get_children())
        for r in rows:
            self._tree.insert('', 'end', values=(
                r[0],
                (r[1] or '').upper(),
                (r[2] or '').upper(),
                r[3] or '',
                r[4] or '',
                r[5] or '',
                r[6] or '',
                r[7] or '',
                r[8] or '',
                r[9] or '',
                r[10] or '',
                r[11] or '',
            ))
        self._info_label.config(text=f"Encontrados: {len(rows)} empleados")

    def _mostrar_todos(self):
        self._info_label.config(text="Cargando todos...")
        self.window.update_idletasks()
        def tarea():
            try:
                nombres_cargo = {}
                nombres_depto = {}
                with self.app.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(
                        "SELECT EMPLEADO, APELLIDOS, NOMBRES, CEDULA, CARGO, "
                        "'' as CARGO_NOM, DEPTO, '' as DEPTO_NOM, SUELDO, "
                        "TELEFONO, emp_mail, ESTADO "
                        f"FROM RPEMPLEA WHERE {SQL_FILTER} ORDER BY APELLIDOS, NOMBRES"
                    )
                    rows = cur.fetchall()
                    try:
                        cur2 = self.conn.cursor()
                        cur2.execute("SELECT CODIGO, NOMBRE FROM DBTABLAS WHERE TIPO='FNC' AND CODEMP='10'")
                        for r in cur2.fetchall():
                            nombres_cargo[str(r[0]).strip()] = r[1]
                        cur2.execute("SELECT CODIGO, NOMBRE FROM DBTABLAS WHERE TIPO='DPT' AND CODEMP='10'")
                        for r in cur2.fetchall():
                            nombres_depto[str(r[0]).strip()] = r[1]
                    except Exception as e:
                        LOG.error("Error cargando nombres de cargo/depto en busqueda avanzada: %s", e)
                resultados = []
                for r in rows:
                    r = list(r)
                    r[5] = nombres_cargo.get(str(r[4]).strip(), '') if r[4] else ''
                    r[7] = nombres_depto.get(str(r[6]).strip(), '') if r[6] else ''
                    try:
                        if r[8]:
                            r[8] = f"${float(r[8]):,.2f}"
                    except Exception:
                        pass
                    resultados.append(r)
                self.window.after(0, lambda: self._mostrar_resultados(resultados))
            except Exception as e:
                self.window.after(0, lambda: self._info_label.config(text="Error"))
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _limpiar(self):
        self._apellido_var.set("")
        self._nombre_var.set("")
        self._cedula_var_b.set("")
        self._estado_var_b.set("TODOS")
        self._depto_var_b.set("")
        self._cargo_var_b.set("")
        self._tree.delete(*self._tree.get_children())
        self._info_label.config(text="")

    def _seleccionar(self):
        sel = self._tree.selection()
        if not sel:
            messagebox.showwarning(
                "SELECCIÓN",
                "═══════════════════════════════════════════════\n"
                "  Seleccione un empleado de la lista primero.\n"
                "═══════════════════════════════════════════════"
            )
            return
        item = self._tree.item(sel[0])
        cod = item['values'][0]
        nom = f"{item['values'][1]} {item['values'][2]}".strip()
        if messagebox.askyesno(
            "CONFIRMAR SELECCIÓN",
            "═══════════════════════════════════════════════\n"
            f"  ¿Cargar este empleado?\n\n"
            f"  {nom}\n"
            f"  Código: {cod}\n"
            "═══════════════════════════════════════════════"
        ):
            self.app._codigo_var.set(str(cod))
            self.window.destroy()
            self.app._buscar_por_codigo()

    def _exportar_excel(self):
        items = self._tree.get_children()
        if not items:
            messagebox.showwarning(
                "EXPORTAR",
                "═══════════════════════════════════════════════\n"
                "  No hay datos para exportar.\n"
                "  Realice una búsqueda primero.\n"
                "═══════════════════════════════════════════════"
            )
            return
        def tarea():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = Workbook()
                ws = wb.active
                ws.title = "EMPLEADOS"
                headers = ['Código', 'Apellidos', 'Nombres', 'Cédula', 'Cargo',
                           'Nombre Cargo', 'Depto', 'Nombre Depto',
                           'Sueldo', 'Teléfono', 'Email', 'Estado']
                hf = Font(bold=True, color="FFFFFF")
                hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font = hf
                    c.fill = hfill
                    c.alignment = Alignment(horizontal='center')
                for ri, item in enumerate(items, 2):
                    vals = self._tree.item(item)['values']
                    for ci, v in enumerate(vals, 1):
                        ws.cell(row=ri, column=ci, value=v)
                for ci in range(1, len(headers) + 1):
                    ws.column_dimensions[chr(64 + ci)].width = 14
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fn = f"BUSQUEDA_EMPLEADOS_{ts}.xlsx"
                wb.save(fn)
                self.window.after(0, lambda: messagebox.showinfo(
                    "EXPORTADO",
                    "═══════════════════════════════════════════════\n"
                    f"  Archivo creado:\n  {fn}\n"
                    f"  Filas: {len(items)}\n"
                    "═══════════════════════════════════════════════"
                ))
            except ImportError:
                self.window.after(0, lambda: messagebox.showerror(
                    "ERROR", "Requiere openpyxl:\npip install openpyxl"))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════
# Vista Completa — Todos los Empleados
# ═══════════════════════════════════════════════════════════════════
class VistaCompletaWindow:
    def __init__(self, parent, conn, conn_lock):
        self.conn = conn
        self.conn_lock = conn_lock
        self.window = parent
        self._crear_interfaz()
        self._cargar_empleados()

    def _crear_interfaz(self):
        frame = ttk.Frame(self.window)
        frame.pack(fill='both', expand=True, padx=8, pady=8)

        # Controles
        ctrl = tk.Frame(frame, bg=COL_BG)
        ctrl.pack(fill='x', pady=(0, 6))
        tk.Label(ctrl, text="Búsqueda:", font=FONT_LABEL, bg=COL_BG).pack(side='left')
        self._busq_var = tk.StringVar()
        e = ttk.Entry(ctrl, textvariable=self._busq_var, width=30)
        e.pack(side='left', padx=(6, 10))
        e.bind('<KeyRelease>', lambda ev: self._filtrar_busqueda())
        ttk.Button(ctrl, text="Refrescar", command=self._cargar_empleados).pack(side='left', padx=2)
        ttk.Button(ctrl, text="Exportar Excel", command=self._exportar_excel).pack(side='left', padx=2)

        # Treeview
        cols = ('cod', 'ced', 'ape', 'nom', 'cargo', 'depto', 'est', 'sueldo', 'tel', 'email')
        self._tree = ttk.Treeview(frame, columns=cols, show='headings', height=18)
        self._tree.heading('cod', text='Cód.')
        self._tree.heading('ced', text='Cédula')
        self._tree.heading('ape', text='Apellidos')
        self._tree.heading('nom', text='Nombres')
        self._tree.heading('cargo', text='Cargo')
        self._tree.heading('depto', text='Depto')
        self._tree.heading('est', text='Est.')
        self._tree.heading('sueldo', text='Sueldo')
        self._tree.heading('tel', text='Teléfono')
        self._tree.heading('email', text='Email')

        self._tree.column('cod', width=50, anchor='c')
        self._tree.column('ced', width=80, anchor='c')
        self._tree.column('ape', width=100)
        self._tree.column('nom', width=100)
        self._tree.column('cargo', width=80)
        self._tree.column('depto', width=70)
        self._tree.column('est', width=50, anchor='c')
        self._tree.column('sueldo', width=80, anchor='e')
        self._tree.column('tel', width=90)
        self._tree.column('email', width=120)

        vsb = ttk.Scrollbar(frame, orient='vertical', command=self._tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        # Estadísticas
        self._stats_lbl = tk.Label(frame, text="Cargando...", font=FONT_SMALL,
                                    fg=COL_TEXT, bg=COL_BG, anchor='w')
        self._stats_lbl.pack(fill='x', pady=(6, 0))

    def _cargar_empleados(self):
        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT EMPLEADO, CEDULA, APELLIDOS, NOMBRES, CARGO, DEPTO, ESTADO, SUELDO, TELEFONO, emp_mail "
                               f"FROM RPEMPLEA WHERE {SQL_FILTER} ORDER BY APELLIDOS, NOMBRES")
                    rows = cur.fetchall()
                self.window.after(0, lambda: self._mostrar_empleados(rows))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_empleados(self, rows):
        self._tree.delete(*self._tree.get_children())
        total, activos, nomina = 0, 0, 0
        for r in rows:
            cod, ced, ape, nom, cargo, depto, est, sueldo, tel, mail = r
            sueldo_fmt = f"${float(sueldo):,.2f}" if sueldo else "$0.00"
            try:
                nomina += float(sueldo) if sueldo else 0
            except Exception:
                pass
            self._tree.insert('', 'end', values=(cod, ced or '', (ape or '').upper(),
                                                 (nom or '').upper(), cargo or '', depto or '',
                                                 est or '', sueldo_fmt, tel or '', mail or ''))
            total += 1
            if est and est.upper().startswith('ACT'):
                activos += 1
        inactivos = total - activos
        promedio = nomina / activos if activos > 0 else 0
        self._stats_lbl.config(text=f"Total: {total}  |  Activos: {activos}  |  Inactivos: {inactivos}  |  "
                                    f"Nómina: ${nomina:,.2f}  |  Promedio: ${promedio:,.2f}")

    def _filtrar_busqueda(self):
        termino = self._busq_var.get().lower()
        for item in self._tree.get_children():
            valores = self._tree.item(item)['values']
            mostrar = any(termino in str(v).lower() for v in valores)
            self._tree.reattach(item, '', 'end') if mostrar else self._tree.detach(item)

    def _exportar_excel(self):
        items = self._tree.get_children()
        if not items:
            messagebox.showwarning("EXPORTAR", "No hay datos para exportar.")
            return
        def tarea():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = Workbook()
                ws = wb.active
                ws.title = "EMPLEADOS"
                headers = ['Cód.', 'Cédula', 'Apellidos', 'Nombres', 'Cargo', 'Depto',
                          'Est.', 'Sueldo', 'Teléfono', 'Email']
                hf = Font(bold=True, color="FFFFFF")
                hfill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=ci, value=h)
                    c.font, c.fill = hf, hfill
                    c.alignment = Alignment(horizontal='center')
                for ri, item in enumerate(items, 2):
                    for ci, v in enumerate(self._tree.item(item)['values'], 1):
                        ws.cell(row=ri, column=ci, value=v)
                for ci in range(1, len(headers) + 1):
                    ws.column_dimensions[chr(64 + ci)].width = 14
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                fn = f"EMPLEADOS_COMPLETO_{ts}.xlsx"
                wb.save(fn)
                self.window.after(0, lambda: messagebox.showinfo("EXPORTADO",
                    f"Archivo: {fn}\nFilas: {len(items)}"))
            except ImportError:
                self.window.after(0, lambda: messagebox.showerror("ERROR", "Requiere openpyxl:\npip install openpyxl"))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", msg))
        threading.Thread(target=tarea, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════
# Exportador de Catálogos (DBTABLAS) — adaptado de SACAR_TABLAS_SEC.pyw
# ═══════════════════════════════════════════════════════════════════
class ExportadorCatalogosWindow:
    """Exporta tipos de DBTABLAS (CODEMP='10') a Excel, una hoja por tipo
    más una hoja 'TODOS'. Reutiliza la conexión ya abierta de la app
    (self.conn) en vez de crear una nueva — el script original tenía un
    connection string roto (SERVER=SERVER\\server, placeholder sin
    reemplazar)."""

    def __init__(self, parent, conn, conn_lock):
        self.window = parent
        self.conn = conn
        self.conn_lock = conn_lock
        self.check_vars = {}
        self._crear_interfaz()
        self._cargar_tipos()

    def _crear_interfaz(self):
        frame = ttk.Frame(self.window)
        frame.pack(fill='both', expand=True, padx=14, pady=14)

        tk.Label(frame, text="📤 Exportar Catálogos", font=FONT_TITLE,
                 fg=COL_HEADER, bg=COL_BG).pack(anchor='w')
        tk.Label(frame, text="Tipos de DBTABLAS filtrados por CODEMP='10'",
                 font=FONT_SMALL, fg=COL_GRAY, bg=COL_BG).pack(anchor='w', pady=(0, 10))

        tipos_frame = ttk.LabelFrame(frame, text="Seleccione los TIPOS a exportar", padding=8)
        tipos_frame.pack(fill='both', expand=True, pady=(0, 10))

        self._catalogos_canvas = canvas = tk.Canvas(tipos_frame, bg=COL_BG, highlightthickness=0)
        vsb = ttk.Scrollbar(tipos_frame, orient='vertical', command=canvas.yview)
        self._frame_checks = tk.Frame(canvas, bg=COL_BG)
        self._frame_checks.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=self._frame_checks, anchor='nw')
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        _bind_scroll_rueda(canvas, canvas)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=(0, 8))
        ttk.Button(btn_frame, text="Solo SEC y DPT", command=self._solo_sec_dpt).pack(side='left', padx=(0, 4))
        ttk.Button(btn_frame, text="Todos", command=self._seleccionar_todos).pack(side='left', padx=4)
        ttk.Button(btn_frame, text="Ninguno", command=self._deseleccionar_todos).pack(side='left', padx=4)

        ttk.Button(frame, text="⬇ EXPORTAR A EXCEL", command=self._exportar,
                   style='Accent.TButton').pack(fill='x', pady=(0, 8))

        self.status = tk.StringVar(value="Cargando tipos...")
        tk.Label(frame, textvariable=self.status, font=FONT_SMALL,
                 fg=COL_GRAY, bg=COL_BG, anchor='w').pack(fill='x')

    def _cargar_tipos(self):
        def tarea():
            try:
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute("SELECT DISTINCT TIPO FROM DBTABLAS WHERE CODEMP = '10' ORDER BY TIPO")
                    tipos = [r[0] for r in cur.fetchall()]
                self.window.after(0, lambda: self._mostrar_tipos(tipos))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): self.status.set(f"Error: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()

    def _mostrar_tipos(self, tipos):
        tipos_defecto = ('SEC', 'DPT')
        for i, tipo in enumerate(tipos):
            var = tk.BooleanVar(value=(tipo in tipos_defecto))
            self.check_vars[tipo] = var
            cb = tk.Checkbutton(self._frame_checks, text=tipo, variable=var, font=FONT_DEFAULT,
                                 bg=COL_BG, fg=COL_TEXT, selectcolor=COL_ACCENT,
                                 activebackground=COL_BG, activeforeground=COL_TEXT)
            cb.grid(row=i // 3, column=i % 3, sticky='w', padx=10, pady=2)
        self.status.set(f"Tipos cargados: {len(tipos)}")
        _bind_scroll_rueda(self._frame_checks, self._catalogos_canvas)

    def _solo_sec_dpt(self):
        for tipo, var in self.check_vars.items():
            var.set(tipo in ('SEC', 'DPT'))

    def _seleccionar_todos(self):
        for var in self.check_vars.values():
            var.set(True)

    def _deseleccionar_todos(self):
        for var in self.check_vars.values():
            var.set(False)

    def _exportar(self):
        tipos_sel = [t for t, v in self.check_vars.items() if v.get()]
        if not tipos_sel:
            messagebox.showwarning("Aviso", "Seleccione al menos un TIPO")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        tipos_txt = "_".join(tipos_sel[:3]) + ("_etc" if len(tipos_sel) > 3 else "")
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"DBTABLAS_{tipos_txt}_{ts}.xlsx")
        if not archivo:
            return

        self.status.set("Exportando...")

        def tarea():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                tipos_str = ",".join(f"'{t}'" for t in tipos_sel)
                with self.conn_lock:
                    cur = self.conn.cursor()
                    cur.execute(f"SELECT TIPO, CODIGO, NOMBRE, FACTOR, T_C, T_P FROM DBTABLAS "
                                f"WHERE CODEMP='10' AND TIPO IN ({tipos_str}) ORDER BY TIPO, CODIGO")
                    cols = [c[0] for c in cur.description]
                    rows = cur.fetchall()

                if not rows:
                    self.window.after(0, lambda: messagebox.showwarning("Aviso", "No se encontraron datos"))
                    self.window.after(0, lambda: self.status.set("Sin datos"))
                    return

                wb = Workbook()
                wb.remove(wb.active)
                por_tipo = {}
                for r in rows:
                    por_tipo.setdefault(r[0], []).append(r)

                def escribir_hoja(ws, filas):
                    for ci, h in enumerate(cols, 1):
                        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)
                    for ri, r in enumerate(filas, 2):
                        for ci, v in enumerate(r, 1):
                            ws.cell(row=ri, column=ci, value=v)

                for tipo in tipos_sel:
                    filas_tipo = por_tipo.get(tipo)
                    if filas_tipo:
                        escribir_hoja(wb.create_sheet(tipo[:31]), filas_tipo)
                escribir_hoja(wb.create_sheet('TODOS'), rows)
                wb.save(archivo)

                n = len(rows)
                self.window.after(0, lambda: self.status.set(f"Exportado: {n} registros"))
                self.window.after(0, lambda: messagebox.showinfo(
                    "Éxito", f"Archivo creado:\n{archivo}\n\nRegistros: {n}\nTipos: {', '.join(tipos_sel)}"))
                self.window.after(0, lambda: os.startfile(os.path.dirname(archivo) or '.'))
            except ImportError:
                self.window.after(0, lambda: messagebox.showerror("Error", "Requiere openpyxl:\npip install openpyxl"))
            except Exception as e:
                self.window.after(0, lambda msg=str(e): messagebox.showerror("Error", f"Error al exportar: {msg}"))
                self.window.after(0, lambda msg=str(e): self.status.set(f"Error: {msg}"))
        threading.Thread(target=tarea, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    try:
        ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo_insevig.ico')
        if os.path.exists(ico):
            root.iconbitmap(default=ico)
    except Exception:
        pass
    app = SistemaGestionEmpleados10(root)
    root.mainloop()


if __name__ == "__main__":
    main()
