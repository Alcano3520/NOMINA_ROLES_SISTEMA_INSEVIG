"""Estado del envío de roles por correo (Fase 5). Reemplaza Outlook COM."""

from __future__ import annotations

import asyncio
import datetime as dt
import io

import reflex as rx

from core.db.health import FUENTE_SQLSERVER
from core.jobs.runner import JobRunner, get_runner, leer_job
from insevig_web.states.auth_state import AuthState
from insevig_web.states.datasource_state import DataSourceState

_TERMINALES = {"ok", "error", "cancelado"}


class EnvioState(rx.State):
    periodo: str = ""
    intervalo: str = "2"
    cc: str = ""
    destinatarios: list[dict] = []  # {'empleado','nombre','cedula','email'}
    errores: list[str] = []

    job: int = 0
    status: str = ""
    msg: str = ""

    # ── Plantilla del correo ───────────────────────────────────────────
    plantilla_asunto: str = ""
    plantilla_html: str = ""
    plantilla_msg: str = ""
    preview_html: str = ""

    @rx.event
    def on_load(self):
        if not self.periodo:
            self.periodo = dt.date.today().strftime("%Y-%m")
        if not self.plantilla_html:
            yield EnvioState.cargar_plantilla

    @rx.event
    async def cargar_plantilla(self):
        from core.parametros import get_email_plantilla

        p = await asyncio.to_thread(get_email_plantilla)
        self.plantilla_asunto, self.plantilla_html = p["asunto"], p["html"]

    @rx.event
    def set_plantilla_asunto(self, v: str):
        self.plantilla_asunto = v

    @rx.event
    def set_plantilla_html(self, v: str):
        self.plantilla_html = v

    @rx.event
    async def guardar_plantilla(self):
        auth = await self.get_state(AuthState)
        if "roles:enviar_email" not in auth.permisos_flat:
            return rx.toast.error("Sin permiso.")
        from core.parametros import set_email_plantilla

        await asyncio.to_thread(
            set_email_plantilla, self.plantilla_asunto, self.plantilla_html
        )
        self.plantilla_msg = "Plantilla guardada."
        return rx.toast.success("Plantilla guardada.")

    @rx.event
    def previsualizar(self):
        from core.email.service import render_plantilla

        d = self.destinatarios[0] if self.destinatarios else {}
        periodo = self.periodo or dt.date.today().strftime("%Y-%m")
        anio, _, mes = periodo.partition("-")
        ctx = {
            "StrNombres": d.get("nombre") or "Juan Pérez",
            "StrCedula": d.get("cedula") or "0912345678",
            "StrEmpleado": d.get("empleado") or "1234",
            "mes": mes or "01", "anio": anio, "año": anio,
        }
        try:
            asunto = render_plantilla(self.plantilla_asunto, ctx)
            cuerpo = render_plantilla(self.plantilla_html, ctx)
        except Exception as e:  # noqa: BLE001
            self.preview_html = ""
            self.plantilla_msg = f"Error en la plantilla: {e}"
            return
        self.plantilla_msg = ""
        self.preview_html = f"<p style='color:#555'><b>Asunto:</b> {asunto}</p><hr>{cuerpo}"

    @rx.event
    def set_periodo(self, v: str):
        self.periodo = v.strip()

    @rx.event
    def set_intervalo(self, v: str):
        self.intervalo = v

    @rx.event
    def set_cc(self, v: str):
        self.cc = v

    @rx.event
    async def subir_lista(self, files: list[rx.UploadFile]):
        if not files:
            return
        import openpyxl

        datos = await files[0].read()
        wb = openpyxl.load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            self.errores = ["Archivo vacío."]
            return
        hdr = [str(h).strip().lower() if h else "" for h in filas[0]]

        def idx(*names):
            for n in names:
                if n in hdr:
                    return hdr.index(n)
            return -1

        i_emp, i_nom, i_ced, i_mail = (
            idx("empleado", "codigo"),
            idx("nombre", "apellidos_nombres", "nombres"),
            idx("cedula"),
            idx("email", "emp_mail", "correo"),
        )
        if i_mail < 0:
            self.errores = ["Falta una columna de email (email / emp_mail / correo)."]
            return
        dest, errs = [], []
        for n, f in enumerate(filas[1:], 2):
            mail = f[i_mail] if i_mail < len(f) else None
            if not mail or "@" not in str(mail):
                errs.append(f"Fila {n}: email inválido.")
                continue
            dest.append(
                {
                    "empleado": str(f[i_emp]).strip() if i_emp >= 0 and i_emp < len(f) and f[i_emp] else "",
                    "nombre": str(f[i_nom]).strip() if i_nom >= 0 and i_nom < len(f) and f[i_nom] else "",
                    "cedula": str(f[i_ced]).strip() if i_ced >= 0 and i_ced < len(f) and f[i_ced] else "",
                    "email": str(mail).strip(),
                }
            )
        self.destinatarios = dest[:1000]
        self.errores = errs[:50]

    @rx.event
    async def enviar(self):
        auth = await self.get_state(AuthState)
        if "roles:enviar_email" not in auth.permisos_flat:
            return rx.toast.error("Sin permiso.")
        if not self.destinatarios:
            return rx.toast.error("Sube primero la lista.")
        ds = await self.get_state(DataSourceState)
        fuente = ds.fuente_por_modulo.get("roles", FUENTE_SQLSERVER)
        periodo = self.periodo or dt.date.today().strftime("%Y-%m")
        anio, mes = periodo.split("-")
        dest = list(self.destinatarios)
        intervalo = float(self.intervalo or 2)
        cc = [c.strip() for c in self.cc.split(",") if c.strip()]
        usuario = auth.username
        pl_html, pl_asunto = self.plantilla_html, self.plantilla_asunto

        def _fn(ctx):
            from core.datos.service import datos_empleado
            from core.email.envio_lote import job_enviar_roles
            from core.pdf.rol_pago import OpcionesRol, rol_pago_pdf

            ctx.progreso(0, len(dest), "Generando PDFs…")
            preparados = []
            for d in dest:
                ident = d["empleado"] or d["cedula"] or d["nombre"]
                emp = datos_empleado(periodo, ident, fuente)
                pdf = rol_pago_pdf(emp, OpcionesRol(fecha_desde=f"01/{mes}/{anio}", fecha_hasta=f"28/{mes}/{anio}")) if emp else b""
                preparados.append({**d, "pdf": pdf})
            job_enviar_roles(
                ctx, preparados, mes=mes, anio=anio, intervalo_seg=intervalo, cc=cc,
                html_plantilla=pl_html, asunto=pl_asunto,
            )

        self.job = get_runner().encolar(
            "envio_roles", {"periodo": periodo, "n": len(dest)}, creado_por=usuario, fn=_fn
        )
        self.status = "pendiente"
        return EnvioState.vigilar

    @rx.event(background=True)
    async def vigilar(self):
        for _ in range(7200):
            async with self:
                jid = self.job
            j = leer_job(jid)
            if j is None:
                return
            async with self:
                self.status = j.status
                self.msg = j.message
            if j.status in _TERMINALES:
                return
            await asyncio.sleep(1)

    # ── Log de envíos ──────────────────────────────────────────────────
    log: list[dict] = []

    @rx.event
    async def cargar_log(self):
        jid = self.job
        if not jid:
            return

        def _leer():
            import sqlmodel

            from core.db import appdb
            from core.db.models import EmailSendLog

            with appdb.session() as s:
                filas = s.exec(
                    sqlmodel.select(EmailSendLog).where(EmailSendLog.job_id == jid)
                ).all()
            return [
                {"empleado": f.employee_code, "email": f.email, "estado": f.status,
                 "error": f.error[:120], "enviado": str(f.sent_at or "")[:19]}
                for f in filas
            ]

        self.log = await asyncio.to_thread(_leer)

    @rx.event
    def cancelar(self):
        if self.job:
            JobRunner.cancelar(self.job)
