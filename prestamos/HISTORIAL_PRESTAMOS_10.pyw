import os, sys

# TLS 1.0 para SQL Server 2008 R2 (debe ir antes de import pyodbc)
def _bundle_dir():
    """Carpeta donde están los archivos bundleados (sys._MEIPASS en EXE, carpeta del script en dev)."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def _app_dir():
    """Carpeta junto al EXE (o junto al .pyw en dev). Aquí van configs que persisten."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

if os.name != 'nt':
    os.environ.setdefault('OPENSSL_CONF', os.path.join(_bundle_dir(), 'openssl_legacy.cnf'))
else:
    os.environ.setdefault('OPENSSL_CONF', os.path.join(_bundle_dir(), 'openssl_legacy.cnf'))

# Ícono en la barra de tareas de Windows (debe ir ANTES de crear cualquier ventana)
try:
    import ctypes
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('insevig.prestamos.1.0')
except Exception:
    pass

import json
import pyodbc
import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import locale
import subprocess
import tempfile
import time
import pandas as pd
import warnings

# Suprimir warning específico de pandas con pyodbc
warnings.filterwarnings('ignore', message='.*pandas only supports SQLAlchemy.*')

class ConsultorPrestamos:
    def __init__(self):
        # Configurar locale para formato de moneda
        try:
            locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
        except:
            try:
                locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')
            except:
                pass
        
        # Parámetros de conexión SQL Server
        self.server = 'SERVER\\server'
        self.database = 'insevig'
        self.username = 'sa'
        self.password = 'puntosoft123*'
        
        # Ruta de la base de datos SQLite
        self.sqlite_path = r"\\server\Respaldo 2017\Base\Saldo_prestamos_driver.db"
        
        # Lista para almacenar datos completos y sin filtrar
        self.datos_completos = []
        self.todos_los_movimientos_originales = []
        self.empleado_actual = None
        
        self.setup_gui()
    
    def conectar_bd(self):
        """Establece conexión con la base de datos SQL Server"""
        try:
            conn_str = (
                f'DRIVER={{ODBC Driver 17 for SQL Server}};'
                f'SERVER={self.server};'
                f'DATABASE={self.database};'
                f'UID={self.username};'
                f'PWD={self.password};'
                f'Encrypt=No;'
                f'TrustServerCertificate=yes;'
            )
            return pyodbc.connect(conn_str)
        except Exception as e:
            messagebox.showerror("Error de Conexión SQL Server", f"No se pudo conectar a la base de datos:\n{e}")
            return None
    
    def conectar_sqlite(self):
        """Establece conexión con la base de datos SQLite"""
        try:
            ruta = self._resolver_ruta_sqlite()
            if not os.path.exists(ruta):
                messagebox.showwarning("Advertencia", f"No se encontró la base de datos SQLite en:\n{ruta}")
                return None
            return sqlite3.connect(ruta)
        except Exception as e:
            messagebox.showerror("Error de Conexión SQLite", f"No se pudo conectar a SQLite:\n{e}")
            return None

    def _resolver_ruta_sqlite(self):
        """Resuelve la ruta de la BD SQLite según la plataforma"""
        if sys.platform == 'win32':
            return self.sqlite_path

        ruta_normalizada = self.sqlite_path.replace('\\', '/')
        if os.path.exists(ruta_normalizada):
            return ruta_normalizada

        ruta_montaje = '/mnt/server/Base/Saldo_prestamos_driver.db'
        if os.path.exists(ruta_montaje):
            return ruta_montaje

        try:
            cache = os.path.join(tempfile.gettempdir(), 'Saldo_prestamos_driver_cache.db')
            if not os.path.exists(cache) or (time.time() - os.path.getmtime(cache)) > 3600:
                subprocess.run(
                    ['smbclient', '//192.168.2.115/Respaldo 2017', '-N',
                     '-c', f'cd Base; get Saldo_prestamos_driver.db {cache}'],
                    capture_output=True, timeout=30, check=True
                )
            if os.path.exists(cache):
                return cache
        except Exception:
            pass

        return ruta_normalizada
    
    def exportar_saldos_prestamos_excel(self):
        """
        Exporta los saldos de préstamos a Excel igual que el programa SALDO_PRESTAMOS.pyw
        """
        try:
            # Mostrar progreso
            ventana_progreso = tk.Toplevel(self.root)
            ventana_progreso.title("Exportando Saldos...")
            ventana_progreso.geometry("400x150")
            ventana_progreso.resizable(False, False)
            ventana_progreso.transient(self.root)
            ventana_progreso.grab_set()
            
            # Centrar ventana
            ventana_progreso.update_idletasks()
            x = (ventana_progreso.winfo_screenwidth() // 2) - (400 // 2)
            y = (ventana_progreso.winfo_screenheight() // 2) - (150 // 2)
            ventana_progreso.geometry(f"400x150+{x}+{y}")
            
            frame_progreso = ttk.Frame(ventana_progreso, padding="20")
            frame_progreso.pack(fill="both", expand=True)
            
            ttk.Label(frame_progreso, text="💰 Exportando Saldos de Préstamos", 
                     font=("Segoe UI", 12, "bold")).pack(pady=(0, 10))
            
            lbl_estado = ttk.Label(frame_progreso, text="Conectando a la base de datos...", 
                                  font=("Segoe UI", 10))
            lbl_estado.pack(pady=5)
            
            progress_bar = ttk.Progressbar(frame_progreso, mode='indeterminate')
            progress_bar.pack(fill="x", pady=10)
            progress_bar.start()
            
            ventana_progreso.update()
            
            # Establecer conexión
            lbl_estado.config(text="Estableciendo conexión...")
            ventana_progreso.update()
            
            conn = self.conectar_bd()
            if not conn:
                ventana_progreso.destroy()
                return
            
            # Ejecutar consulta
            lbl_estado.config(text="Ejecutando consulta para extraer saldos...")
            ventana_progreso.update()
            
            query = """
            SELECT 
                i.[FECHA], 
                i.[EMPLEADO], 
                e.[APELLIDOS] + ' ' + e.[NOMBRES] AS NOMBRE_COMPLETO,
                e.[CEDULA],
                i.[VALOR], 
                i.[OBSERV]
            FROM [insevig].[dbo].[RPINGDES] i
            LEFT JOIN [insevig].[dbo].[RPEMPLEA] e ON i.[EMPLEADO] = e.[EMPLEADO]
            WHERE i.[CLASE] = 205
            """
            
            df = pd.read_sql(query, conn)
            conn.close()
            
            # Formatear CEDULA a string de 10 dígitos
            if 'CEDULA' in df.columns:
                df['CEDULA'] = df['CEDULA'].apply(lambda x: self.formatear_cedula(x))
            
            if df.empty:
                progress_bar.stop()
                ventana_progreso.destroy()
                messagebox.showinfo("Sin Datos", "No se encontraron saldos de préstamos (CLASE = 205)")
                return
            
            # Crear DataFrame consolidado
            lbl_estado.config(text="Procesando datos y creando consolidado...")
            ventana_progreso.update()
            
            df_consolidado = df.groupby(['EMPLEADO', 'NOMBRE_COMPLETO', 'CEDULA'])['VALOR'].sum().reset_index()
            df_consolidado.rename(columns={'VALOR': 'VALOR_TOTAL'}, inplace=True)
            
            # Solicitar ubicación de archivo
            progress_bar.stop()
            lbl_estado.config(text="Seleccione ubicación para guardar...")
            ventana_progreso.update()
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_sugerido = f"SALDOS_PRESTAMOS_CLASE205_{timestamp}.xlsx"
            
            archivo_excel = filedialog.asksaveasfilename(
                title="Guardar Saldos de Préstamos",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                initialfile=nombre_sugerido,
                parent=ventana_progreso
            )
            
            if not archivo_excel:
                ventana_progreso.destroy()
                return
            
            # Crear archivo Excel
            lbl_estado.config(text="Creando archivo Excel...")
            progress_bar.start()
            ventana_progreso.update()
            
            with pd.ExcelWriter(archivo_excel, engine='xlsxwriter') as writer:
                # Hoja detallada
                df.to_excel(writer, sheet_name='Detalle_CLASE205', index=False)
                
                # Hoja consolidada
                df_consolidado.to_excel(writer, sheet_name='Consolidado_CLASE205', index=False)
                
                # Obtener workbook para aplicar formato
                workbook = writer.book
                
                # Formato para encabezados
                header_format = workbook.add_format({
                    'bold': True,
                    'text_wrap': True,
                    'valign': 'top',
                    'fg_color': '#D7E4BC',
                    'border': 1
                })
                
                # Formato para números
                money_format = workbook.add_format({
                    'num_format': '$#,##0.00',
                    'align': 'right'
                })
                
                # Aplicar formato a hoja detallada
                worksheet_detalle = writer.sheets['Detalle_CLASE205']
                for col_num, value in enumerate(df.columns.values):
                    worksheet_detalle.write(0, col_num, value, header_format)
                
                def _col_width(series, col_name):
                    try:
                        w = series.fillna('').astype(str).str.len().max()
                        return max(int(w) if w == w else 0, len(str(col_name))) + 2
                    except Exception:
                        return len(str(col_name)) + 4

                # Ajustar ancho de columnas y aplicar formato de moneda en hoja detallada
                for i, col in enumerate(df.columns):
                    if col == 'VALOR':
                        worksheet_detalle.set_column(i, i, 15, money_format)
                    else:
                        worksheet_detalle.set_column(i, i, min(_col_width(df[col], col), 50))

                # Aplicar formato a hoja consolidada
                worksheet_consolidado = writer.sheets['Consolidado_CLASE205']
                for col_num, value in enumerate(df_consolidado.columns.values):
                    worksheet_consolidado.write(0, col_num, value, header_format)

                # Ajustar ancho de columnas y aplicar formato de moneda en hoja consolidada
                for i, col in enumerate(df_consolidado.columns):
                    if col == 'VALOR_TOTAL':
                        worksheet_consolidado.set_column(i, i, 18, money_format)
                    else:
                        worksheet_consolidado.set_column(i, i, min(_col_width(df_consolidado[col], col), 50))
            
            progress_bar.stop()
            ventana_progreso.destroy()
            
            # Mostrar resultado
            total_empleados = len(df_consolidado)
            total_registros = len(df)
            saldo_total = df_consolidado['VALOR_TOTAL'].sum()
            
            mensaje_resultado = f"""✅ SALDOS EXPORTADOS EXITOSAMENTE
            
📊 Estadísticas:
• Total de empleados con saldos: {total_empleados}
• Total de registros detallados: {total_registros}
• Saldo total general: ${saldo_total:,.2f}

📁 Archivo creado:
{os.path.basename(archivo_excel)}

📋 El archivo contiene 2 hojas:
• Detalle_CLASE205: Todos los registros
• Consolidado_CLASE205: Saldo por empleado"""
            
            result = messagebox.askyesnocancel("Exportación Completada", 
                                             f"{mensaje_resultado}\n\n¿Desea abrir el archivo?")
            
            if result is True:  # Sí, abrir archivo
                if sys.platform == 'win32':
                    os.startfile(archivo_excel)
                elif sys.platform == 'darwin':
                    os.system(f'open "{archivo_excel}"')
                else:
                    os.system(f'xdg-open "{archivo_excel}"')
            elif result is None:  # Cancelar - mostrar ubicación
                messagebox.showinfo("Archivo Guardado", f"Archivo guardado en:\n{archivo_excel}")
            
        except ImportError:
            messagebox.showerror("Error", 
                               "Para exportar a Excel necesita instalar:\n" +
                               "pip install pandas openpyxl xlsxwriter")
        except Exception as e:
            if 'ventana_progreso' in locals():
                ventana_progreso.destroy()
            messagebox.showerror("Error", f"Error al exportar saldos:\n{str(e)}")
    
    def buscar_empleados_por_nombre(self, apellidos=None, nombres=None):
        """Busca empleados por nombre y apellido usando LIKE"""
        conn = self.conectar_bd()
        if not conn:
            return pd.DataFrame()
        
        try:
            cursor = conn.cursor()
            query = """
            SELECT [EMPLEADO], [APELLIDOS], [NOMBRES], [CEDULA], [CARGO]
            FROM [insevig].[dbo].[RPEMPLEA]
            WHERE 1=1
            """
            conditions = []
            params = []

            if apellidos and apellidos.strip():
                conditions.append("UPPER([APELLIDOS]) LIKE UPPER(?)")
                params.append(f"%{apellidos.strip()}%")

            if nombres and nombres.strip():
                # Dividir nombres para buscar cada parte
                nombres_parts = nombres.strip().split()
                for parte in nombres_parts:
                    if len(parte) >= 2:  # Solo buscar partes de al menos 2 caracteres
                        conditions.append("UPPER([NOMBRES]) LIKE UPPER(?)")
                        params.append(f"%{parte}%")

            if not conditions:
                conn.close()
                return pd.DataFrame()

            query += " AND " + " AND ".join(conditions)
            query += " ORDER BY [APELLIDOS], [NOMBRES]"

            cursor.execute(query, params)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            
            conn.close()
            return pd.DataFrame.from_records(rows, columns=columns)

        except Exception as e:
            conn.close()
            messagebox.showerror("Error", f"Error al buscar empleados:\n{e}")
            return pd.DataFrame()
    
    def abrir_buscador_nombres(self):
        """Abre la ventana de búsqueda por nombres"""
        ventana_buscador = tk.Toplevel(self.root)
        ventana_buscador.title("🔍 Buscar Empleado por Nombre")
        ventana_buscador.geometry("800x600")
        ventana_buscador.resizable(True, True)
        
        # Centrar ventana
        ventana_buscador.transient(self.root)
        ventana_buscador.grab_set()
        
        ventana_buscador.update_idletasks()
        x = (ventana_buscador.winfo_screenwidth() // 2) - (800 // 2)
        y = (ventana_buscador.winfo_screenheight() // 2) - (600 // 2)
        ventana_buscador.geometry(f"800x600+{x}+{y}")
        
        # Frame principal
        frame_principal = ttk.Frame(ventana_buscador, padding="15")
        frame_principal.pack(fill="both", expand=True)
        
        # Título
        titulo_label = ttk.Label(frame_principal, text="🔍 BUSCAR EMPLEADO POR NOMBRE", 
                               font=("Segoe UI", 16, "bold"), foreground="#2c3e50")
        titulo_label.pack(pady=(0, 20))
        
        # Frame de búsqueda
        busqueda_frame = ttk.LabelFrame(frame_principal, text=" 📝 Criterios de Búsqueda ", padding="15")
        busqueda_frame.pack(fill="x", pady=(0, 15))
        
        # Campos de búsqueda en dos columnas
        campos_frame = ttk.Frame(busqueda_frame)
        campos_frame.pack(fill="x")
        
        # Apellidos
        ttk.Label(campos_frame, text="Apellidos:", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=0, sticky="e", padx=(0, 10), pady=5)
        entry_apellidos = ttk.Entry(campos_frame, width=25, font=("Segoe UI", 11))
        entry_apellidos.grid(row=0, column=1, sticky="w", pady=5)
        
        # Nombres
        ttk.Label(campos_frame, text="Nombres:", font=("Segoe UI", 11, "bold")).grid(
            row=1, column=0, sticky="e", padx=(0, 10), pady=5)
        entry_nombres = ttk.Entry(campos_frame, width=25, font=("Segoe UI", 11))
        entry_nombres.grid(row=1, column=1, sticky="w", pady=5)
        
        # Botones de búsqueda
        botones_frame = ttk.Frame(busqueda_frame)
        botones_frame.pack(fill="x", pady=(15, 0))
        
        def realizar_busqueda():
            apellidos = entry_apellidos.get().strip()
            nombres = entry_nombres.get().strip()
            
            if not apellidos and not nombres:
                messagebox.showwarning("Advertencia", "Debe ingresar al menos apellidos o nombres para buscar")
                return
            
            if apellidos and len(apellidos) < 2:
                messagebox.showwarning("Advertencia", "Los apellidos deben tener al menos 2 caracteres")
                return
            
            if nombres and len(nombres) < 2:
                messagebox.showwarning("Advertencia", "Los nombres deben tener al menos 2 caracteres")
                return
            
            # Limpiar resultados anteriores
            for item in tree_resultados.get_children():
                tree_resultados.delete(item)
            
            try:
                # Buscar empleados
                lbl_estado.config(text="🔍 Buscando empleados...", foreground="#3498db")
                ventana_buscador.update()

                df_resultados = self.buscar_empleados_por_nombre(apellidos, nombres)

                if df_resultados.empty:
                    try:
                        lbl_estado.config(text="❌ No se encontraron empleados con esos criterios", foreground="#e74c3c")
                    except tk.TclError:
                        pass  # Ventana cerrada
                    messagebox.showinfo("Sin resultados", "No se encontraron empleados que coincidan con los criterios de búsqueda")
                    return

                # Mostrar resultados
                for index, row in df_resultados.iterrows():
                    tree_resultados.insert("", "end", values=(
                        row["EMPLEADO"],
                        row["APELLIDOS"],
                        row["NOMBRES"],
                        self.formatear_cedula(row["CEDULA"]),
                        row["CARGO"] if pd.notna(row["CARGO"]) else ""
                    ))

                total_encontrados = len(df_resultados)
                try:
                    lbl_estado.config(text=f"✅ Se encontraron {total_encontrados} empleado(s)", foreground="#27ae60")
                except tk.TclError:
                    pass  # Ventana cerrada
            except Exception as e:
                print(f"Error en búsqueda: {e}")
        
        def limpiar_busqueda():
            entry_apellidos.delete(0, tk.END)
            entry_nombres.delete(0, tk.END)
            for item in tree_resultados.get_children():
                tree_resultados.delete(item)
            lbl_estado.config(text="Ingrese criterios de búsqueda", foreground="#7f8c8d")
            entry_apellidos.focus()
        
        btn_buscar = ttk.Button(botones_frame, text="🔍 BUSCAR", command=realizar_busqueda)
        btn_buscar.pack(side="left", padx=(0, 10))
        
        btn_limpiar = ttk.Button(botones_frame, text="🧹 LIMPIAR", command=limpiar_busqueda)
        btn_limpiar.pack(side="left")
        
        # Frame de resultados
        resultados_frame = ttk.LabelFrame(frame_principal, text=" 📋 Resultados de la Búsqueda ", padding="10")
        resultados_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # TreeView para resultados
        columns_resultados = ("CODIGO", "APELLIDOS", "NOMBRES", "CEDULA", "CARGO")
        tree_resultados = ttk.Treeview(resultados_frame, columns=columns_resultados, show="headings", height=15)
        
        # Configurar encabezados
        headers_config = {
            "CODIGO": {"text": "CÓDIGO", "width": 80, "anchor": "center"},
            "APELLIDOS": {"text": "APELLIDOS", "width": 150, "anchor": "w"},
            "NOMBRES": {"text": "NOMBRES", "width": 150, "anchor": "w"},
            "CEDULA": {"text": "CÉDULA", "width": 120, "anchor": "center"},
            "CARGO": {"text": "CARGO", "width": 200, "anchor": "w"}
        }
        
        for col, config in headers_config.items():
            tree_resultados.heading(col, text=config["text"])
            tree_resultados.column(col, width=config["width"], anchor=config["anchor"])
        
        # Scrollbars para resultados
        scrollbar_v_res = ttk.Scrollbar(resultados_frame, orient="vertical", command=tree_resultados.yview)
        scrollbar_h_res = ttk.Scrollbar(resultados_frame, orient="horizontal", command=tree_resultados.xview)
        tree_resultados.configure(yscrollcommand=scrollbar_v_res.set, xscrollcommand=scrollbar_h_res.set)
        
        tree_resultados.pack(side="left", fill="both", expand=True)
        scrollbar_v_res.pack(side="right", fill="y")
        scrollbar_h_res.pack(side="bottom", fill="x")
        
        # Estilo para el TreeView
        tree_resultados.tag_configure('selected_row', background='#3498db', foreground='white')
        
        # Frame de estado y botones finales
        estado_frame = ttk.Frame(frame_principal)
        estado_frame.pack(fill="x")
        
        lbl_estado = ttk.Label(estado_frame, text="Ingrese criterios de búsqueda", 
                              font=("Segoe UI", 10, "italic"), foreground="#7f8c8d")
        lbl_estado.pack(side="left")
        
        def seleccionar_empleado():
            selection = tree_resultados.selection()
            if not selection:
                messagebox.showwarning("Advertencia", "Por favor seleccione un empleado de la lista")
                return
            
            # Obtener datos del empleado seleccionado
            item = selection[0]
            valores = tree_resultados.item(item)['values']
            codigo_empleado = valores[0]
            nombre_completo = f"{valores[1]} {valores[2]}"
            
            # Colocar el código en el campo de búsqueda principal
            self.entry_empleado.delete(0, tk.END)
            self.entry_empleado.insert(0, str(codigo_empleado))
            
            # Cerrar ventana de búsqueda
            ventana_buscador.destroy()
            
            # Mostrar confirmación y ejecutar búsqueda automáticamente
            messagebox.showinfo("Empleado Seleccionado", 
                              f"Empleado seleccionado:\n{nombre_completo}\nCódigo: {codigo_empleado}")
            
            # Ejecutar búsqueda automáticamente
            self.buscar_prestamos()
        
        def cancelar():
            ventana_buscador.destroy()
        
        # Botones finales
        botones_finales = ttk.Frame(estado_frame)
        botones_finales.pack(side="right")
        
        btn_seleccionar = ttk.Button(botones_finales, text="✅ SELECCIONAR EMPLEADO", 
                                   command=seleccionar_empleado, style="Success.TButton")
        btn_seleccionar.pack(side="left", padx=(0, 10))
        
        btn_cancelar = ttk.Button(botones_finales, text="❌ CANCELAR", command=cancelar)
        btn_cancelar.pack(side="left")
        
        # Eventos
        def on_double_click(event):
            seleccionar_empleado()
        
        tree_resultados.bind("<Double-1>", on_double_click)
        
        # Permitir búsqueda con Enter
        entry_apellidos.bind("<Return>", lambda e: realizar_busqueda())
        entry_nombres.bind("<Return>", lambda e: realizar_busqueda())
        
        # Enfocar primer campo
        entry_apellidos.focus()
        
        # Atajos de teclado
        ventana_buscador.bind("<Escape>", lambda e: cancelar())
        ventana_buscador.bind("<F5>", lambda e: realizar_busqueda())
    
    def ejecutar_query(self, query, parametros=None):
        """Ejecuta una query en SQL Server y devuelve los resultados"""
        conn = self.conectar_bd()
        if not conn:
            return []
        
        try:
            cursor = conn.cursor()
            if parametros:
                cursor.execute(query, parametros)
            else:
                cursor.execute(query)
            
            # Obtener nombres de columnas
            columns = [column[0] for column in cursor.description]
            
            # Obtener datos
            rows = cursor.fetchall()
            
            # Convertir a lista de diccionarios
            result = []
            for row in rows:
                result.append(dict(zip(columns, row)))
            
            conn.close()
            return result
        except Exception as e:
            conn.close()
            messagebox.showerror("Error", f"Error al ejecutar consulta:\n{e}")
            return []
    
    def obtener_datos_empleado(self, codigo_empleado):
        """Obtiene los datos básicos del empleado"""
        query = """
        SELECT [EMPLEADO], [APELLIDOS], [NOMBRES], [CEDULA]
        FROM [insevig].[dbo].[RPEMPLEA]
        WHERE [EMPLEADO] = ?
        """
        resultado = self.ejecutar_query(query, [codigo_empleado])
        
        if not resultado:
            return None
        
        emp = resultado[0]
        return {
            'empleado': emp['EMPLEADO'],
            'nombre': f"{emp['APELLIDOS']} {emp['NOMBRES']}",
            'cedula': self.formatear_cedula(emp['CEDULA'])
        }
    
    def obtener_todos_saldos(self):
        """Obtiene todos los empleados con saldo de préstamos (RPINGDES), con situación si existe."""
        conn = self.conectar_bd()
        if not conn:
            return []

        def _fetch(cursor, con_situacion):
            extra = ", ISNULL(e.SITUACION,'') AS SITUACION" if con_situacion else ", '' AS SITUACION"
            cursor.execute(f"""
            SELECT i.EMPLEADO,
                   ISNULL(e.APELLIDOS + ' ' + e.NOMBRES, 'Empleado ' + CAST(i.EMPLEADO AS VARCHAR)) AS NOMBRE,
                   ISNULL(SUM(i.VALOR), 0) AS SALDO{extra}
            FROM [insevig].[dbo].[RPINGDES] i
            LEFT JOIN [insevig].[dbo].[RPEMPLEA] e
                   ON i.EMPLEADO = e.EMPLEADO AND e.CODEMP='10' AND e.CODSUC='10'
            WHERE i.CLASE = 205 AND i.CODEMP='10' AND i.CODSUC='10'
            GROUP BY i.EMPLEADO, e.APELLIDOS, e.NOMBRES{', e.SITUACION' if con_situacion else ''}
            HAVING SUM(i.VALOR) > 0
            ORDER BY SALDO DESC
            """)
            return cursor.fetchall(), [d[0] for d in cursor.description]

        try:
            cursor = conn.cursor()
            try:
                rows, columns = _fetch(cursor, con_situacion=True)
                self._panel_tiene_situacion = True
            except Exception:
                conn.close()
                conn = self.conectar_bd()
                cursor = conn.cursor()
                rows, columns = _fetch(cursor, con_situacion=False)
                self._panel_tiene_situacion = False

            conn.close()
            resultados = []
            for row in rows:
                d = dict(zip(columns, row))
                resultados.append({
                    'codigo':    str(d['EMPLEADO']).strip(),
                    'nombre':    d['NOMBRE'].strip(),
                    'saldo':     d['SALDO'],
                    'situacion': str(d.get('SITUACION', '')).strip().upper(),
                })
            return resultados
        except Exception as e:
            try:
                conn.close()
            except Exception:
                pass
            messagebox.showerror("Error", f"Error al obtener saldos:\n{e}")
            return []
    
    def obtener_datos_rpingdes_combinados(self, codigo_empleado):
        """Obtiene saldos y datos detallados de RPINGDES en UNA sola consulta"""
        query = """
        SELECT [NUMERO], [VALOR], [OBSERV], [FECHA], [CONCEPTO]
        FROM [insevig].[dbo].[RPINGDES]
        WHERE [EMPLEADO] = ? AND [CLASE] = 205
        ORDER BY [NUMERO], [FECHA]
        """
        resultado = self.ejecutar_query(query, [codigo_empleado])
        
        saldos_por_numero = {}
        for row in resultado:
            num = row['NUMERO']
            saldos_por_numero[num] = saldos_por_numero.get(num, 0) + row['VALOR']
        
        return saldos_por_numero, resultado
    
    def obtener_saldo_por_numero(self, codigo_empleado):
        saldos, _ = self.obtener_datos_rpingdes_combinados(codigo_empleado)
        return saldos
    
    def obtener_datos_rpingdes(self, codigo_empleado):
        _, datos = self.obtener_datos_rpingdes_combinados(codigo_empleado)
        return datos
    
    def obtener_numeros_excluir(self):
        """Lista de números de egresos que ya están en el historial SQLite"""
        return [
            '27958', '28215', '28592', '29301', '29633', '29790', 
            '30062', '30437', '30691', '30928', '31777', '32211', 
            '32721', '33089', '33634', '33944', '33964', '34483', 
            '34492', '34797', '35168', '35616', '35923'
        ]
    
    def obtener_movimientos_sistema_filtrados(self, codigo_empleado):
        """Obtiene movimientos del sistema excluyendo los ya migrados a SQLite"""
        numeros_excluir = self.obtener_numeros_excluir()
        
        if not numeros_excluir:
            query = """
            SELECT [NUMERO], [FECHA], [VALOR], [CONCEPTO], [OBSERV]
            FROM [insevig].[dbo].[RPHISTOR]
            WHERE [EMPLEADO] = ? AND [CLASE] = 205
            ORDER BY [NUMERO] ASC, [FECHA] ASC
            """
            return self.ejecutar_query(query, [codigo_empleado])
        
        placeholders = ','.join(['?' for _ in numeros_excluir])
        query = f"""
        SELECT [NUMERO], [FECHA], [VALOR], [CONCEPTO], [OBSERV]
        FROM [insevig].[dbo].[RPHISTOR]
        WHERE [EMPLEADO] = ? AND [CLASE] = 205 
        AND [NUMERO] NOT IN ({placeholders})
        ORDER BY [NUMERO] ASC, [FECHA] ASC
        """
        
        parametros = [codigo_empleado] + numeros_excluir
        return self.ejecutar_query(query, parametros)
    
    def convertir_fecha_sqlite(self, fecha_str):
        """Convierte una fecha de SQLite al formato datetime correcto"""
        if not fecha_str:
            return datetime.now()
        
        if isinstance(fecha_str, datetime):
            return fecha_str
        
        formatos_fecha = [
            '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', 
            '%d/%m/%y', '%Y-%m-%d %H:%M:%S'
        ]
        
        for formato in formatos_fecha:
            try:
                return datetime.strptime(str(fecha_str).strip(), formato)
            except ValueError:
                continue
        
        print(f"⚠️  No se pudo convertir fecha: '{fecha_str}' - usando fecha actual")
        return datetime.now()
    
    def obtener_historial_sqlite(self, codigo_empleado, solo_cuadre=False):
        """Obtiene movimientos del historial SQLite para el empleado específico"""
        conn = self.conectar_sqlite()
        if not conn:
            return [] if not solo_cuadre else set()
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
            SELECT fecha, ingreso, egreso, concepto, tipo, numero_fila
            FROM historial_prestamos 
            WHERE codigo_empleado = ?
            ORDER BY fecha ASC
            """, [codigo_empleado])
            
            if solo_cuadre:
                valores = set()
                for row in cursor.fetchall():
                    if row[4] in ('CUADRE', 'CRUZE') and row[2] > 0:
                        valores.add(round(row[2], 2))
                conn.close()
                return valores
            
            resultados = []
            for row in cursor.fetchall():
                fecha_convertida = self.convertir_fecha_sqlite(row[0])
                es_cuadre = row[4] in ('CUADRE', 'CRUZE')
                
                if row[1] > 0:  # INGRESO
                    resultados.append({
                        'FECHA': fecha_convertida,
                        'VALOR': row[1],
                        'OBSERV': row[3] if row[3] else "",
                        'TIPO': 'INGRESO',
                        'NUMERO': f"HIST_{row[5]}",
                        'ORIGEN': 'HISTORICO',
                        'ES_CUADRE': es_cuadre
                    })
                elif row[2] > 0:  # EGRESO
                    resultados.append({
                        'FECHA': fecha_convertida,
                        'VALOR': row[2],
                        'OBSERV': row[3] if row[3] else "",
                        'TIPO': 'EGRESO',
                        'NUMERO': f"HIST_{row[5]}",
                        'ORIGEN': 'HISTORICO',
                        'ES_CUADRE': es_cuadre
                    })
            
            conn.close()
            return resultados
            
        except Exception as e:
            print(f"Error al leer SQLite: {e}")
            conn.close()
            return [] if not solo_cuadre else set()
    
    def obtener_movimientos_completos(self, codigo_empleado, movimientos_sqlite=None):
        """Obtiene TODOS los movimientos combinando SQL Server (filtrado) y SQLite"""
        movimientos_sistema = self.obtener_movimientos_sistema_filtrados(codigo_empleado)
        if movimientos_sqlite is None:
            movimientos_sqlite = self.obtener_historial_sqlite(codigo_empleado)
        
        todos_los_movimientos = []
        
        # Agregar movimientos del sistema actual
        for mov in movimientos_sistema:
            todos_los_movimientos.append({
                'NUMERO': mov['NUMERO'],
                'FECHA': mov['FECHA'],
                'VALOR': mov['VALOR'],
                'CONCEPTO': mov['CONCEPTO'],
                'OBSERV': mov['OBSERV'],
                'ORIGEN': 'SISTEMA'
            })
        
        # Agregar movimientos históricos de SQLite
        for mov in movimientos_sqlite:
            todos_los_movimientos.append({
                'NUMERO': mov['NUMERO'],
                'FECHA': mov['FECHA'],
                'VALOR': mov['VALOR'] if mov['TIPO'] == 'EGRESO' else -mov['VALOR'],
                'CONCEPTO': mov['OBSERV'],
                'OBSERV': mov['OBSERV'],
                'ORIGEN': mov['ORIGEN'],
                'ES_CUADRE': mov.get('ES_CUADRE', False)
            })
        
        return todos_los_movimientos
    
    # Observaciones genéricas que no aportan información real
    _OBS_GENERICAS = frozenset({
        'PRESTAMOS COMPANIA', 'PRESTAMOS COMPAÑIA',
        'PRESTAMO COMPANIA',  'PRESTAMO COMPAÑIA',
        'PRESTAMOS EMPRESA',  'PRESTAMO EMPRESA',
    })

    def _mejor_observacion_prestamo(self, numero, movimientos, datos_rpingdes):
        """Devuelve la mejor observacion disponible para un INGRESO sintetico.

        Busca en este orden:
        1. RPINGDES.OBSERV para ese numero (si no es generica)
        2. Movimientos RPHISTOR ordenados: primero los de VALOR negativo
           (desembolso original) que suelen tener mejor descripcion
        3. Fallback: lo que sea que tenga el primer movimiento
        """
        def _limpiar(v):
            return str(v or '').strip()

        def _util(obs):
            return bool(obs) and obs.upper().rstrip('. ') not in self._OBS_GENERICAS

        # 1. RPINGDES
        for d in datos_rpingdes:
            if str(d['NUMERO']) == str(numero):
                cand = _limpiar(d.get('OBSERV')) or _limpiar(d.get('CONCEPTO'))
                if _util(cand):
                    return cand
                break

        # 2. Todos los movimientos: negativos primero (= desembolso original)
        for mov in sorted(movimientos, key=lambda m: m.get('VALOR', 0)):
            cand = _limpiar(mov.get('OBSERV')) or _limpiar(mov.get('CONCEPTO'))
            if _util(cand):
                return cand

        # 3. Fallback sin filtro
        m0 = movimientos[0] if movimientos else {}
        return _limpiar(m0.get('OBSERV')) or _limpiar(m0.get('CONCEPTO'))

    def agrupar_prestamos_por_numero(self, movimientos):
        """Agrupa los movimientos por número de préstamo"""
        prestamos_agrupados = {}
        
        for mov in movimientos:
            numero = mov['NUMERO']
            if numero not in prestamos_agrupados:
                prestamos_agrupados[numero] = {
                    'numero': numero,
                    'fecha_prestamo': mov['FECHA'],
                    'movimientos': [],
                    'total_egreso': 0,
                    'origen': mov.get('ORIGEN', 'SISTEMA')
                }
            
            prestamos_agrupados[numero]['movimientos'].append(mov)
            if mov['VALOR'] > 0:
                prestamos_agrupados[numero]['total_egreso'] += abs(mov['VALOR'])
        
        return prestamos_agrupados
    
    @staticmethod
    def formatear_cedula(valor):
        """Convierte CEDULA de float/int a string de 10 dígitos con ceros a la izquierda"""
        if valor is None:
            return ""
        try:
            num = int(float(str(valor).strip()))
            return f"{num:010d}"
        except (ValueError, TypeError):
            return str(valor).strip()
    
    def formatear_moneda(self, valor):
        """Formatea un valor como moneda con mejor presentación"""
        if valor == 0:
            return "$0.00"
        return f"${valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    def formatear_fecha(self, fecha):
        """Formatea una fecha"""
        if isinstance(fecha, datetime):
            return fecha.strftime('%d-%b-%Y')
        return str(fecha)
    
    def aplicar_filtros(self):
        """Aplica los filtros seleccionados a los datos"""
        if not self.todos_los_movimientos_originales:
            return
        
        # Limpiar TreeView
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Obtener valores de filtros
        filtro_tipo = self.combo_tipo.get()
        filtro_origen = self.combo_origen.get()
        filtro_numero = self.entry_numero_filtro.get().strip()
        filtro_observacion = self.entry_observacion_filtro.get().strip().lower()
        
        # Filtros de fecha
        fecha_desde = None
        fecha_hasta = None
        try:
            if self.entry_fecha_desde.get().strip():
                fecha_desde = datetime.strptime(self.entry_fecha_desde.get().strip(), '%d/%m/%Y')
            if self.entry_fecha_hasta.get().strip():
                fecha_hasta = datetime.strptime(self.entry_fecha_hasta.get().strip(), '%d/%m/%Y')
                fecha_hasta = fecha_hasta.replace(hour=23, minute=59, second=59)
        except ValueError:
            messagebox.showwarning("Error de Fecha", "Formato de fecha incorrecto. Use DD/MM/AAAA")
            return
        
        # Filtros de monto
        monto_min = None
        monto_max = None
        try:
            if self.entry_monto_min.get().strip():
                monto_min = float(self.entry_monto_min.get().strip().replace(',', ''))
            if self.entry_monto_max.get().strip():
                monto_max = float(self.entry_monto_max.get().strip().replace(',', ''))
        except ValueError:
            messagebox.showwarning("Error de Monto", "Los montos deben ser números válidos")
            return
        
        # Aplicar filtros
        movimientos_filtrados = []
        for mov in self.todos_los_movimientos_originales:
            # Filtro por tipo
            if filtro_tipo != "Todos":
                if (filtro_tipo == "Ingresos" and mov['tipo'] != 'INGRESO') or \
                   (filtro_tipo == "Egresos" and mov['tipo'] != 'EGRESO'):
                    continue
            
            # Filtro por origen
            if filtro_origen != "Todos":
                if (filtro_origen == "Sistema Actual" and mov.get('origen') != 'SISTEMA') or \
                   (filtro_origen == "Histórico SQLite" and mov.get('origen') != 'HISTORICO'):
                    continue
            
            # Filtro por número
            if filtro_numero and filtro_numero not in str(mov['numero']):
                continue
            
            # Filtro por observación
            if filtro_observacion and filtro_observacion not in mov['observacion_completa'].lower():
                continue
            
            # Filtro por fecha
            if fecha_desde and mov['fecha'] < fecha_desde:
                continue
            if fecha_hasta and mov['fecha'] > fecha_hasta:
                continue
            
            # Filtro por monto
            valor_absoluto = abs(self.convertir_moneda_a_numero(mov['ingreso'] if mov['ingreso'] else mov['egreso']))
            if monto_min and valor_absoluto < monto_min:
                continue
            if monto_max and valor_absoluto > monto_max:
                continue
            
            movimientos_filtrados.append(mov)
        
        # Mostrar resultados filtrados
        self.mostrar_movimientos_en_tree(movimientos_filtrados)
        
        # Actualizar información (contar solo registros visibles, excluyendo CUADRE/CRUZE ocultos)
        total_mostrados = sum(1 for m in movimientos_filtrados if not m.get('es_cuadre', False))
        total_originales_visibles = sum(1 for m in self.todos_los_movimientos_originales if not m.get('es_cuadre', False))
        info_filtro = f"📊 Mostrando {total_mostrados} de {total_originales_visibles} registros"
        
        if total_mostrados != total_originales_visibles:
            info_filtro += " (FILTRADO)"
            self.lbl_filtros.config(text=info_filtro, foreground="#e74c3c")
        else:
            self.lbl_filtros.config(text=info_filtro, foreground="#27ae60")
    
    def limpiar_filtros(self):
        """Limpia todos los filtros y muestra todos los datos"""
        self.combo_tipo.set("Todos")
        self.combo_origen.set("Todos")
        self.entry_numero_filtro.delete(0, tk.END)
        self.entry_observacion_filtro.delete(0, tk.END)
        self.entry_fecha_desde.delete(0, tk.END)
        self.entry_fecha_hasta.delete(0, tk.END)
        self.entry_monto_min.delete(0, tk.END)
        self.entry_monto_max.delete(0, tk.END)
        
        if self.todos_los_movimientos_originales:
            self.aplicar_filtros()
    
    def mostrar_movimientos_en_tree(self, movimientos_a_mostrar):
        """Muestra los movimientos en el TreeView"""
        # Limpiar TreeView y datos completos
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.datos_completos = []
        
        saldo_progresivo = 0
        contador = 1
        
        for mov_data in movimientos_a_mostrar:
            fecha_str = self.formatear_fecha(mov_data['fecha'])
            numero_display = mov_data['numero']
            observ_completa = mov_data['observacion_completa']
            observ_display = observ_completa[:77] + "..." if len(observ_completa) > 80 else observ_completa
            
            if mov_data['tipo'] == 'INGRESO':
                saldo_progresivo += self.convertir_moneda_a_numero(mov_data['ingreso'])
                valor_ingreso = mov_data['ingreso']
                valor_egreso = ""
                tags = ('ingreso_historico' if mov_data.get('origen') == 'HISTORICO' else 'ingreso',)
            else:
                saldo_progresivo -= self.convertir_moneda_a_numero(mov_data['egreso'])
                valor_ingreso = ""
                valor_egreso = mov_data['egreso']
                tags = ('egreso_historico' if mov_data.get('origen') == 'HISTORICO' else 'egreso',)
            
            es_cuadre = mov_data.get('es_cuadre', False)
            
            if not es_cuadre:
                self.tree.insert("", "end", values=(
                    contador, fecha_str, valor_ingreso, valor_egreso,
                    numero_display, observ_display, mov_data['tipo'],
                    self.formatear_moneda(saldo_progresivo)
                ), tags=tags)
                
                self.datos_completos.append({
                    'numero_fila': contador,
                    'fecha': fecha_str,
                    'ingreso': valor_ingreso,
                    'egreso': valor_egreso,
                    'numero': str(numero_display),
                    'observacion_completa': observ_completa,
                    'tipo': mov_data['tipo'],
                    'saldo': self.formatear_moneda(saldo_progresivo),
                    'es_historico': mov_data.get('origen') == 'HISTORICO'
                })
                
                contador += 1
        
        self.root.after(100, self.ajustar_columnas_automatico)
    
    def buscar_prestamos(self):
        """Busca y muestra los préstamos del empleado ordenados SOLO por fecha"""
        codigo_empleado = self.entry_empleado.get().strip()
        
        if not codigo_empleado:
            messagebox.showwarning("Advertencia", "Por favor ingrese el código del empleado")
            return
        
        # Limpiar resultados anteriores
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.datos_completos = []
        self.todos_los_movimientos_originales = []
        
        self.lbl_info.config(text="Buscando...")
        self.root.update()
        
        # Obtener datos del empleado
        datos_empleado = self.obtener_datos_empleado(codigo_empleado)
        if not datos_empleado:
            messagebox.showerror("Error", "Empleado no encontrado")
            self.lbl_info.config(text="")
            return
        
        self.empleado_actual = datos_empleado
        
        # Obtener todos los datos necesarios (una sola llamada a SQLite y RPINGDES)
        movimientos_sqlite = self.obtener_historial_sqlite(codigo_empleado)
        valores_cuadre = set()
        for mov in movimientos_sqlite:
            if mov.get('ES_CUADRE') and mov['TIPO'] == 'EGRESO':
                valores_cuadre.add(round(mov['VALOR'], 2))
        
        saldos_por_numero, datos_rpingdes = self.obtener_datos_rpingdes_combinados(codigo_empleado)
        saldo_actual_total = sum(saldos_por_numero.values())
        # Guardar para uso de _preparar_contexto_ia (valores reales de RPINGDES)
        self.empleado_actual['saldo_actual']      = saldo_actual_total
        self.empleado_actual['saldos_por_numero'] = dict(saldos_por_numero)
        movimientos = self.obtener_movimientos_completos(codigo_empleado, movimientos_sqlite)
        
        if not movimientos and saldo_actual_total == 0:
            messagebox.showinfo("Información", "No se encontraron préstamos para este empleado")
            self.lbl_info.config(text="")
            return
        
        # Procesar movimientos para crear lista completa
        prestamos_agrupados = self.agrupar_prestamos_por_numero(movimientos)
        todos_los_movimientos = []
        
        # Agregar INGRESOS
        for numero, prestamo_info in prestamos_agrupados.items():
            if prestamo_info['origen'] == 'HISTORICO':
                mov_ingreso = None
                for mov in prestamo_info['movimientos']:
                    if mov['VALOR'] < 0:
                        mov_ingreso = mov
                        break
                
                if mov_ingreso:
                    todos_los_movimientos.append({
                        'fecha': mov_ingreso['FECHA'],
                        'tipo': 'INGRESO',
                        'valor': abs(mov_ingreso['VALOR']),
                        'numero': mov_ingreso['NUMERO'],
                        'observ': mov_ingreso['OBSERV'],
                        'origen': 'HISTORICO',
                        'es_cuadre': mov_ingreso.get('ES_CUADRE', False)
                    })
            else:
                total_descontado_numero = prestamo_info['total_egreso']
                saldo_pendiente_numero = saldos_por_numero.get(numero, 0)
                valor_total_prestamo = total_descontado_numero + saldo_pendiente_numero
                primer_movimiento = prestamo_info['movimientos'][0]
                
                fecha_ingreso = prestamo_info['fecha_prestamo']
                for d in datos_rpingdes:
                    if str(d['NUMERO']) == str(numero):
                        fecha_ingreso = d['FECHA']
                        break
                
                ingreso_es_cuadre = round(valor_total_prestamo, 2) in valores_cuadre
                obs_ing = self._mejor_observacion_prestamo(
                    numero, prestamo_info['movimientos'], datos_rpingdes)
                todos_los_movimientos.append({
                    'fecha': fecha_ingreso,
                    'tipo': 'INGRESO',
                    'valor': valor_total_prestamo,
                    'numero': primer_movimiento['NUMERO'],
                    'observ': obs_ing,
                    'origen': 'SISTEMA',
                    'es_cuadre': ingreso_es_cuadre
                })
        
        # Agregar préstamos solo en RPINGDES
        for numero, saldo in saldos_por_numero.items():
            if numero not in prestamos_agrupados and saldo > 0:
                observ_real = ""
                fecha_real = datetime.now()

                for dato in datos_rpingdes:
                    if dato['NUMERO'] == numero:
                        fecha_real = dato['FECHA']
                        cand = (str(dato.get('OBSERV') or '').strip()
                                or str(dato.get('CONCEPTO') or '').strip())
                        obs_up = cand.upper().rstrip('. ')
                        observ_real = cand if obs_up not in self._OBS_GENERICAS else ""
                        break

                if not observ_real:
                    observ_real = self._mejor_observacion_prestamo(
                        numero, [], datos_rpingdes)
                
                todos_los_movimientos.append({
                    'fecha': fecha_real,
                    'tipo': 'INGRESO',
                    'valor': saldo,
                    'numero': numero,
                    'observ': observ_real,
                    'origen': 'SISTEMA',
                    'es_cuadre': False
                })
        
        # Agregar EGRESOS
        for mov in movimientos:
            if mov['VALOR'] > 0:
                origen = mov.get('ORIGEN', 'SISTEMA')
                todos_los_movimientos.append({
                    'fecha': mov['FECHA'],
                    'tipo': 'EGRESO',
                    'valor': abs(mov['VALOR']),
                    'numero': mov['NUMERO'],
                    'observ': mov['OBSERV'] if mov['OBSERV'] else "",
                    'origen': origen,
                    'es_cuadre': mov.get('ES_CUADRE', False)
                })
        
        # Ordenar por fecha
        todos_los_movimientos.sort(key=lambda x: x['fecha'])
        
        # Convertir a formato para filtros
        movimientos_para_filtros = []
        for mov in todos_los_movimientos:
            fecha_str = self.formatear_fecha(mov['fecha'])
            raw_numero = mov['numero']
            try:
                num_limpio = str(int(float(str(raw_numero))))
            except (ValueError, TypeError):
                num_limpio = str(raw_numero)
            numero_display = num_limpio
            if mov.get('origen') == 'HISTORICO':
                numero_display += " [H]"
            
            observ_completa = mov['observ'] if mov['observ'] else ""
            
            if mov['tipo'] == 'INGRESO':
                valor_ingreso = self.formatear_moneda(mov['valor'])
                valor_egreso = ""
            else:
                valor_ingreso = ""
                valor_egreso = self.formatear_moneda(mov['valor'])
            
            movimientos_para_filtros.append({
                'fecha': mov['fecha'],
                'tipo': mov['tipo'],
                'numero': numero_display,
                'observacion_completa': observ_completa,
                'ingreso': valor_ingreso,
                'egreso': valor_egreso,
                'origen': mov.get('origen', 'SISTEMA'),
                'es_cuadre': mov.get('es_cuadre', False)
            })
        
        # Guardar movimientos originales para filtros
        self.todos_los_movimientos_originales = movimientos_para_filtros
        
        # Mostrar información del empleado
        info_partes = [
            f"👤 {datos_empleado['nombre']}",
            f"🆔 {datos_empleado['cedula']}",
            f"💰 SALDO: {self.formatear_moneda(saldo_actual_total)}"
        ]
        
        if movimientos_sqlite:
            info_partes.append(f"📁 HISTÓRICOS: {len(movimientos_sqlite)} registros")
        
        total_movimientos = len(todos_los_movimientos)
        if total_movimientos > 0:
            info_partes.append(f"📊 TOTAL: {total_movimientos}")
        
        info_texto = " • ".join(info_partes)
        self.lbl_info.config(text=info_texto, foreground="#2c3e50")
        
        # Aplicar filtros (inicialmente muestra todos)
        self.aplicar_filtros()
    
    def convertir_moneda_a_numero(self, valor_texto):
        """Convierte un valor monetario formateado a número float"""
        if not valor_texto or str(valor_texto).strip() == "":
            return 0.0
        
        try:
            texto_limpio = str(valor_texto).strip().replace("$", "").strip()
            if not texto_limpio:
                return 0.0
            
            if "," in texto_limpio and "." in texto_limpio:
                pos_coma = texto_limpio.rfind(",")
                pos_punto = texto_limpio.rfind(".")
                
                if pos_coma > pos_punto:
                    texto_limpio = texto_limpio.replace(".", "").replace(",", ".")
                else:
                    texto_limpio = texto_limpio.replace(",", "")
            elif "," in texto_limpio:
                comas = texto_limpio.count(",")
                pos_coma = texto_limpio.rfind(",")
                
                if comas > 1 or (len(texto_limpio) - pos_coma) != 3:
                    texto_limpio = texto_limpio.replace(",", "")
                else:
                    texto_limpio = texto_limpio.replace(",", ".")
            elif "." in texto_limpio:
                puntos = texto_limpio.count(".")
                pos_punto = texto_limpio.rfind(".")
                
                if puntos > 1 or (len(texto_limpio) - pos_punto) != 3:
                    texto_limpio = texto_limpio.replace(".", "")
            
            return float(texto_limpio)
            
        except (ValueError, AttributeError) as e:
            print(f"Error convirtiendo '{valor_texto}': {e}")
            return 0.0
    
    def exportar_excel(self):
        """Exporta los datos actuales a Excel con datos del empleado, observaciones completas y hoja de descuentos"""
        if not self.datos_completos:
            messagebox.showwarning("Advertencia", "No hay datos para exportar")
            return
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from tkinter import filedialog
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_sugerido = f"Historial_Prestamos_{timestamp}.xlsx"
            
            archivo = filedialog.asksaveasfilename(
                title="Guardar Reporte de Préstamos",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
                initialfile=nombre_sugerido
            )
            
            if not archivo:
                return
            
            emp = getattr(self, 'empleado_actual', None)
            
            # --- HOJA 1: Historial completo ---
            data = []
            for d in self.datos_completos:
                fecha_dt = datetime.strptime(d['fecha'], '%d-%b-%Y') if d['fecha'] else datetime.now()
                fecha_formateada = fecha_dt.strftime('%d/%m/%Y')
                data.append({
                    '#': d['numero_fila'],
                    'FECHA': fecha_formateada,
                    'INGRESO': self.convertir_moneda_a_numero(d['ingreso']),
                    'EGRESO': self.convertir_moneda_a_numero(d['egreso']),
                    'NUMERO': d['numero'],
                    'OBSERVACIONES': d['observacion_completa'],
                    'TIPO': d['tipo'],
                    'SALDO': self.convertir_moneda_a_numero(d['saldo'])
                })
            
            df_historial = pd.DataFrame(data)
            
            # --- HOJA 2: igual que Historial, pero OBSERVACIONES de EGRESOS = "DESCONTADO MES AÑO" ---
            meses_es = {
                1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
                5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
                9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
            }

            # Egresos del mismo mes/año se consolidan en UNA sola fila
            buf_eg   = None   # acumulador del periodo actual
            filas_d  = []     # resultado final

            def _flush_eg(buf, filas):
                if buf:
                    numeros = ', '.join(sorted(buf['nums'])) if len(buf['nums']) <= 3 else '(varios)'
                    filas.append({
                        '_tipo':  'EGRESO',
                        '_fecha': buf['fecha'],
                        '_ing':   '',
                        '_eg':    buf['total'],
                        '_num':   numeros,
                        '_obs':   buf['label'],
                        '_saldo': buf['saldo'],
                    })

            for d in self.datos_completos:
                fecha_dt = datetime.strptime(d['fecha'], '%d-%b-%Y') if d['fecha'] else datetime.now()
                if d['tipo'] == 'EGRESO':
                    clave   = (fecha_dt.year, fecha_dt.month)
                    mes_str = meses_es[fecha_dt.month]
                    label   = f"DESCONTADO {mes_str} {fecha_dt.year}"
                    eg_val  = self.convertir_moneda_a_numero(d['egreso'])
                    if buf_eg is None or buf_eg['clave'] != clave:
                        _flush_eg(buf_eg, filas_d)
                        buf_eg = {'clave': clave, 'label': label, 'total': 0.0,
                                  'fecha': fecha_dt, 'saldo': 0.0, 'nums': set()}
                    buf_eg['total'] += eg_val
                    buf_eg['saldo']  = self.convertir_moneda_a_numero(d['saldo'])
                    buf_eg['nums'].add(str(d['numero']))
                else:
                    _flush_eg(buf_eg, filas_d)
                    buf_eg = None
                    filas_d.append({
                        '_tipo':  d['tipo'],
                        '_fecha': fecha_dt,
                        '_ing':   self.convertir_moneda_a_numero(d['ingreso']),
                        '_eg':    '',
                        '_num':   d['numero'],
                        '_obs':   d['observacion_completa'],
                        '_saldo': self.convertir_moneda_a_numero(d['saldo']),
                    })
            _flush_eg(buf_eg, filas_d)

            descuentos = []
            for i, r in enumerate(filas_d, start=1):
                descuentos.append({
                    '#':             i,
                    'FECHA':         r['_fecha'].strftime('%d/%m/%Y'),
                    'INGRESO':       r['_ing'],
                    'EGRESO':        r['_eg'],
                    'NUMERO':        r['_num'],
                    'OBSERVACIONES': r['_obs'],
                    'TIPO':          r['_tipo'],
                    'SALDO':         r['_saldo'],
                })

            df_descuentos = pd.DataFrame(descuentos)
            
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Border, Side

            # ── Estilos reutilizables ────────────────────────────────────────
            AZUL_OSC   = "1E3A5F"
            AZUL_CLAR  = "D6E4F0"
            VERDE_OSC  = "1A5C38"
            VERDE_CLAR = "D5F5E3"
            ROJO_CLAR  = "FDECEA"
            GRIS_CLAR  = "F2F3F4"
            BLANCO     = "FFFFFF"

            borde_fino = Border(
                left=Side(style='thin', color="AAAAAA"),
                right=Side(style='thin', color="AAAAAA"),
                top=Side(style='thin', color="AAAAAA"),
                bottom=Side(style='thin', color="AAAAAA"),
            )
            borde_medio = Border(
                left=Side(style='medium', color="1E3A5F"),
                right=Side(style='medium', color="1E3A5F"),
                top=Side(style='medium', color="1E3A5F"),
                bottom=Side(style='medium', color="1E3A5F"),
            )

            ANCHOS_COL = {
                '#': 6, 'FECHA': 14, 'INGRESO': 17, 'EGRESO': 17,
                'SALDO': 17, 'NUMERO': 26, 'TIPO': 13,
                'OBSERVACIONES': 58, 'VALOR': 17, 'DESCRIPCION': 32,
            }
            COLS_DINERO = {'INGRESO', 'EGRESO', 'SALDO', 'VALOR'}
            FMT_MONEDA  = '$#,##0.00'
            FMT_FECHA   = 'DD/MM/YYYY'

            def _estilo_hoja(ws, df, start_row, color_hdr=AZUL_OSC,
                             color_fila_par=AZUL_CLAR, color_fila_impar=BLANCO):
                n_cols = len(df.columns)

                # ── anchos de columna ────────────────────────────────────────
                for i, col in enumerate(df.columns, start=1):
                    letra = get_column_letter(i)
                    if not df.empty:
                        try:
                            w = df[col].fillna('').astype(str).str.len().max()
                            ancho_datos = int(w) if w == w else 0
                        except Exception:
                            ancho_datos = 0
                    else:
                        ancho_datos = 0
                    ancho = ANCHOS_COL.get(col, max(len(str(col)) + 4, ancho_datos + 4))
                    ws.column_dimensions[letra].width = min(ancho, 65)

                # ── encabezado de columnas (start_row) ──────────────────────
                fill_hdr   = PatternFill(start_color=color_hdr, end_color=color_hdr, fill_type="solid")
                font_hdr   = Font(bold=True, color="FFFFFF", size=11)
                alin_hdr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for cell in ws[start_row]:
                    if cell.column > n_cols:
                        break
                    cell.fill    = fill_hdr
                    cell.font    = font_hdr
                    cell.border  = borde_medio
                    cell.alignment = alin_hdr
                ws.row_dimensions[start_row].height = 24

                # ── filas de datos con colores alternos y bordes ─────────────
                fill_par   = PatternFill(start_color=color_fila_par,   end_color=color_fila_par,   fill_type="solid")
                fill_impar = PatternFill(start_color=color_fila_impar, end_color=color_fila_impar, fill_type="solid")
                font_dato  = Font(size=10)

                for row_idx, row in enumerate(ws.iter_rows(min_row=start_row + 1,
                                                            max_row=ws.max_row,
                                                            min_col=1, max_col=n_cols), start=1):
                    fill = fill_par if row_idx % 2 == 0 else fill_impar
                    for cell in row:
                        col_name = df.columns[cell.column - 1]
                        cell.border = borde_fino
                        cell.font   = font_dato
                        cell.fill   = fill
                        if col_name in COLS_DINERO:
                            cell.number_format = FMT_MONEDA
                            cell.alignment = Alignment(horizontal="right")
                        elif col_name == 'FECHA':
                            cell.alignment = Alignment(horizontal="center")
                        elif col_name == '#':
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", wrap_text=False)

                # ── fila de totales ─────────────────────────────────────────
                cols_tot = [c for c in df.columns if c in COLS_DINERO]
                if cols_tot and not df.empty:
                    fill_tot = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                    font_tot = Font(bold=True, color="FFFFFF", size=11)
                    tot_row  = ws.max_row + 1
                    ws.cell(tot_row, 1, value="TOTAL").font   = font_tot
                    ws.cell(tot_row, 1).fill   = fill_tot
                    ws.cell(tot_row, 1).border = borde_medio
                    ws.cell(tot_row, 1).alignment = Alignment(horizontal="center")
                    for i, col in enumerate(df.columns, start=1):
                        cell_t = ws.cell(tot_row, i)
                        cell_t.border = borde_medio
                        cell_t.fill   = fill_tot
                        cell_t.font   = font_tot
                        if col in cols_tot:
                            data_start  = start_row + 1
                            data_end    = tot_row - 1
                            col_letter  = get_column_letter(i)
                            cell_t.value          = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
                            cell_t.number_format  = FMT_MONEDA
                            cell_t.alignment      = Alignment(horizontal="right")
                    ws.row_dimensions[tot_row].height = 20

                # ── congelar encabezado y activar autofiltro ─────────────────
                ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
                ws.auto_filter.ref = ws.cell(start_row, 1).coordinate + ":" + \
                                     ws.cell(start_row, n_cols).coordinate

            def _escribir_banner(ws, emp, font_title):
                if emp:
                    ws.cell(row=1, column=1, value=f"EMPLEADO: {emp['nombre']}").font = Font(bold=True, size=13, color="1E3A5F")
                    ws.cell(row=2, column=1, value=f"CÉDULA: {emp['cedula']}").font   = Font(size=11)
                    ws.cell(row=3, column=1, value=f"CÓDIGO: {emp['empleado']}").font  = Font(size=11)
                    ws.cell(row=4, column=1,
                            value=f"FECHA REPORTE: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=11, italic=True)
                    ws.row_dimensions[1].height = 18

            # Escribir con openpyxl para incluir encabezado del empleado
            with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
                start_row = 5 if emp else 1  # fila donde va el encabezado de columnas
                # Hoja 1: Historial
                df_historial.to_excel(writer, sheet_name='Historial', index=False, startrow=start_row - 1)
                ws = writer.sheets['Historial']
                _escribir_banner(ws, emp, Font(bold=True, size=13))
                _estilo_hoja(ws, df_historial, start_row,
                             color_hdr=AZUL_OSC, color_fila_par=AZUL_CLAR, color_fila_impar=BLANCO)

                # Hoja 2: Descuentos
                if not df_descuentos.empty:
                    df_descuentos.to_excel(writer, sheet_name='DESCUENTOS', index=False, startrow=start_row - 1)
                    ws2 = writer.sheets['DESCUENTOS']
                    _escribir_banner(ws2, emp, Font(bold=True, size=13))
                    _estilo_hoja(ws2, df_descuentos, start_row,
                                 color_hdr=VERDE_OSC, color_fila_par=VERDE_CLAR, color_fila_impar=BLANCO)
            
            messagebox.showinfo("Éxito", f"Datos exportados exitosamente a:\n{os.path.basename(archivo)}")
            
            if messagebox.askyesno("Abrir Archivo", "¿Desea abrir el archivo exportado?"):
                if sys.platform == 'win32':
                    os.startfile(archivo)
                elif sys.platform == 'darwin':
                    os.system(f'open "{archivo}"')
                else:
                    os.system(f'xdg-open "{archivo}"')
            
        except ImportError:
            messagebox.showerror("Error", "Para exportar a Excel necesita instalar:\npip install pandas openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar:\n{str(e)}")
    
    def mostrar_detalle_fila(self, event):
        """Muestra una ventana con los detalles completos de la fila seleccionada"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        valores = self.tree.item(item)['values']
        
        if len(valores) < 8 or not hasattr(self, 'datos_completos'):
            return
        
        numero_fila = valores[0]
        datos_fila = None
        for dato in self.datos_completos:
            if dato['numero_fila'] == numero_fila:
                datos_fila = dato
                break
        
        if not datos_fila:
            numero_valor = valores[4]
            datos_fila = {
                'numero_fila': valores[0],
                'fecha': valores[1],
                'ingreso': valores[2],
                'egreso': valores[3],
                'numero': str(numero_valor),
                'observacion_completa': valores[5],
                'tipo': valores[6],
                'saldo': valores[7],
                'es_historico': "[H]" in str(numero_valor)
            }
        
        self.crear_ventana_detalle(
            datos_fila['numero_fila'], datos_fila['fecha'], datos_fila['ingreso'],
            datos_fila['egreso'], str(datos_fila['numero']).replace(" [H]", ""),
            datos_fila['observacion_completa'], datos_fila['tipo'],
            datos_fila['saldo'], datos_fila['es_historico']
        )
    
    def crear_ventana_detalle(self, num_fila, fecha, ingreso, egreso, numero, 
                            observacion, tipo, saldo, es_historico):
        """Crea una ventana elegante con los detalles del movimiento"""
        ventana_detalle = tk.Toplevel(self.root)
        ventana_detalle.title(f"📋 Detalles del Movimiento #{num_fila}")
        ventana_detalle.geometry("700x600")
        ventana_detalle.resizable(True, True)
        ventana_detalle.transient(self.root)
        
        ventana_detalle.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 700) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 600) // 2
        ventana_detalle.geometry(f"700x600+{x}+{y}")
        
        ventana_detalle.grab_set()
        
        frame_principal = ttk.Frame(ventana_detalle, padding="20")
        frame_principal.pack(fill="both", expand=True)
        
        titulo_texto = f"💰 DETALLE DEL {tipo}"
        if es_historico:
            titulo_texto += " 📁 [HISTÓRICO]"
        
        titulo_label = ttk.Label(frame_principal, text=titulo_texto, 
                               font=("Segoe UI", 16, "bold"), foreground="#2c3e50")
        titulo_label.pack(pady=(0, 20))
        
        observacion_resumida = (str(observacion).strip()[:80] + "…" if len(str(observacion).strip()) > 80 
                                else str(observacion).strip()) if observacion and str(observacion).strip() else "—"
        
        info_frame = ttk.LabelFrame(frame_principal, text=" ℹ️ Información General ", padding="15")
        info_frame.pack(fill="x", pady=(0, 15))
        
        info_data = [
            ("📅 FECHA:", fecha),
            ("🔢 NÚMERO:", numero),
            ("📊 POSICIÓN:", f"#{num_fila}"),
            ("💰 VALOR:", ingreso if ingreso else egreso),
            ("📈 SALDO RESULTANTE:", saldo),
            ("🏷️ TIPO:", tipo),
            ("📍 ORIGEN:", "Histórico (SQLite)" if es_historico else "Sistema Actual (SQL Server)"),
            ("📝 OBSERVACIÓN:", observacion_resumida),
        ]
        
        for i, (etiqueta, valor) in enumerate(info_data):
            ttk.Label(info_frame, text=etiqueta, font=("Segoe UI", 10, "bold")).grid(
                row=i, column=0, sticky="w", pady=3, padx=(0, 15))
            lbl_valor = ttk.Label(info_frame, text=str(valor), font=("Segoe UI", 10))
            lbl_valor.grid(row=i, column=1, sticky="w", pady=3)
            if i == 7 and observacion and str(observacion).strip():
                lbl_valor.configure(foreground="#2980B9")
        
        obs_frame = ttk.LabelFrame(frame_principal, text=" 📝 Observación Completa ", padding="15")
        obs_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        text_frame = ttk.Frame(obs_frame)
        text_frame.pack(fill="both", expand=True)
        
        text_obs = tk.Text(text_frame, wrap=tk.WORD, font=("Segoe UI", 12), 
                          height=12, relief="solid", borderwidth=1,
                          bg="#ffffff", fg="#2c3e50", padx=10, pady=10)
        
        scrollbar_text = ttk.Scrollbar(text_frame, orient="vertical", command=text_obs.yview)
        text_obs.configure(yscrollcommand=scrollbar_text.set)
        
        observacion_texto = str(observacion).strip() if observacion and str(observacion).strip() and str(observacion).strip().lower() not in ['none', 'null', ''] else "Sin observaciones registradas"
        text_obs.insert("1.0", observacion_texto)
        
        text_obs.pack(side="left", fill="both", expand=True)
        scrollbar_text.pack(side="right", fill="y")
        text_obs.config(state="disabled")
        
        botones_frame = ttk.Frame(frame_principal)
        botones_frame.pack(fill="x", pady=(15, 0))
        
        def copiar_observacion():
            try:
                ventana_detalle.clipboard_clear()
                ventana_detalle.clipboard_append(observacion_texto)
                btn_copiar.config(text="✅ COPIADO")
                ventana_detalle.after(2000, lambda: btn_copiar.config(text="📋 COPIAR OBSERVACIÓN"))
            except Exception as e:
                btn_copiar.config(text="❌ ERROR")
                ventana_detalle.after(2000, lambda: btn_copiar.config(text="📋 COPIAR OBSERVACIÓN"))
        
        btn_copiar = ttk.Button(botones_frame, text="📋 COPIAR OBSERVACIÓN", 
                               command=copiar_observacion)
        btn_copiar.pack(side="left", padx=(0, 10))
        
        btn_cerrar = ttk.Button(botones_frame, text="❌ CERRAR", 
                               command=ventana_detalle.destroy)
        btn_cerrar.pack(side="right")
        
        btn_cerrar.focus()
        ventana_detalle.bind("<Escape>", lambda e: ventana_detalle.destroy())
        ventana_detalle.bind("<Return>", lambda e: ventana_detalle.destroy())
    
    # ══════════════════════════════════════════════════════════════════════
    #  ANÁLISIS CON IA (Groq / OpenRouter / Ollama)
    # ══════════════════════════════════════════════════════════════════════

    # Config por defecto — funciona sin ia_config.json externo
    _IA_CONFIG_DEFAULT = {
        "proveedor": "openrouter",
        "api_key":   "<OPENROUTER_API_KEY>",  # Cargar desde config/ia_config.json
        "modelo":    "openai/gpt-oss-20b:free",
        "ollama_url": "http://localhost:11434",
    }

    def _ruta_config_ia(self):
        # Guarda junto al EXE (o junto al .pyw en dev), no en el temp de extracción
        return os.path.join(_app_dir(), 'ia_config.json')

    def _cargar_config_ia(self):
        try:
            with open(self._ruta_config_ia(), 'r', encoding='utf-8') as f:
                cfg = json.load(f)
            # Rellenar claves faltantes con los defaults
            for k, v in self._IA_CONFIG_DEFAULT.items():
                cfg.setdefault(k, v)
            return cfg
        except Exception:
            return dict(self._IA_CONFIG_DEFAULT)

    def _guardar_config_ia(self, config):
        try:
            with open(self._ruta_config_ia(), 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def _preparar_contexto_ia(self, desde=None, hasta=None):
        import math
        emp       = getattr(self, 'empleado_actual', None)
        data_full = self.datos_completos
        if not data_full:
            return ""

        nombre = emp['nombre'] if emp else "Empleado desconocido"
        cedula = emp.get('cedula', '-') if emp else '-'

        MESES_ES = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
                    7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}

        def _fdt(d):
            try:
                return datetime.strptime(d['fecha'], '%d-%b-%Y')
            except Exception:
                return None

        # ── filtrar por período si se especifica ────────────────────────
        data = []
        for d in data_full:
            fdt = _fdt(d)
            if desde and fdt and fdt < desde:
                continue
            if hasta and fdt and fdt > hasta:
                continue
            data.append(d)
        if not data:
            return ""

        periodo_filtro_txt = ""
        if desde or hasta:
            d_str = desde.strftime('%d/%m/%Y') if desde else '(inicio)'
            h_str = hasta.strftime('%d/%m/%Y') if hasta else '(hoy)'
            periodo_filtro_txt = f"\n(Análisis filtrado al período {d_str} — {h_str})"

        # ── saldo real desde RPINGDES (fuente de verdad) ────────────────
        saldo_act    = emp.get('saldo_actual')          if emp else None
        saldos_x_num = emp.get('saldos_por_numero', {}) if emp else {}
        total_eg  = sum(self.convertir_moneda_a_numero(d['egreso'])  for d in data if d['tipo'] == 'EGRESO')
        total_ing = sum(self.convertir_moneda_a_numero(d['ingreso']) for d in data if d['tipo'] == 'INGRESO')
        if saldo_act is None:
            saldo_act = max(0.0, total_ing - total_eg)

        # ── período global ──────────────────────────────────────────────
        fechas_todas = [f for d in data for f in [_fdt(d)] if f]
        fecha_inicio_str = min(fechas_todas).strftime('%d/%m/%Y') if fechas_todas else '-'
        fecha_ultimo_str = max(fechas_todas).strftime('%d/%m/%Y') if fechas_todas else '-'

        # ── períodos globales con descuento ─────────────────────────────
        periodos_eg: dict = {}
        for d in data:
            if d['tipo'] == 'EGRESO':
                fdt = _fdt(d)
                if fdt:
                    k = (fdt.year, fdt.month)
                    periodos_eg[k] = periodos_eg.get(k, 0) + self.convertir_moneda_a_numero(d['egreso'])
        n_periodos   = len(periodos_eg)
        prom_mensual = (total_eg / n_periodos) if n_periodos else 0

        # ── último mes con descuento y meses sin pago ───────────────────
        hoy = datetime.now()
        if periodos_eg:
            ultimo_k      = max(periodos_eg.keys())
            meses_sin     = (hoy.year - ultimo_k[0]) * 12 + (hoy.month - ultimo_k[1])
            ultimo_eg_str = datetime(ultimo_k[0], ultimo_k[1], 1).strftime('%B %Y')
        else:
            meses_sin     = 0
            ultimo_eg_str = '-'

        # ── resumen detallado por préstamo ──────────────────────────────
        prestamos: dict = {}
        for d in data:
            num = str(d['numero'])
            if num not in prestamos:
                prestamos[num] = {'ing': 0.0, 'eg': 0.0,
                                  'fecha_otorgado': None,
                                  'concepto': '',
                                  'cuotas': []}  # (year, month, amount)
            fdt     = _fdt(d)
            obs_raw = str(d.get('observacion_completa', '')).strip()

            if d['tipo'] == 'INGRESO':
                prestamos[num]['ing'] += self.convertir_moneda_a_numero(d['ingreso'])
                if fdt and prestamos[num]['fecha_otorgado'] is None:
                    prestamos[num]['fecha_otorgado'] = fdt
                if not prestamos[num]['concepto'] and len(obs_raw) > 3:
                    prestamos[num]['concepto'] = obs_raw[:90]
            else:
                amt = self.convertir_moneda_a_numero(d['egreso'])
                prestamos[num]['eg'] += amt
                if fdt:
                    prestamos[num]['cuotas'].append((fdt.year, fdt.month, amt))
                if (not prestamos[num]['concepto']
                        and len(obs_raw) > 4
                        and not obs_raw.upper().startswith('DESCONTADO')):
                    prestamos[num]['concepto'] = obs_raw[:90]

        detalle_lines = []
        for num, p in sorted(prestamos.items()):
            # saldo real RPINGDES; calculado solo para prestamos solo en SQLite
            if num in saldos_x_num:
                saldo_p = saldos_x_num[num]
            else:
                saldo_p = max(0.0, p['ing'] - p['eg'])
            cancelado = saldo_p <= 0.01

            f_oto        = p['fecha_otorgado'].strftime('%d/%m/%Y') if p['fecha_otorgado'] else '?'
            concepto_txt = f' "{p["concepto"]}"' if p['concepto'] else ''

            cuotas_ord = sorted(set((y, m) for y, m, _ in p['cuotas']))
            n_cuotas   = len(cuotas_ord)
            cuota_prom = (p['eg'] / n_cuotas) if n_cuotas else 0

            # brechas entre cuotas
            brechas = 0
            for i in range(1, len(cuotas_ord)):
                prev, cur = cuotas_ord[i-1], cuotas_ord[i]
                diff = (cur[0] - prev[0]) * 12 + (cur[1] - prev[1])
                if diff > 1:
                    brechas += diff - 1
            brecha_txt = f", {brechas} mes(es) sin descuento en el periodo" if brechas else ", pagos continuos"

            if cancelado:
                ult_k  = cuotas_ord[-1] if cuotas_ord else None
                f_ult  = f"{MESES_ES[ult_k[1]]}-{ult_k[0]}" if ult_k else '?'
                linea  = (f"  - Prestamo {num}{concepto_txt}: otorgado {f_oto} por ${p['ing']:,.2f} — "
                          f"CANCELADO (ultima cuota {f_ult}, {n_cuotas} cuotas de ~${cuota_prom:,.0f}/mes)")
            else:
                prim_k   = cuotas_ord[0]  if cuotas_ord else None
                ult_k    = cuotas_ord[-1] if cuotas_ord else None
                f_primer = f"{MESES_ES[prim_k[1]]}-{prim_k[0]}" if prim_k else '?'
                f_ult    = f"{MESES_ES[ult_k[1]]}-{ult_k[0]}"   if ult_k  else '?'
                meses_r  = math.ceil(saldo_p / cuota_prom) if cuota_prom > 0 else '?'
                linea    = (f"  - Prestamo {num}{concepto_txt}: otorgado {f_oto} por ${p['ing']:,.2f} — "
                            f"primer descuento {f_primer}, ultimo {f_ult} "
                            f"({n_cuotas} cuotas ~${cuota_prom:,.0f}/mes{brecha_txt}) — "
                            f"saldo pendiente ${saldo_p:,.2f}, aprox. {meses_r} mes(es) para cancelar")
            detalle_lines.append(linea)

        sin_pago_txt = (f"\n  ATENCION: {meses_sin} mes(es) sin descuento desde el ultimo pago.\n"
                        if meses_sin > 1 else "")

        return (
            f"Empleado: {nombre}  |  Cédula: {cedula}\n"
            f"Período analizado: {fecha_inicio_str} al {fecha_ultimo_str}{periodo_filtro_txt}\n\n"
            f"RESUMEN FINANCIERO GLOBAL:\n"
            f"  Total recibido en préstamos : ${total_ing:,.2f}\n"
            f"  Total descontado en nómina  : ${total_eg:,.2f}\n"
            f"  Saldo pendiente actual      : ${saldo_act:,.2f}\n"
            f"  Meses con descuento         : {n_periodos}\n"
            f"  Promedio mensual descontado : ${prom_mensual:,.2f}\n"
            f"  Último mes con descuento    : {ultimo_eg_str}\n"
            f"{sin_pago_txt}"
            f"\n*** TOTAL DEUDA HOY (OFICIAL, usa SOLO este número para el total): ${saldo_act:,.2f} ***\n"
            f"\nDETALLE POR PRÉSTAMO:\n" + "\n".join(detalle_lines)
        )

    def _llamar_ia(self, config, contexto):
        import urllib.request, urllib.error, json as _json

        SYSTEM = (
            "Eres asistente de nomina. Escribe en espanol muy sencillo, con frases cortas, "
            "como si le explicaras a alguien que no sabe de contabilidad. "
            "Escribe DOS parrafos sin titulos ni listas. "
            "PARRAFO 1 — prestamos que aun debe: por CADA prestamo activo escribe OBLIGATORIAMENTE "
            "en este orden: (a) de que fue exactamente segun el concepto dado, "
            "(b) el mes y anio en que se lo otorgaron, "
            "(c) cuanto debe todavia de ese prestamo, "
            "(d) en que mes y anio aproximado terminaria de pagarlo al ritmo actual. "
            "Si lleva meses sin descuento en algun prestamo, dilo. "
            "PARRAFO 2 — lo que ya cancelo: por cada prestamo cancelado di de que fue y cuando lo termino de pagar. "
            "Ultima frase del parrafo 2: copia EXACTAMENTE el monto que aparece en "
            "'TOTAL DEUDA HOY (OFICIAL)' del contexto — no sumes ni calcules, usa ese numero tal cual. "
            "Usa palabras simples. Nada de terminos contables. No omitas ningun prestamo activo."
        )

        proveedor = config.get('proveedor', 'groq')

        if proveedor == 'ollama':
            url    = config.get('ollama_url', 'http://localhost:11434').rstrip('/') + '/api/chat'
            modelo = config.get('modelo', 'llama3')
            payload = _json.dumps({
                "model":    modelo,
                "messages": [
                    {"role": "system", "content": SYSTEM},
                    {"role": "user",   "content": contexto},
                ],
                "stream": False,
            }).encode()
            req = urllib.request.Request(url, data=payload,
                                         headers={"Content-Type": "application/json"})
            try:
                with urllib.request.urlopen(req, timeout=60) as r:
                    resp = _json.loads(r.read())
            except urllib.error.HTTPError as e:
                raise Exception(f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')[:300]}")
            return resp['message']['content'].strip()

        # Groq o OpenRouter (formato OpenAI)
        if proveedor == 'openrouter':
            url     = 'https://openrouter.ai/api/v1/chat/completions'
            modelo_cfg = config.get('modelo', '').strip()
            # Lista de fallback — se prueba en orden hasta que uno responde
            FALLBACK = [
                modelo_cfg,
                'openai/gpt-oss-120b:free',
                'openai/gpt-oss-20b:free',
                'meta-llama/llama-3.3-70b-instruct:free',
                'meta-llama/llama-3.2-3b-instruct:free',
                'nvidia/nemotron-nano-9b-v2:free',
                'nousresearch/hermes-3-llama-3.1-405b:free',
            ]
            modelos = list(dict.fromkeys(m for m in FALLBACK if m))  # sin duplicados ni vacíos
        else:
            url     = 'https://api.groq.com/openai/v1/chat/completions'
            modelos = [config.get('modelo', 'llama-3.1-8b-instant')]

        api_key = config.get('api_key', '')
        headers_api = {
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {api_key}",
            "HTTP-Referer":  "https://github.com/insevig/historial-prestamos",
            "X-Title":       "Historial Prestamos INSEVIG",
        }

        ultimo_error = "Sin respuesta"
        for modelo in modelos:
            payload = _json.dumps({
                "model":       modelo,
                "messages":    [
                    {"role": "system", "content": SYSTEM},
                    {"role": "user",   "content": contexto},
                ],
                "temperature": 0.4,
                "max_tokens":  950,
            }).encode()
            req = urllib.request.Request(url, data=payload, headers=headers_api)
            try:
                with urllib.request.urlopen(req, timeout=45) as r:
                    resp = _json.loads(r.read())
                if 'choices' in resp:
                    return resp['choices'][0]['message']['content'].strip()
                ultimo_error = str(resp.get('error', resp))
            except urllib.error.HTTPError as e:
                cuerpo = e.read().decode('utf-8', errors='replace')
                try:
                    detalle = _json.loads(cuerpo).get('error', {}).get('message', cuerpo)
                except Exception:
                    detalle = cuerpo
                ultimo_error = f"HTTP {e.code} [{modelo}]: {detalle[:200]}"
                if e.code not in (429, 500, 502, 503):
                    raise Exception(ultimo_error)
            except Exception as ex:
                ultimo_error = str(ex)

        raise Exception(f"Todos los modelos fallaron.\nÚltimo error: {ultimo_error}")

    def _mostrar_analisis_ia(self, texto):
        import threading, subprocess, sys, re

        v = tk.Toplevel(self.root)
        v.title("Analisis IA — Historial de Prestamos")
        v.geometry("700x520")
        v.resizable(True, True)
        v.transient(self.root)
        v.grab_set()

        emp    = getattr(self, 'empleado_actual', None)
        nombre = emp['nombre'] if emp else ""
        ttk.Label(v, text=f"Analisis: {nombre}",
                  font=("Segoe UI", 11, "bold"), foreground="#1E3A5F").pack(pady=(10, 3))

        # ── controles TTS (se packean ANTES que frame_txt para garantizar visibilidad) ──
        tts_handle = [None]   # [subprocess.Popen | pyttsx3 engine | None]

        def _texto_limpio():
            # quitar emojis y caracteres no pronunciables
            return re.sub(r'[^\w\s\.,;:\-\$\%\(\)/]', ' ', texto, flags=re.UNICODE)

        def _detener():
            obj = tts_handle[0]
            tts_handle[0] = None   # primero a None para que loops lean el cambio
            if obj is None:
                return
            try:
                if hasattr(obj, 'kill'):         # subprocess.Popen → kill duro
                    obj.kill()
                elif hasattr(obj, 'terminate'):  # subprocess.Popen → terminar
                    obj.terminate()
                elif hasattr(obj, 'stop'):       # pyttsx3 engine
                    try: obj.stop()
                    except Exception: pass
                    try: obj.endLoop()
                    except Exception: pass
            except Exception:
                pass
            try:
                btn_leer.config(state="normal")
                btn_parar.config(state="disabled")
                lbl_estado_tts.config(text="")
            except Exception:
                pass

        def _leer():
            _detener()
            btn_leer.config(state="disabled")
            btn_parar.config(state="normal")
            lbl_estado_tts.config(text="Leyendo...")

            def _preparar_tts(texto):
                """Convierte montos y símbolos a texto legible para TTS."""
                import re as _re
                def _fmt_monto(m):
                    entero = _re.sub(r'[,.]', '', m.group(1))
                    dec    = m.group(2)
                    if dec == '00':
                        return entero + ' dólares'
                    return entero + ' dólares con ' + dec + ' centavos'
                # $1,234.56 o $1.234,56
                texto = _re.sub(r'\$\s*([\d,.]+)[.,](\d{2})\b', _fmt_monto, texto)
                texto = _re.sub(r'\$\s*([\d,.]+)',
                                lambda m: _re.sub(r'[,.]', '', m.group(1)) + ' dólares', texto)
                # 1,234 → 1234 (separador de miles residual)
                texto = _re.sub(r'(\d),(\d{3})\b', r'\1\2', texto)
                # 200.00 → 200 (decimales triviales)
                texto = _re.sub(r'\b(\d+)\.00\b', r'\1', texto)
                # 5.5% → 5.5 por ciento
                texto = _re.sub(r'(\d+(?:\.\d+)?)\s*%', r'\1 por ciento', texto)
                return texto

            def _run():
                t = _preparar_tts(_texto_limpio())
                exito = False
                _frozen = getattr(sys, 'frozen', False)
                _NO_WIN = 0x08000000  # CREATE_NO_WINDOW

                # ── Windows: pyttsx3 primero (SAPI5 in-process, sin ventana) ─────
                if sys.platform == 'win32':
                    try:
                        import pyttsx3
                        v.after(0, lambda: lbl_estado_tts.config(text="Leyendo..."))
                        engine = pyttsx3.init()
                        voices = engine.getProperty('voices') or []
                        for voz in voices:
                            vid = (voz.id + voz.name).lower()
                            if any(x in vid for x in ('spanish', 'espanol', 'es_', '_es', 'es-')):
                                engine.setProperty('voice', voz.id)
                                break
                        engine.setProperty('rate',   145)
                        engine.setProperty('volume', 1.0)
                        tts_handle[0] = engine
                        engine.say(t[:4500])
                        engine.runAndWait()
                        exito = True
                    except Exception:
                        pass

                # ── Windows fallback: PowerShell SAPI sin ventana ─────────────────
                if not exito and sys.platform == 'win32':
                    try:
                        v.after(0, lambda: lbl_estado_tts.config(text="Leyendo..."))
                        safe = (t[:4500]
                                .replace('\\', ' ').replace('"', ' ')
                                .replace("'", ' ').replace('\n', ' ')
                                .replace('<', ' ').replace('>', ' '))
                        ps = (
                            "Add-Type -AssemblyName System.Speech; "
                            "$s=New-Object System.Speech.Synthesis.SpeechSynthesizer; "
                            "$s.Rate=-1; $s.Volume=100; "
                            "try{$v=$s.GetInstalledVoices()|"
                            "Where-Object{$_.VoiceInfo.Culture -like 'es*'}|"
                            "Select-Object -First 1;"
                            "if($v){$s.SelectVoice($v.VoiceInfo.Name)}}catch{}; "
                            f'$s.Speak("{safe}")'
                        )
                        proc = subprocess.Popen(
                            ['powershell', '-WindowStyle', 'Hidden',
                             '-NonInteractive', '-NoProfile', '-Command', ps],
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                            creationflags=_NO_WIN,
                        )
                        tts_handle[0] = proc
                        proc.wait()
                        exito = (proc.returncode == 0)
                    except Exception:
                        pass

                # ── Linux/Mac intento 1: gtts + pygame (Google, voz natural) ─────
                if not exito and sys.platform != 'win32':
                    try:
                        v.after(0, lambda: lbl_estado_tts.config(text="Cargando voz..."))
                        try:
                            from gtts import gTTS
                            import pygame as _pg
                        except ImportError:
                            if not _frozen:
                                v.after(0, lambda: lbl_estado_tts.config(text="Instalando voz..."))
                                subprocess.run(
                                    [sys.executable, '-m', 'pip', 'install', 'gtts', 'pygame', '-q'],
                                    capture_output=True, timeout=120
                                )
                                from gtts import gTTS
                                import pygame as _pg
                            else:
                                raise

                        import tempfile, os as _os, time as _t
                        tmp_mp3 = tempfile.mktemp(suffix='.mp3')
                        v.after(0, lambda: lbl_estado_tts.config(text="Leyendo..."))
                        gTTS(text=t[:4500], lang='es', slow=False).save(tmp_mp3)
                        _pg.mixer.pre_init(44100, -16, 1, 512)
                        _pg.mixer.init()
                        _pg.mixer.music.load(tmp_mp3)
                        _pg.mixer.music.play()
                        import time as _t
                        while _pg.mixer.music.get_busy():
                            if tts_handle[0] is None:
                                _pg.mixer.music.stop()
                                break
                            _t.sleep(0.15)
                        _pg.mixer.quit()
                        try: _os.unlink(tmp_mp3)
                        except Exception: pass
                        exito = True
                    except Exception:
                        pass

                # ── Linux intento 2: espeak-ng (offline, respaldo) ────────────────
                if not exito and sys.platform != 'win32':
                    for exe in ('espeak-ng', 'espeak'):
                        try:
                            proc = subprocess.Popen(
                                [exe, '-v', 'es+m3', '-s', '135', '-a', '200', '-p', '38'],
                                stdin=subprocess.PIPE,
                                stdout=subprocess.DEVNULL,
                                stderr=subprocess.DEVNULL,
                            )
                            tts_handle[0] = proc
                            proc.stdin.write(t.encode('utf-8', errors='replace'))
                            proc.stdin.close()
                            proc.wait()
                            exito = True
                            break
                        except FileNotFoundError:
                            continue
                        except Exception:
                            break

                # ── Linux intento 3: reproductores con gtts mp3 ───────────────────
                if not exito and sys.platform != 'win32':
                    try:
                        import tempfile, os as _os
                        tmp_mp3 = tempfile.mktemp(suffix='.mp3')
                        from gtts import gTTS
                        gTTS(text=t[:4500], lang='es', slow=False).save(tmp_mp3)
                        for player, args in [
                            ('mpg123',  [tmp_mp3]),
                            ('mpg321',  [tmp_mp3]),
                            ('ffplay',  ['-nodisp', '-autoexit', tmp_mp3]),
                            ('mplayer', [tmp_mp3]),
                        ]:
                            try:
                                proc = subprocess.Popen(
                                    [player] + args,
                                    stdout=subprocess.DEVNULL,
                                    stderr=subprocess.DEVNULL,
                                )
                                tts_handle[0] = proc
                                proc.wait()
                                exito = True
                                break
                            except FileNotFoundError:
                                continue
                        try: _os.unlink(tmp_mp3)
                        except Exception: pass
                    except Exception:
                        pass

                if not exito:
                    if sys.platform == 'win32':
                        msg = ("No se pudo activar la voz.\n\n"
                               "Verifique que Windows tiene voces instaladas:\n"
                               "Configuracion > Hora e idioma > Voz")
                        v.after(0, lambda m=msg: messagebox.showwarning("Voz no disponible", m))

                tts_handle[0] = None
                try:
                    btn_leer.config(state="normal")
                    btn_parar.config(state="disabled")
                    lbl_estado_tts.config(text="")
                except Exception:
                    pass

            threading.Thread(target=_run, daemon=True).start()

        def _cerrar():
            _detener()
            v.destroy()

        ctrl = ttk.Frame(v)
        ctrl.pack(side="bottom", pady=(4, 10))

        btn_leer  = ttk.Button(ctrl, text="Leer en voz alta", command=_leer,
                               style="Success.TButton", width=18)
        btn_leer.pack(side="left", padx=(0, 6))

        btn_parar = ttk.Button(ctrl, text="Parar", command=_detener,
                               state="disabled", width=8)
        btn_parar.pack(side="left", padx=(0, 6))

        lbl_estado_tts = ttk.Label(ctrl, text="", font=("Segoe UI", 9), foreground="#2980b9")
        lbl_estado_tts.pack(side="left", padx=(0, 12))

        ttk.Button(ctrl, text="Cerrar", command=_cerrar).pack(side="left")

        frame_txt = ttk.Frame(v)
        frame_txt.pack(fill="both", expand=True, padx=15, pady=(0, 4))

        txt = tk.Text(frame_txt, wrap="word", font=("Segoe UI", 11),
                      bg="#f7f9fc", relief="flat", bd=0,
                      padx=10, pady=10, state="normal")
        sb = ttk.Scrollbar(frame_txt, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)
        txt.insert("1.0", texto)
        txt.config(state="disabled")

        v.bind("<Escape>", lambda e: _cerrar())
        v.protocol("WM_DELETE_WINDOW", _cerrar)

    def _configurar_ia(self):
        config = self._cargar_config_ia()

        v = tk.Toplevel(self.root)
        v.title("⚙️ Configurar IA")
        v.geometry("480x310")
        v.resizable(False, False)
        v.transient(self.root)
        v.grab_set()

        f = ttk.Frame(v, padding=20)
        f.pack(fill="both", expand=True)

        # Proveedor
        ttk.Label(f, text="Proveedor:", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", pady=4)
        prov_var = tk.StringVar(value=config.get('proveedor', 'groq'))
        combo_prov = ttk.Combobox(f, textvariable=prov_var,
                                   values=["groq", "openrouter", "ollama"],
                                   state="readonly", width=18)
        combo_prov.grid(row=0, column=1, sticky="w", padx=10)

        # API Key
        ttk.Label(f, text="API Key:", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky="w", pady=4)
        key_var = tk.StringVar(value=config.get('api_key', ''))
        entry_key = ttk.Entry(f, textvariable=key_var, width=38, show="*")
        entry_key.grid(row=1, column=1, sticky="w", padx=10)

        # Modelo
        ttk.Label(f, text="Modelo:", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="w", pady=4)
        modelo_var = tk.StringVar(value=config.get('modelo', ''))
        entry_modelo = ttk.Entry(f, textvariable=modelo_var, width=38)
        entry_modelo.grid(row=2, column=1, sticky="w", padx=10)

        # URL Ollama
        ttk.Label(f, text="URL Ollama:", font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="w", pady=4)
        url_var = tk.StringVar(value=config.get('ollama_url', 'http://localhost:11434'))
        entry_url = ttk.Entry(f, textvariable=url_var, width=38)
        entry_url.grid(row=3, column=1, sticky="w", padx=10)

        defaults = {
            'groq':        'llama-3.1-8b-instant',
            'openrouter':  'openai/gpt-oss-20b:free',
            'ollama':      'llama3',
        }

        def _on_prov_change(event=None):
            prov = prov_var.get()
            if not modelo_var.get() or modelo_var.get() in defaults.values():
                modelo_var.set(defaults.get(prov, ''))
            entry_url.config(state="normal" if prov == "ollama" else "disabled")
            entry_key.config(state="disabled" if prov == "ollama" else "normal")

        combo_prov.bind("<<ComboboxSelected>>", _on_prov_change)
        _on_prov_change()

        hint_texts = {
            'groq':       "API key gratuita en: console.groq.com",
            'openrouter': "API key gratuita en: openrouter.ai/keys",
            'ollama':     "Sin API key — requiere Ollama corriendo localmente",
        }
        lbl_hint = ttk.Label(f, text=hint_texts.get(prov_var.get(), ''),
                              font=("Segoe UI", 9), foreground="#2980b9")
        lbl_hint.grid(row=4, column=0, columnspan=2, sticky="w", pady=(8, 0))

        def _upd_hint(event=None):
            lbl_hint.config(text=hint_texts.get(prov_var.get(), ''))
        combo_prov.bind("<<ComboboxSelected>>", lambda e: (_on_prov_change(), _upd_hint()))

        def _guardar():
            self._guardar_config_ia({
                'proveedor':  prov_var.get(),
                'api_key':    key_var.get().strip(),
                'modelo':     modelo_var.get().strip(),
                'ollama_url': url_var.get().strip(),
            })
            v.destroy()
            messagebox.showinfo("Guardado", "Configuración de IA guardada.")

        bf = ttk.Frame(f)
        bf.grid(row=5, column=0, columnspan=2, pady=(16, 0))
        ttk.Button(bf, text="💾 Guardar", command=_guardar,
                   style="Primary.TButton").pack(side="left", padx=5)
        ttk.Button(bf, text="Cancelar", command=v.destroy).pack(side="left", padx=5)

    def _dialogo_periodo_ia(self):
        """Muestra un diálogo para seleccionar el período a analizar y una pregunta opcional.
        Retorna dict {'desde': datetime|None, 'hasta': datetime|None, 'pregunta': str}
        o None si el usuario canceló."""
        data = self.datos_completos
        fechas_disp = []
        for d in data:
            try:
                fechas_disp.append(datetime.strptime(d['fecha'], '%d-%b-%Y'))
            except Exception:
                pass

        f_min = min(fechas_disp) if fechas_disp else datetime(2000, 1, 1)
        f_max = max(fechas_disp) if fechas_disp else datetime.now()

        resultado = {}

        v = tk.Toplevel(self.root)
        v.title("🤖 Configurar análisis IA")
        v.geometry("480x310")
        v.resizable(False, False)
        v.transient(self.root)
        v.grab_set()

        f = ttk.Frame(v, padding=18)
        f.pack(fill="both", expand=True)

        ttk.Label(f, text="Período a analizar", font=("Segoe UI", 10, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 6))

        ttk.Label(f, text="Desde:").grid(row=1, column=0, sticky="w")
        desde_var = tk.StringVar(value=f_min.strftime('%d/%m/%Y'))
        ttk.Entry(f, textvariable=desde_var, width=13).grid(row=1, column=1, sticky="w", padx=(4, 14))

        ttk.Label(f, text="Hasta:").grid(row=1, column=2, sticky="w")
        hasta_var = tk.StringVar(value=f_max.strftime('%d/%m/%Y'))
        ttk.Entry(f, textvariable=hasta_var, width=13).grid(row=1, column=3, sticky="w", padx=(4, 0))

        # botones de atajo de período
        bf2 = ttk.Frame(f)
        bf2.grid(row=2, column=0, columnspan=4, sticky="w", pady=(8, 2))

        hoy = datetime.now()

        def _set_todo():
            desde_var.set(f_min.strftime('%d/%m/%Y'))
            hasta_var.set(f_max.strftime('%d/%m/%Y'))

        def _set_ultimo_anio():
            desde_var.set(datetime(hoy.year - 1, hoy.month, 1).strftime('%d/%m/%Y'))
            hasta_var.set(f_max.strftime('%d/%m/%Y'))

        def _set_ultimo_semestre():
            mes_ini = hoy.month - 6
            anio_ini = hoy.year
            if mes_ini <= 0:
                mes_ini += 12
                anio_ini -= 1
            desde_var.set(datetime(anio_ini, mes_ini, 1).strftime('%d/%m/%Y'))
            hasta_var.set(f_max.strftime('%d/%m/%Y'))

        def _set_anio_actual():
            desde_var.set(datetime(hoy.year, 1, 1).strftime('%d/%m/%Y'))
            hasta_var.set(f_max.strftime('%d/%m/%Y'))

        ttk.Button(bf2, text="Todo el historial",  command=_set_todo,           width=17).pack(side="left", padx=(0, 4))
        ttk.Button(bf2, text=f"Año {hoy.year}",    command=_set_anio_actual,    width=10).pack(side="left", padx=(0, 4))
        ttk.Button(bf2, text="Último año",          command=_set_ultimo_anio,    width=10).pack(side="left", padx=(0, 4))
        ttk.Button(bf2, text="Último semestre",     command=_set_ultimo_semestre,width=14).pack(side="left")

        ttk.Separator(f, orient="horizontal").grid(row=3, column=0, columnspan=4, sticky="ew", pady=10)

        ttk.Label(f, text="Pregunta específica (opcional):",
                  font=("Segoe UI", 10, "bold")).grid(row=4, column=0, columnspan=4, sticky="w")
        ttk.Label(f, text='Ej: "¿cuándo termina de pagar?" o "resumen para el gerente"',
                  font=("Segoe UI", 9), foreground="#555").grid(row=5, column=0, columnspan=4, sticky="w")
        pregunta_var = tk.StringVar()
        ttk.Entry(f, textvariable=pregunta_var, width=52).grid(
            row=6, column=0, columnspan=4, sticky="ew", pady=(4, 0))

        lbl_err = ttk.Label(f, text="", foreground="red", font=("Segoe UI", 9))
        lbl_err.grid(row=7, column=0, columnspan=4, sticky="w", pady=(4, 0))

        def _analizar():
            def _parse(s, campo):
                s = s.strip()
                if not s:
                    return None
                for fmt in ('%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y', '%Y-%m-%d'):
                    try:
                        return datetime.strptime(s, fmt)
                    except ValueError:
                        pass
                lbl_err.config(text=f"Fecha inválida en '{campo}'. Use DD/MM/AAAA.")
                return "ERROR"

            d_val = _parse(desde_var.get(), "Desde")
            h_val = _parse(hasta_var.get(), "Hasta")
            if d_val == "ERROR" or h_val == "ERROR":
                return
            if d_val and h_val and d_val > h_val:
                lbl_err.config(text="'Desde' no puede ser mayor que 'Hasta'.")
                return
            resultado['desde']    = d_val
            resultado['hasta']    = h_val
            resultado['pregunta'] = pregunta_var.get().strip()
            v.destroy()

        bf3 = ttk.Frame(f)
        bf3.grid(row=8, column=0, columnspan=4, pady=(14, 0))
        ttk.Button(bf3, text="🤖 Analizar", command=_analizar,
                   style="Primary.TButton").pack(side="left", padx=(0, 8))
        ttk.Button(bf3, text="Cancelar", command=v.destroy).pack(side="left")
        v.bind("<Return>", lambda e: _analizar())
        v.bind("<Escape>", lambda e: v.destroy())

        v.wait_window()
        return resultado if resultado else None

    def analizar_prestamos_ia(self):
        if not self.datos_completos:
            messagebox.showwarning("Sin datos", "Primero busque un empleado para analizar.")
            return

        config = self._cargar_config_ia()
        if not config.get('proveedor'):
            if not messagebox.askyesno("Configurar IA",
                                        "No hay IA configurada todavía.\n¿Desea configurarla ahora?"):
                return
            self._configurar_ia()
            config = self._cargar_config_ia()
            if not config.get('proveedor'):
                return

        if config.get('proveedor') != 'ollama' and not config.get('api_key', '').strip():
            messagebox.showwarning("Sin API Key",
                                   "Ingrese un API Key en ⚙️ Configurar IA antes de analizar.")
            self._configurar_ia()
            return

        # ── diálogo de período y pregunta ───────────────────────────────
        params = self._dialogo_periodo_ia()
        if params is None:
            return  # usuario canceló

        desde    = params.get('desde')
        hasta    = params.get('hasta')
        pregunta = params.get('pregunta', '')

        contexto = self._preparar_contexto_ia(desde=desde, hasta=hasta)
        if not contexto:
            messagebox.showwarning("Sin datos", "No hay movimientos en el período seleccionado.")
            return
        if pregunta:
            contexto += f"\n\nPregunta específica del usuario: {pregunta}"

        v_carga = tk.Toplevel(self.root)
        v_carga.title("Analizando…")
        v_carga.geometry("320x110")
        v_carga.resizable(False, False)
        v_carga.transient(self.root)
        v_carga.grab_set()
        ttk.Label(v_carga, text="🤖  Consultando IA, un momento…",
                  font=("Segoe UI", 11)).pack(expand=True, pady=(18, 4))
        pb = ttk.Progressbar(v_carga, mode='indeterminate')
        pb.pack(fill="x", padx=30, pady=6)
        pb.start()
        v_carga.update()

        try:
            resultado = self._llamar_ia(config, contexto)
            v_carga.destroy()
            self._mostrar_analisis_ia(resultado)
        except Exception as e:
            v_carga.destroy()
            messagebox.showerror("Error IA",
                                  f"No se pudo obtener el análisis:\n\n{e}\n\n"
                                  "Verifique el API key y el proveedor en ⚙️ Configurar IA.")

    def ajustar_columnas_automatico(self):
        """Ajusta automáticamente el ancho de las columnas basado en el contenido"""
        LIMITES = {
            '#':            (40,   60),
            'FECHA':        (95,  120),
            'INGRESO':      (90,  115),
            'EGRESO':       (90,  115),
            'SALDO MENSUAL':(100, 125),
            'NUMERO':       (90,  120),
            'TIPO':         (75,  100),
            'OBSERV':       (350, 800),
        }
        CHARS_PX = {'#': 10, 'FECHA': 10, 'INGRESO': 10, 'EGRESO': 10,
                    'SALDO MENSUAL': 10, 'NUMERO': 9, 'TIPO': 10, 'OBSERV': 8}

        col_list = list(self.tree['columns'])
        for col in col_list:
            header_text = self.tree.heading(col)['text']
            min_w, max_w = LIMITES.get(col, (80, 400))
            px = CHARS_PX.get(col, 9)
            best = len(header_text) * 12 + 30

            for item in self.tree.get_children():
                values = self.tree.item(item)['values']
                idx = col_list.index(col)
                if idx < len(values):
                    w = len(str(values[idx])) * px + 30
                    if w > best:
                        best = w

            self.tree.column(col, width=int(min(max(best, min_w), max_w)))

        if self.tree.get_children():
            self.tree.see(self.tree.get_children()[0])
    
    def verificar_conexiones(self):
        """Verifica el estado de las conexiones"""
        estado_sql = "❌"
        estado_sqlite = "❌"
        
        conn_sql = self.conectar_bd()
        if conn_sql:
            try:
                cursor = conn_sql.cursor()
                cursor.execute("SELECT COUNT(*) FROM [insevig].[dbo].[RPEMPLEA]")
                total_empleados = cursor.fetchone()[0]
                estado_sql = f"✅ ({total_empleados} empleados)"
                conn_sql.close()
            except:
                estado_sql = "❌"
        
        conn_sqlite = self.conectar_sqlite()
        if conn_sqlite:
            try:
                cursor = conn_sqlite.cursor()
                cursor.execute("SELECT COUNT(*) FROM historial_prestamos")
                total_historicos = cursor.fetchone()[0]
                estado_sqlite = f"✅ ({total_historicos} registros)"
                conn_sqlite.close()
            except:
                estado_sqlite = "❌"
        
        return estado_sql, estado_sqlite
    
    def mostrar_estado_conexiones(self):
        """Muestra ventana con estado detallado de las conexiones"""
        estado_sql, estado_sqlite = self.verificar_conexiones()
        
        ventana_estado = tk.Toplevel(self.root)
        ventana_estado.title("Estado de Conexiones")
        ventana_estado.geometry("500x350")
        ventana_estado.resizable(False, False)
        
        ventana_estado.transient(self.root)
        ventana_estado.grab_set()
        
        frame_principal = ttk.Frame(ventana_estado, padding="20")
        frame_principal.pack(fill="both", expand=True)
        
        ttk.Label(frame_principal, text="Estado de las Conexiones", font=("Arial", 14, "bold")).pack(pady=(0, 20))
        
        # SQL Server
        ttk.Label(frame_principal, text="SQL Server (Sistema Principal):", font=("Arial", 12, "bold")).pack(anchor="w")
        ttk.Label(frame_principal, text=f"Servidor: {self.server}", font=("Arial", 10)).pack(anchor="w", padx=20)
        ttk.Label(frame_principal, text=f"Base de datos: {self.database}", font=("Arial", 10)).pack(anchor="w", padx=20)
        ttk.Label(frame_principal, text=f"Estado: {estado_sql}", font=("Arial", 10)).pack(anchor="w", padx=20)
        
        ttk.Separator(frame_principal, orient='horizontal').pack(fill="x", pady=15)
        
        # SQLite
        ttk.Label(frame_principal, text="SQLite (Historial):", font=("Arial", 12, "bold")).pack(anchor="w")
        ttk.Label(frame_principal, text=f"Archivo: {self.sqlite_path}", font=("Arial", 10)).pack(anchor="w", padx=20)
        ttk.Label(frame_principal, text=f"Estado: {estado_sqlite}", font=("Arial", 10)).pack(anchor="w", padx=20)
        
        ttk.Separator(frame_principal, orient='horizontal').pack(fill="x", pady=15)
        
        # Números excluidos
        ttk.Label(frame_principal, text="Números Excluidos:", font=("Arial", 12, "bold")).pack(anchor="w")
        numeros_excluir = self.obtener_numeros_excluir()
        texto_numeros = ", ".join(numeros_excluir[:10]) + f"... (Total: {len(numeros_excluir)})"
        ttk.Label(frame_principal, text=texto_numeros, font=("Arial", 10)).pack(anchor="w", padx=20)
        
        ttk.Button(frame_principal, text="Cerrar", command=ventana_estado.destroy).pack(pady=20)
    
    def cargar_panel_saldos(self):
        """Carga la lista de empleados con saldo en el panel lateral"""
        try:
            self.todos_saldos = self.obtener_todos_saldos()
            self.mostrar_panel_saldos(self.todos_saldos)
        except Exception as e:
            print(f"Error cargando panel de saldos: {e}")
    
    def mostrar_panel_saldos(self, datos):
        """Muestra los datos filtrados en el panel de saldos"""
        for item in self.panel_saldos_tree.get_children():
            self.panel_saldos_tree.delete(item)
        for emp in datos:
            self.panel_saldos_tree.insert("", "end", values=(
                emp['codigo'],
                emp['nombre'],
                self.formatear_moneda(emp['saldo'])
            ))
    
    def filtrar_panel_saldos(self):
        """Filtra la lista de empleados del panel lateral por texto y estado."""
        texto       = self.entry_filtro_saldo.get().strip().lower()
        solo_act    = getattr(self, 'var_solo_activos', None)
        filtrar_act = solo_act.get() if solo_act else False

        resultado = self.todos_saldos
        if filtrar_act and getattr(self, '_panel_tiene_situacion', False):
            resultado = [e for e in resultado if e.get('situacion') == 'ACT']
        if texto:
            resultado = [e for e in resultado
                         if texto in e['codigo'] or texto in e['nombre'].lower()]
        self.mostrar_panel_saldos(resultado)
    
    def seleccionar_empleado_panel(self, event=None):
        """Carga los prestamos del empleado seleccionado en el panel"""
        selection = self.panel_saldos_tree.selection()
        if not selection:
            return
        valores = self.panel_saldos_tree.item(selection[0])['values']
        if not valores:
            return
        codigo = str(valores[0]).strip()
        self.entry_empleado.delete(0, tk.END)
        self.entry_empleado.insert(0, codigo)
        self.buscar_prestamos()
    
    def setup_gui(self):
        """Configura la interfaz gráfica con sistema de filtros"""
        self.root = tk.Tk()
        self.root.title("Consultor de Historial de Prestamos v10 - INSEVIG")
        self.root.geometry("1600x900")

        # Ícono de la ventana
        try:
            ico = os.path.join(_bundle_dir(), 'src', 'logo_insevig.ico')
            png = os.path.join(_bundle_dir(), 'src', 'logo_insevig.png')
            if os.path.exists(ico):
                self.root.iconbitmap(default=ico)
            elif os.path.exists(png):
                self._logo_img = tk.PhotoImage(file=png)
                self.root.iconphoto(True, self._logo_img)
        except Exception:
            pass

        # Cerrar splash de PyInstaller una vez que la ventana está lista
        try:
            import pyi_splash
            pyi_splash.close()
        except ImportError:
            pass

        self.configurar_estilos()
        
        try:
            self.root.state('zoomed')
        except:
            self.root.attributes('-zoomed', True)
        
        # ── BANDA SUPERIOR: logo | título | filtros | búsqueda ─────────────
        frame_superior = ttk.Frame(self.root)
        frame_superior.pack(fill="x", padx=15, pady=(5, 3))

        # LOGO (extremo izquierdo)
        logo_frame = ttk.Frame(frame_superior)
        logo_frame.pack(side="left", padx=(0, 12))
        ttk.Label(logo_frame, text="🔷🔹", font=("Arial", 18)).pack(anchor="w")
        ttk.Label(logo_frame, text="PEREIRA",
                 font=("Segoe UI", 13, "bold"), foreground="#4472C4").pack(anchor="w")
        ttk.Label(logo_frame, text="SYSTEMS",
                 font=("Segoe UI", 9), foreground="#7F8C8D").pack(anchor="w")

        # BÚSQUEDA Y BOTONES (extremo derecho — se packea antes para reservar espacio)
        busqueda_frame = ttk.LabelFrame(frame_superior, text=" Busqueda y Acciones ", padding="5")
        busqueda_frame.pack(side="right")

        primera_linea = ttk.Frame(busqueda_frame)
        primera_linea.pack(pady=(0, 3))
        ttk.Label(primera_linea, text="Empleado:", font=("Segoe UI", 10, "bold")).pack(side="left", padx=(0, 4))
        self.entry_empleado = ttk.Entry(primera_linea, width=12, font=("Segoe UI", 11, "bold"))
        self.entry_empleado.pack(side="left", padx=(0, 4))
        self.entry_empleado.bind("<Return>", lambda e: self.buscar_prestamos())
        ttk.Button(primera_linea, text="NOMBRES",
                  command=self.abrir_buscador_nombres,
                  style="Primary.TButton").pack(side="left", padx=(0, 4))
        ttk.Button(primera_linea, text="BUSCAR",
                  command=self.buscar_prestamos).pack(side="left")

        segunda_linea = ttk.Frame(busqueda_frame)
        segunda_linea.pack()
        ttk.Button(segunda_linea, text="SALDOS",
                  command=self.exportar_saldos_prestamos_excel,
                  style="Success.TButton").pack(side="left", padx=(0, 4))
        ttk.Button(segunda_linea, text="EXCEL",
                  command=self.exportar_excel).pack(side="left", padx=(0, 4))
        ttk.Button(segunda_linea, text="IA",
                  command=self.analizar_prestamos_ia,
                  style="Primary.TButton").pack(side="left", padx=(0, 3))
        ttk.Button(segunda_linea, text="CFG",
                  command=self._configurar_ia,
                  width=4).pack(side="left")

        # FILTROS (a la izquierda de búsqueda — se packea segundo desde la derecha)
        filtros_frame = ttk.LabelFrame(frame_superior, text=" Filtros ", padding="5")
        filtros_frame.pack(side="right", padx=(0, 8))

        fila1 = ttk.Frame(filtros_frame)
        fila1.pack(fill="x", pady=(0, 3))

        ttk.Label(fila1, text="Tipo:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.combo_tipo = ttk.Combobox(fila1, values=["Todos", "Ingresos", "Egresos"],
                                      state="readonly", width=7, font=("Segoe UI", 8))
        self.combo_tipo.set("Todos")
        self.combo_tipo.pack(side="left", padx=(0, 5))
        self.combo_tipo.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtros())

        ttk.Label(fila1, text="Origen:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.combo_origen = ttk.Combobox(fila1, values=["Todos", "Sistema Actual", "Historico SQLite"],
                                        state="readonly", width=9, font=("Segoe UI", 8))
        self.combo_origen.set("Todos")
        self.combo_origen.pack(side="left", padx=(0, 5))
        self.combo_origen.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtros())

        ttk.Label(fila1, text="Obs:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.entry_observacion_filtro = ttk.Entry(fila1, width=11, font=("Segoe UI", 8))
        self.entry_observacion_filtro.pack(side="left", padx=(0, 5))
        self.entry_observacion_filtro.bind("<KeyRelease>", lambda e: self.root.after(500, self.aplicar_filtros))

        ttk.Button(fila1, text="X", command=self.limpiar_filtros,
                  style="Warning.TButton", width=2).pack(side="left")

        fila2 = ttk.Frame(filtros_frame)
        fila2.pack(fill="x")

        ttk.Label(fila2, text="Num:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.entry_numero_filtro = ttk.Entry(fila2, width=6, font=("Segoe UI", 8))
        self.entry_numero_filtro.pack(side="left", padx=(0, 5))
        self.entry_numero_filtro.bind("<KeyRelease>", lambda e: self.root.after(500, self.aplicar_filtros))

        ttk.Label(fila2, text="Desde:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.entry_fecha_desde = ttk.Entry(fila2, width=9, font=("Segoe UI", 8))
        self.entry_fecha_desde.pack(side="left", padx=(0, 5))
        self.entry_fecha_desde.bind("<KeyRelease>", lambda e: self.root.after(1000, self.aplicar_filtros))

        ttk.Label(fila2, text="Hasta:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.entry_fecha_hasta = ttk.Entry(fila2, width=9, font=("Segoe UI", 8))
        self.entry_fecha_hasta.pack(side="left", padx=(0, 5))
        self.entry_fecha_hasta.bind("<KeyRelease>", lambda e: self.root.after(1000, self.aplicar_filtros))

        ttk.Label(fila2, text="Min:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.entry_monto_min = ttk.Entry(fila2, width=6, font=("Segoe UI", 8))
        self.entry_monto_min.pack(side="left", padx=(0, 4))
        self.entry_monto_min.bind("<KeyRelease>", lambda e: self.root.after(1000, self.aplicar_filtros))

        ttk.Label(fila2, text="Max:", font=("Segoe UI", 8)).pack(side="left", padx=(0, 2))
        self.entry_monto_max = ttk.Entry(fila2, width=6, font=("Segoe UI", 8))
        self.entry_monto_max.pack(side="left", padx=(0, 4))
        self.entry_monto_max.bind("<KeyRelease>", lambda e: self.root.after(1000, self.aplicar_filtros))

        self.lbl_filtros = ttk.Label(fila2, text="", font=("Segoe UI", 8, "italic"))
        self.lbl_filtros.pack(side="left", padx=(4, 0))

        # TÍTULO (centro — se packea al último para rellenar el espacio restante)
        titulo_frame = ttk.Frame(frame_superior)
        titulo_frame.pack(side="left", expand=True, fill="x", padx=(10, 0))
        ttk.Label(titulo_frame, text="CONSULTOR DE PRESTAMOS",
                 font=("Segoe UI", 16, "bold"), foreground="#2c3e50").pack(anchor="center")
        ttk.Label(titulo_frame, text="Sistema Integrado de Prestamos  •  F5: buscar por nombre  •  DD/MM/AAAA",
                 font=("Segoe UI", 9), foreground="#7f8c8d").pack(anchor="center")

        ttk.Separator(self.root, orient='horizontal').pack(fill="x", pady=(3, 2))

        # Frame de información del empleado
        self.frame_info = ttk.LabelFrame(self.root, text=" Informacion del Empleado ", padding="4")
        self.frame_info.pack(fill="x", padx=15, pady=(2, 4))

        self.lbl_info = ttk.Label(self.frame_info,
                                 text="Ingrese un codigo de empleado o use NOMBRES para buscar...",
                                 font=("Segoe UI", 11, "bold"), foreground="#34495e")
        self.lbl_info.pack()
        
        # Frame principal con PanedWindow (treeview + panel saldos)
        frame_principal_contenido = ttk.Frame(self.root)
        frame_principal_contenido.pack(fill="both", expand=True, padx=15, pady=(0, 8))
        
        paned = ttk.PanedWindow(frame_principal_contenido, orient=tk.HORIZONTAL)
        paned.pack(fill="both", expand=True)
        
        # --- PANEL IZQUIERDO: TreeView de movimientos ---
        frame_tree = ttk.LabelFrame(paned, text=" 📋 Historial de Movimientos (💡 Doble clic para ver detalles) ", padding="5")
        paned.add(frame_tree, weight=4)
        
        columns = ('#', 'FECHA', 'INGRESO', 'EGRESO', 'NUMERO', 'OBSERV', 'TIPO', 'SALDO MENSUAL')
        self.tree = ttk.Treeview(frame_tree, columns=columns, show='headings', height=25, cursor="hand2")
        
        headers_config = {
            '#':            {'text': '#',             'width': 45,  'anchor': 'center', 'stretch': False},
            'FECHA':        {'text': 'FECHA',         'width': 100, 'anchor': 'center', 'stretch': False},
            'INGRESO':      {'text': 'INGRESO ($)',   'width': 100, 'anchor': 'e',      'stretch': False},
            'EGRESO':       {'text': 'EGRESO ($)',    'width': 100, 'anchor': 'e',      'stretch': False},
            'NUMERO':       {'text': 'NÚMERO',        'width': 100, 'anchor': 'center', 'stretch': False},
            'OBSERV':       {'text': 'OBSERVACIONES', 'width': 420, 'anchor': 'w',      'stretch': True},
            'TIPO':         {'text': 'TIPO',          'width': 80,  'anchor': 'center', 'stretch': False},
            'SALDO MENSUAL':{'text': 'SALDO ($)',     'width': 110, 'anchor': 'e',      'stretch': False},
        }

        for col, config in headers_config.items():
            self.tree.heading(col, text=config['text'])
            self.tree.column(col, width=config['width'], anchor=config['anchor'],
                             minwidth=50, stretch=config['stretch'])
        
        self.configurar_tags_tree()
        
        scrollbar_v = ttk.Scrollbar(frame_tree, orient="vertical", command=self.tree.yview)
        scrollbar_h = ttk.Scrollbar(frame_tree, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_v.set, xscrollcommand=scrollbar_h.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        scrollbar_v.grid(row=0, column=1, sticky='ns', pady=5)
        scrollbar_h.grid(row=1, column=0, sticky='ew', padx=5)
        
        frame_tree.grid_rowconfigure(0, weight=1)
        frame_tree.grid_columnconfigure(0, weight=1)
        
        self.tree.bind("<Double-1>", self.mostrar_detalle_fila)
        
        # --- PANEL DERECHO: Lista de empleados con saldo ---
        self.frame_panel_saldos = ttk.LabelFrame(paned, text=" 👥 Empleados con Saldo ", padding="5")
        paned.add(self.frame_panel_saldos, weight=2)
        
        # Buscador + filtro activos del panel
        frame_busq = ttk.Frame(self.frame_panel_saldos)
        frame_busq.pack(fill="x", pady=(0, 3))

        ttk.Label(frame_busq, text="Buscar:", font=("Segoe UI", 9)).pack(side="left", padx=(0, 3))
        self.var_solo_activos = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_busq, text="Act.", variable=self.var_solo_activos,
                        command=self.filtrar_panel_saldos).pack(side="right")
        self.entry_filtro_saldo = ttk.Entry(frame_busq, font=("Segoe UI", 9))
        self.entry_filtro_saldo.pack(side="left", fill="x", expand=True)
        self.entry_filtro_saldo.bind("<KeyRelease>", lambda e: self.root.after(300, self.filtrar_panel_saldos))
        
        # Treeview del panel
        columnas_saldo = ('codigo', 'nombre', 'saldo')
        self.panel_saldos_tree = ttk.Treeview(self.frame_panel_saldos, columns=columnas_saldo, show='headings', height=30)
        self.panel_saldos_tree.heading('codigo', text='CÓD.')
        self.panel_saldos_tree.heading('nombre', text='NOMBRE')
        self.panel_saldos_tree.heading('saldo', text='SALDO ($)')
        self.panel_saldos_tree.column('codigo', width=52,  anchor='center', stretch=False)
        self.panel_saldos_tree.column('nombre', width=210, anchor='w',      stretch=True)
        self.panel_saldos_tree.column('saldo',  width=110, anchor='e',      stretch=False)
        
        scroll_panel = ttk.Scrollbar(self.frame_panel_saldos, orient="vertical", command=self.panel_saldos_tree.yview)
        self.panel_saldos_tree.configure(yscrollcommand=scroll_panel.set)
        
        self.panel_saldos_tree.pack(side="left", fill="both", expand=True)
        scroll_panel.pack(side="right", fill="y")
        
        self.panel_saldos_tree.bind("<Double-1>", self.seleccionar_empleado_panel)
        self.panel_saldos_tree.bind("<Return>", self.seleccionar_empleado_panel)
        
        self.todos_saldos = []
        self.cargar_panel_saldos()
        
        # Atajos de teclado globales
        self.root.bind("<F5>", lambda e: self.abrir_buscador_nombres())
        self.root.bind("<Control-f>", lambda e: self.abrir_buscador_nombres())
        
        self.entry_empleado.focus()
    
    def configurar_estilos(self):
        """Configura estilos personalizados para la interfaz (paleta INTERFAZ_GRAFICA_RRHH.txt)"""
        COL_HEADER  = '#1E3A5F'
        COL_ACCENT  = '#2980B9'
        COL_OK      = '#27AE60'
        COL_PEND    = '#E67E22'

        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure("Primary.TButton", background=COL_ACCENT, foreground="white", font=("Segoe UI", 9, "bold"), padding=(8, 4))
        style.configure("Success.TButton", background=COL_OK, foreground="white", font=("Segoe UI", 9, "bold"), padding=(8, 4))
        style.configure("Warning.TButton", background=COL_PEND, foreground="white", font=("Segoe UI", 9, "bold"), padding=(6, 3))
        
        style.configure("Treeview", background="white", foreground="#2C3E50", rowheight=28, fieldbackground="white", font=("Segoe UI", 10))
        style.configure("Treeview.Heading", background=COL_HEADER, foreground="white", font=("Segoe UI", 11, "bold"), padding=(5, 8))
        style.map("Treeview", background=[('selected', COL_ACCENT)])
    
    def configurar_tags_tree(self):
        """Configura los tags del TreeView con colores mejorados"""
        self.tree.tag_configure('ingreso', background='#ffffff', foreground='#2c3e50', font=("Segoe UI", 11, "bold"))
        self.tree.tag_configure('egreso', background='#ffeaea', foreground='#c0392b', font=("Segoe UI", 11))
        self.tree.tag_configure('ingreso_historico', background='#e8f4fd', foreground='#2980b9', font=("Segoe UI", 11, "bold"))
        self.tree.tag_configure('egreso_historico', background='#fff4e6', foreground='#d68910', font=("Segoe UI", 11))
    
    def ejecutar(self):
        """Ejecuta la aplicación"""
        self.root.mainloop()

def main():
    print("🚀 Iniciando Consultor de Historial de Préstamos v10 con Búsqueda por Nombres + Saldos...")
    print("📊 Funcionalidades: SQL Server + SQLite + Filtros + Búsqueda por Nombres + Exportar Saldos")
    print("💡 Atajos: F5 = Buscar por nombres, Enter = Buscar préstamos")
    print("💾 NUEVO: Botón 'SALDOS' para exportar saldos de préstamos a Excel")
    print("📅 v10: Fecha del INGRESO tomada de RPINGDES (original) en lugar de RPHISTOR (descuento)")
    app = ConsultorPrestamos()
    app.ejecutar()

if __name__ == "__main__":
    main()
